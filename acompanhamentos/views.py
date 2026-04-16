from typing import Any
import uuid
from django.shortcuts import render
from django.views.generic import (
    ListView,
    CreateView,
    DetailView,
    UpdateView,
    DeleteView,
)
from django.urls import reverse_lazy
from django.shortcuts import redirect
from django.forms import inlineformset_factory
from django.db import transaction
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
import logging
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta, time
import json
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404
from decimal import Decimal, InvalidOperation
import requests
from django.conf import settings
from openpyxl.styles import Alignment
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView
from .models import (
    registrodeagenteacompanhamento,
    registroacompanhamento,
    registroacompanhamentoagente,
    registroderesposavelagenteacompanhamento,
    FranquiaAgente,
    Cliente
)

from .forms import (
    RegistroAgente,
    RegistroAcompanhamentoForm,
    RegistroAcompanhamentoAgenteForm,
    RegistroAcompanhamentoAgenteCreateFormSet,
    RegistroAcompanhamentoAgenteUpdateFormSet,
    RegistroResponsavelAgente,
    RegistroAcompanhamentoFromRequisicaoForm,
    BaseAgenteObrigatorioFormSet,
    RegistroAcompanhamentoEditForm,
    RegistroAcompanhamentoAgenteEditFormSet,
)

from typing import Any
from django.views.generic import (
    ListView,
    CreateView,
    DetailView,
    UpdateView,
    DeleteView,
)

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from .utils import gerar_link_app_missao
from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import get_object_or_404
from .api.supabase_client import get_supabase
from .gs_acionamento_service import gs_acionamento
from django.db.models import Case, When, Value, IntegerField
import logging
logger = logging.getLogger(__name__)

import logging
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from .api.supabase_client import get_supabase
from .models import registroacompanhamento
from .gs_sync_utils import notificar_gs_mudanca_status

logger = logging.getLogger(__name__)
from datetime import datetime, timedelta, date, time as dtime
from django.utils import timezone as tz
from .utils import sync_acompanhamento_to_supabase, delete_supabase_mission
from .gs_sync_utils import notificar_gs_verificado

@csrf_exempt
@require_GET
def sync_status_from_supabase(request):
    """
    Sincroniza status_acompanhamento no Django com o status real do Supabase.
    Consulta todas as missões ativas (não concluídas) e atualiza se divergente.

    GET /acompanhamentos/api/sync-status-supabase/
    """
    sb = get_supabase()

    # Busca acompanhamentos Django ativos com mission_id
    acompanhamentos = list(
        registroacompanhamento.objects
        .exclude(supabase_mission_id__isnull=True)
        .exclude(status_acompanhamento__in=["concluido", "verificado"])
        .values_list("id", "supabase_mission_id", "status_acompanhamento")
    )

    if not acompanhamentos:
        return JsonResponse({"success": True, "message": "Nenhuma missão ativa", "updated": 0})

    mission_ids = [str(a[1]) for a in acompanhamentos]
    django_map = {str(a[1]): {"id": a[0], "status": a[2]} for a in acompanhamentos}

    updated = 0
    errors = 0

    # Consulta Supabase em batches
    for i in range(0, len(mission_ids), 50):
        batch = mission_ids[i:i + 50]

        try:
            res = (
                sb.table("missions_control")
                .select("id, status")
                .in_("id", batch)
                .execute()
            )
            supabase_missions = res.data or []
        except Exception as e:
            logger.error(f"Erro Supabase batch {i}: {e}")
            errors += 1
            continue

        for mission in supabase_missions:
            sb_id = mission.get("id")
            sb_status = (mission.get("status") or "").strip()

            if not sb_id or not sb_status:
                continue

            django_info = django_map.get(sb_id)
            if not django_info:
                continue

            if django_info["status"] != sb_status:
                try:
                    acomp = registroacompanhamento.objects.get(pk=django_info["id"])
                    acomp.status_acompanhamento = sb_status
                    acomp.save(update_fields=["status_acompanhamento"])
                    updated += 1

                    # ✅ Notificar GS se status relevante (em_andamento, concluido)
                    notificar_gs_mudanca_status(django_info["id"], sb_status)

                except Exception as e:
                    logger.error(f"Erro update/notify acomp #{django_info['id']}: {e}")
                    errors += 1

    return JsonResponse({
        "success": True,
        "checked": len(mission_ids),
        "updated": updated,
        "errors": errors,
    })


# views.py (helpers)

def _build_clone_acompanhamento(base, *, tipo_servico, nome_user):
    """
    Cria um NOVO registroacompanhamento copiando os campos do form base.
    Cada clone vira 1 missão no Supabase.
    """
    return registroacompanhamento(
        cliente=base.cliente,
        tipo_servico=tipo_servico or base.tipo_servico,

        origem=base.origem,
        origem_2=base.origem_2,
        origem_3=base.origem_3,
        destino=base.destino,
        destino_2=base.destino_2,
        destino_3=base.destino_3,
        latitude_origem=base.latitude_origem,
        longitude_origem=base.longitude_origem,
        latitude_origem2=base.latitude_origem2,
        longitude_origem2=base.longitude_origem2,
        latitude_origem3=base.latitude_origem3,
        longitude_origem3=base.longitude_origem3,
        latitude_destino=base.latitude_destino,
        longitude_destino=base.longitude_destino,
        latitude_destino_2=base.latitude_destino_2,
        longitude_destino_2=base.longitude_destino_2,
        latitude_destino_3=base.latitude_destino_3,
        longitude_destino_3=base.longitude_destino_3,
        km_estimado_carro=base.km_estimado_carro,
        km_estimado_moto=base.km_estimado_moto,
        raio_cerca=base.raio_cerca if base.raio_cerca is not None else 60,
        raio_cerca_2=base.raio_cerca_2 if base.raio_cerca_2 is not None else 60,
        raio_cerca_3=base.raio_cerca_3 if base.raio_cerca_3 is not None else 60,

        campo_personalizado_titulo=base.campo_personalizado_titulo,
        campo_personalizado_valor=base.campo_personalizado_valor,
        ocorrencia=base.ocorrencia,

        status=base.status or "pendente",
        status_acompanhamento=getattr(base, "status_acompanhamento", None) or "pendente",
        botao_panico=getattr(base, "botao_panico", False),

        # ✅ Copia is_reuso do base
        is_reuso=getattr(base, "is_reuso", False),

        nome_user=nome_user,
    )


def _extract_principais_from_formset(agentes_formset):
    """
    Extrai instâncias NÃO salvas do formset (somente principais).
    Ignora forms vazios e marcados pra DELETE.
    """
    agentes = []
    for f in agentes_formset.forms:
        if not getattr(f, "cleaned_data", None):
            continue
        if f.cleaned_data.get("DELETE"):
            continue
        if not f.has_changed():
            continue

        inst = f.save(commit=False)  # não precisa de parent aqui
        if getattr(inst, "tipo_agente", None) == "carona":
            continue

        agentes.append(inst)
    return agentes


def _criar_caronas(acompanhamento, agente_principal, carona_ids):
    """
    Reaproveita sua lógica atual de carona (copiando dados do principal).
    """
    if not agente_principal:
        return

    for agente_id in carona_ids:
        if not agente_id:
            continue

        agente_obj = registrodeagenteacompanhamento.objects.filter(id=agente_id).first()
        if not agente_obj:
            continue

        registroacompanhamentoagente.objects.create(
            acompanhamento=acompanhamento,
            tipo_agente="carona",
            responsavel_agente=agente_principal.responsavel_agente,
            agente=agente_obj,
            franquia=agente_principal.franquia,
            tipo_servico=agente_principal.tipo_servico,

            placa_agente=agente_principal.placa_agente,
            motorista=agente_principal.motorista,
            placa_motorista=agente_principal.placa_motorista,

            km_inicio=agente_principal.km_inicio,
            km_final=agente_principal.km_final,
            km_total=agente_principal.km_total,

            data_solicitada=agente_principal.data_solicitada,
            horario_solicitado=agente_principal.horario_solicitado,
            data_inicio=agente_principal.data_inicio,
            horario_inicio=agente_principal.horario_inicio,
            data_finalizacao=agente_principal.data_finalizacao,
            horario_finalizacao=agente_principal.horario_finalizacao,

            pedagio=Decimal("0.00"),
            valor_agente=Decimal("0.00"),
        )


def _sync_supabase_safe(request, acompanhamento):
    """
    Sincroniza no Supabase SEM explodir sua criação se der erro de rede/RLS/etc.
    """
    try:
        return sync_acompanhamento_to_supabase(acompanhamento)  # (success, mission_id, error)
    except Exception as e:
        logger.exception("Erro inesperado ao sincronizar com Supabase")
        return (False, None, str(e))


# ------------------------------------------------------
#                 Agente Acompanhamento
# ------------------------------------------------------
class AgenteAcompanhamentoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = registrodeagenteacompanhamento
    form_class = RegistroAgente
    template_name = "agente_acompanhamento_create.html"
    success_url = reverse_lazy("agenteAcompanhamentoList")
    permission_required = "acompanhamentos.add_registrodeagenteacompanhamento"

    def get_initial(self):
        initial = super().get_initial()
        user = getattr(self.request, 'user', None)
        if user:
            initial['nome_user'] = user.get_full_name() or user.username
        return initial

    def form_valid(self, form):
        campos_para_validar = [
            'nome', 'cpf', 'pix', 'banco', 'agencia', 'conta', 'tipo_conta'
        ]

        dados = form.cleaned_data

        # Verifica se TODOS estão vazios
        todos_vazios = all(
            not dados.get(campo)
            for campo in campos_para_validar
        )

        if todos_vazios:
            return redirect("agenteAcompanhamentoCreate")

        instance = form.save(commit=False)

        user = getattr(self.request, 'user', None)
        if user:
            instance.nome_user = user.get_full_name() or user.username

        instance.save()
        return super().form_valid(form)

class AgenteAcompanhamentoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = registrodeagenteacompanhamento
    template_name = "agente_acompanhamento_list.html"
    context_object_name = "agente_acompanhamentos"
    permission_required = "acompanhamentos.view_registrodeagenteacompanhamento"

    def get_queryset(self):
        queryset = super().get_queryset()

        nome = self.request.GET.get("nome")

        if nome:
            queryset = queryset.filter(nome__icontains=nome)

        return queryset.order_by("-criado_em")

class RegistroAgenteAcompanhamentoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):

    model = registrodeagenteacompanhamento
    form_class = RegistroAgente
    template_name = "agente_acompanhamento_create.html"
    success_url = reverse_lazy("agenteAcompanhamentoList")
    permission_required = "acompanhamentos.change_registrodeagenteacompanhamento"

    def form_valid(self, form):
        return super().form_valid(form)

# ------------------------------------------------------
#             Responsável Agente Acompanhamento
# ------------------------------------------------------
class ResponsavelAgenteAcompanhamentoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = registroderesposavelagenteacompanhamento
    template_name = "responsavel_list.html"
    context_object_name = "responsavel_agente_acompanhamentos"
    permission_required = "acompanhamentos.view_registrodeagenteacompanhamento"

    def get_queryset(self):
        queryset = super().get_queryset()

        nome = self.request.GET.get("nome")

        if nome:
            queryset = queryset.filter(nome__icontains=nome)

        return queryset.order_by("-criado_em")

@login_required
@permission_required("acompanhamentos.add_registroderesposavelagenteacompanhamento", raise_exception=True)

@require_POST
def criar_responsavel_agente_ajax(request):
    nome = request.POST.get("nome")

    if not nome:
        return JsonResponse(
            {"success": False, "error": "Nome é obrigatório."},
            status=400
        )

    agente = registroderesposavelagenteacompanhamento.objects.create(
        nome=nome,
        nome_user=request.user.get_full_name() or request.user.username
    )

    return JsonResponse({
        "success": True,
        "id": agente.id,
        "nome": agente.nome
    })

@login_required
@permission_required("acompanhamentos.change_registroderesposavelagenteacompanhamento",raise_exception=True)

@require_POST
def editar_responsavel_agente_ajax(request, pk):
    nome = request.POST.get("nome")

    if not nome:
        return JsonResponse(
            {"success": False, "error": "Nome é obrigatório."},
            status=400
        )

    agente = get_object_or_404(
        registroderesposavelagenteacompanhamento,
        pk=pk
    )

    agente.nome = nome
    agente.nome_user = request.user.get_full_name() or request.user.username
    agente.save(update_fields=["nome", "nome_user", "atualizado_em"])

    return JsonResponse({
        "success": True,
        "id": agente.id,
        "nome": agente.nome
    })

# ------------------------------------------------------
#                   Acompanhamento
# ------------------------------------------------------
def join_values(values):
    return "\n".join(str(v) for v in values if v)


def _is_pronta_resposta_tipo(tipo_servico) -> bool:
    """Identifica se o tipo de servico eh pronta_resposta."""
    try:
        return (getattr(tipo_servico, "tipo_cadastro", "") or "").strip().lower() == "pronta_resposta"
    except Exception:
        return False


def _parse_km_post(value):
    """Converte valor recebido do form para Decimal(2), aceitando virgula ou ponto."""
    raw = (value or "").strip()
    if not raw:
        return None

    normalized = raw.replace(",", ".")
    try:
        return Decimal(normalized).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _resolve_km_estimado_from_post(request, tipo_servico):
    """Le KM estimado enviado pelo front apenas para pronta_resposta."""
    if not _is_pronta_resposta_tipo(tipo_servico):
        return None, None

    km_carro = _parse_km_post(request.POST.get("km_estimado_carro"))
    km_moto = _parse_km_post(request.POST.get("km_estimado_moto"))
    return km_carro, km_moto


def _calcular_rota_google(origem_lat, origem_lng, destino_lat, destino_lng, travel_mode):
    google_maps_api_key = getattr(settings, "GOOGLE_MAPS_API_KEY", "")
    if not google_maps_api_key:
        raise RuntimeError("Google API key nao configurada")

    response = requests.post(
        "https://routes.googleapis.com/directions/v2:computeRoutes",
        headers={
            "Content-Type": "application/json",
            "X-Goog-Api-Key": google_maps_api_key,
            "X-Goog-FieldMask": "routes.distanceMeters,routes.duration",
        },
        json={
            "origin": {
                "location": {
                    "latLng": {
                        "latitude": origem_lat,
                        "longitude": origem_lng,
                    }
                }
            },
            "destination": {
                "location": {
                    "latLng": {
                        "latitude": destino_lat,
                        "longitude": destino_lng,
                    }
                }
            },
            "travelMode": travel_mode,
            "routingPreference": "TRAFFIC_UNAWARE",
        },
        timeout=15,
    )
    response.raise_for_status()

    payload = response.json()
    routes = payload.get("routes") or []
    if not routes:
        return None

    first_route = routes[0]
    distance_meters = first_route.get("distanceMeters")
    if distance_meters is None:
        return None

    return {
        "distance_meters": distance_meters,
        "km": round(float(distance_meters) / 1000, 2),
        "duration": first_route.get("duration"),
    }


def _calcular_estimativa_km_google(origem_lat, origem_lng, destino_lat, destino_lng):
    try:
        carro = _calcular_rota_google(origem_lat, origem_lng, destino_lat, destino_lng, "DRIVE")
    except RuntimeError as exc:
        return {"success": False, "error": str(exc)}
    except requests.RequestException:
        carro = None

    try:
        moto = _calcular_rota_google(origem_lat, origem_lng, destino_lat, destino_lng, "TWO_WHEELER")
    except RuntimeError as exc:
        return {"success": False, "error": str(exc)}
    except requests.RequestException:
        moto = None

    if carro is None and moto is None:
        return {
            "success": False,
            "error": "Falha ao consultar Google Routes API. Verifique chave, faturamento e permissoes da API.",
        }

    return {
        "success": True,
        "carro": carro,
        "moto": moto,
    }


@require_GET
@login_required
def ajax_estimativa_km(request):
    """
    Calcula estimativa de distancia (KM) para carro e moto usando
    a Google Routes API com base nas coordenadas informadas.
    """
    origem_lat = request.GET.get("origem_lat")
    origem_lng = request.GET.get("origem_lng")
    destino_lat = request.GET.get("destino_lat")
    destino_lng = request.GET.get("destino_lng")

    if not all([origem_lat, origem_lng, destino_lat, destino_lng]):
        return JsonResponse({"success": False, "error": "Coordenadas incompletas"}, status=400)

    try:
        origem_lat = float(origem_lat)
        origem_lng = float(origem_lng)
        destino_lat = float(destino_lat)
        destino_lng = float(destino_lng)
    except (TypeError, ValueError):
        return JsonResponse({"success": False, "error": "Coordenadas invalidas"}, status=400)

    resultado = _calcular_estimativa_km_google(
        origem_lat=origem_lat,
        origem_lng=origem_lng,
        destino_lat=destino_lat,
        destino_lng=destino_lng,
    )

    if not resultado.get("success"):
        status_code = 503 if "API key" in str(resultado.get("error", "")) else 502
        return JsonResponse(resultado, status=status_code)

    return JsonResponse(resultado)

def moeda(valor):
    if valor is None:
        return ""
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def agentes_nao_carona(agentes):
    return [a for a in agentes if a.tipo_agente != "carona"]

def join_values_nao_carona(values):
    return "\n".join(str(v) for v in values if v)



# views.py

class AcompanhamentoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = registroacompanhamento
    form_class = RegistroAcompanhamentoForm
    template_name = "acompanhamento_create.html"
    success_url = reverse_lazy("requisicao_solicitacao_list")
    permission_required = "acompanhamentos.add_registroacompanhamento"

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        agentes_formset = context["agentes_formset"]

        if not agentes_formset.is_valid():
            return self.render_to_response(context)

        user = self.request.user
        nome_user = user.get_full_name() or user.username

        # Base (NÃO salva) - vai ser clonada N vezes
        base = form.save(commit=False)
        if not getattr(base, "status", None):
            base.status = "pendente"
        base.nome_user = nome_user

        # 1) pega os agentes principais do formset
        agentes_principais = _extract_principais_from_formset(agentes_formset)
        if not agentes_principais:
            messages.warning(self.request, "Nenhum serviço/agente foi preenchido no formset.")
            return self.render_to_response(context)

        # 2) cria 1 acompanhamento por agente principal
        criados = []  # [(acompanhamento, agente_principal)]
        for agente_inst in agentes_principais:
            tipo_serv = agente_inst.tipo_servico or base.tipo_servico

            novo_acomp = _build_clone_acompanhamento(
                base,
                tipo_servico=tipo_serv,
                nome_user=nome_user
            )
            novo_acomp.save()

            agente_inst.acompanhamento = novo_acomp
            agente_inst.tipo_servico = agente_inst.tipo_servico or tipo_serv
            agente_inst.save()

            criados.append((novo_acomp, agente_inst))

        # 3) caronas: se tiver mais de 1 serviço, aplica só no 1º (pra não duplicar tudo)
        caronas = self.request.POST.getlist("carona_agente[]")
        if caronas:
            alvo_acomp = criados[0][0]
            principal = alvo_acomp.agentes.filter(tipo_agente="principal").first()
            _criar_caronas(alvo_acomp, principal, caronas)

            if len(criados) > 1:
                messages.warning(
                    self.request,
                    "Caronas aplicadas apenas no 1º serviço (porque você criou múltiplos acompanhamentos)."
                )

        # 4) Sync Supabase (1 missão por acompanhamento)
        ok = 0
        mission_ids_ok = []
        for (acomp, _) in criados:
            success, mission_id, error = _sync_supabase_safe(self.request, acomp)
            if success:
                ok += 1
                if mission_id:
                    mission_ids_ok.append(str(mission_id))
            else:
                messages.warning(
                    self.request,
                    f"Acompanhamento #{acomp.id} criado, mas falhou ao sincronizar no Supabase: {error}"
                )

        messages.success(
            self.request,
            f"Foram criados {len(criados)} acompanhamentos (1 por serviço). "
            f"Sincronizados no Supabase: {ok}/{len(criados)}."
        )

        # Se quiser logar os mission_ids pra copiar fácil:
        if mission_ids_ok:
            logger.info("Supabase mission_ids criados: %s", ", ".join(mission_ids_ok))

        return redirect(self.success_url)

class AcompanhamentoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = registroacompanhamento
    template_name = "acompanhamento_list.html"
    context_object_name = "itens"
    permission_required = "acompanhamentos.view_registroacompanhamento"

    TIPOS_CADASTRO = [
        ("", "Todos"),
        ("acompanhamento", "Acompanhamento"),
        ("pronta_resposta", "Pronta Resposta"),
    ]

    def _is_cancelado(self, queryset):
        return queryset.filter(
            Q(status_requisicao__iexact="cancelado") |
            Q(status_requisicao__iexact="cancelada") |
            Q(status_requisicao__iexact="canceladas")
        )

    def _apply_common_filters(self, qs):
        data = self.request.GET.get("data")
        data2 = self.request.GET.get("data2")
        agente = self.request.GET.get("agente")
        cliente = self.request.GET.get("cliente")
        responsavel = self.request.GET.get("responsavel")
        tipo_cadastro = self.request.GET.get("tipo_cadastro")

        if data and data2:
            if data > data2:
                data, data2 = data2, data
            qs = qs.filter(agentes__data_solicitada__range=[data, data2])
        elif data:
            qs = qs.filter(agentes__data_solicitada=data)
        elif data2:
            qs = qs.filter(agentes__data_solicitada=data2)

        if agente:
            qs = qs.filter(agentes__agente__nome__icontains=agente)

        if cliente:
            qs = qs.filter(cliente__nome__icontains=cliente)

        if responsavel:
            qs = qs.filter(agentes__responsavel_agente_id=responsavel)

        if tipo_cadastro in {"acompanhamento", "pronta_resposta"}:
            qs = qs.filter(tipo_servico__tipo_cadastro=tipo_cadastro)

        return qs

    def get_pendentes_queryset(self):
        qs = registroacompanhamento.objects.filter(
            Q(tipo_servico__isnull=True) |

            Q(agentes__responsavel_agente__isnull=True) |
            Q(agentes__responsavel_agente__isnull=True) |

            Q(agentes__agente__isnull=True) |

            Q(agentes__data_solicitada__isnull=True) |
            Q(agentes__horario_solicitado__isnull=True) |

            Q(agentes__data_inicio__isnull=True) |
            Q(agentes__horario_inicio__isnull=True) |

            Q(agentes__data_finalizacao__isnull=True) |
            Q(agentes__horario_finalizacao__isnull=True) |

            Q(agentes__km_inicio__isnull=True) |
            Q(agentes__km_final__isnull=True) |

            Q(agentes__franquia__isnull=True)
        ).distinct()

        return qs.exclude(
            Q(status_requisicao__iexact="cancelado") |
            Q(status_requisicao__iexact="cancelada") |
            Q(status_requisicao__iexact="canceladas")
        )

    def get_queryset(self):
        qs = (
            registroacompanhamento.objects
            .prefetch_related(
                "agentes",
                "agentes__agente",
                "agentes__franquia"
            )
            .select_related("cliente")
        )

        qs = self._apply_common_filters(qs)

        canceladas = self.request.GET.get("canceladas")
        if canceladas == "true":
            return self._is_cancelado(qs).distinct().order_by("-criado_em")

        qs = qs.exclude(
            Q(status_requisicao__iexact="cancelado") |
            Q(status_requisicao__iexact="cancelada") |
            Q(status_requisicao__iexact="canceladas")
        )

        pendente = self.request.GET.get("pendente")
        if pendente == "true":
            return self._apply_common_filters(self.get_pendentes_queryset()).distinct().order_by("-criado_em")

        return qs.distinct().order_by("-criado_em")

    def render_to_response(self, context, **response_kwargs):

        queryset = context["itens"]

        if self.request.GET.get("export") == "excel":
            return self.exportar_excel(queryset)
        
        if self.request.GET.get("export") == "pdf":
            return self.exportar_pdf(queryset)

        return super().render_to_response(context, **response_kwargs)

    def exportar_pdf(self, queryset):
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=(2000, 900),
            topMargin=20,
            leftMargin=20,
            rightMargin=20,
            bottomMargin=20
        )

        elements = []

        styles = getSampleStyleSheet()
        title_style = styles["Title"]

        cell_style = ParagraphStyle(
            name="CellStyle",
            fontName="Helvetica",
            fontSize=7,
            leading=9,
            wordWrap="CJK",
            alignment=0
        )

        title = Paragraph("Relatório de Acompanhamentos", title_style)
        title.alignment = 1
        elements.append(title)
        elements.append(Paragraph("<br/>", styles["Normal"]))

        headers = [
            "Protocolo", "Cliente", "Tipo Serviço", "Origem", "Destino",
            "Responsável", "Agente", "Placa Agente",
            "Motorista", "Placa Motorista",
            "Data Solicitada", "Hora Solicitada",
            "Data Inicial", "Hora Inicial",
            "Data Final", "Hora Final",
            "Hora Total", "Hora Excedente",
            "KM Início", "KM Final", "KM Total", "KM Excedente",
            "Pedágio", "Franquia",
            "Valor Agente(s)",
            "TOTAL",
            "Ocorrência", "Feito Por"
        ]

        data = [headers]

        total_geral = Decimal("0.00")

        for item in queryset:
            agentes = item.agentes.all()
            agentes_validos = agentes_nao_carona(agentes)

            total_geral += item.total_valor_agentes

            row = [
                str(item.id),
                Paragraph(str(item.cliente), cell_style),
                Paragraph(str(item.tipo_servico), cell_style),
                Paragraph(item.origem or "", cell_style),
                Paragraph(item.destino or "", cell_style),

                Paragraph(
                    join_values(a.responsavel_agente.nome for a in agentes_validos),
                    cell_style
                ),
                Paragraph(join_values(str(a.agente) for a in agentes), cell_style),
                Paragraph(join_values(a.placa_agente for a in agentes_validos), cell_style),

                Paragraph(join_values(a.motorista for a in agentes_validos), cell_style),
                Paragraph(join_values(a.placa_motorista for a in agentes_validos), cell_style),

                join_values_nao_carona(a.data_solicitada.strftime("%d/%m/%Y") for a in agentes_validos if a.data_solicitada),
                join_values_nao_carona(a.horario_solicitado.strftime("%H:%M") for a in agentes_validos if a.horario_solicitado),

                join_values_nao_carona(a.data_inicio.strftime("%d/%m/%Y") for a in agentes_validos if a.data_inicio),
                join_values_nao_carona(a.horario_inicio.strftime("%H:%M") for a in agentes_validos if a.horario_inicio),

                join_values_nao_carona(a.data_finalizacao.strftime("%d/%m/%Y") for a in agentes_validos if a.data_finalizacao),
                join_values_nao_carona(a.horario_finalizacao.strftime("%H:%M") for a in agentes_validos if a.horario_finalizacao),

                join_values_nao_carona(str(a.horario_total) for a in agentes_validos if a.horario_total),
                join_values_nao_carona(str(a.horario_excedente) for a in agentes_validos if a.horario_excedente),

                join_values_nao_carona(str(a.km_inicio) for a in agentes_validos if a.km_inicio is not None),
                join_values_nao_carona(str(a.km_final) for a in agentes_validos if a.km_final is not None),
                join_values_nao_carona(str(a.km_total) for a in agentes_validos if a.km_total is not None),
                join_values_nao_carona(str(a.km_excedente) for a in agentes_validos if a.km_excedente),

                join_values_nao_carona(moeda(a.pedagio) for a in agentes_validos if a.pedagio),
                join_values(a.franquia.nome if a.franquia else "" for a in agentes_validos),

                join_values(moeda(a.valor_agente) for a in agentes if a.valor_agente),
                moeda(item.total_valor_agentes),

                Paragraph(item.ocorrencia or "", cell_style),
                Paragraph(item.nome_user or "", cell_style),
            ]

            data.append(row)

        total_row = [""] * len(headers)
        total_row[25] = Paragraph(f"Total Geral: {moeda(total_geral)}", cell_style)
        data.append(total_row)

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

            ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="acompanhamentos.pdf"'
        return response

    def exportar_excel(self, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Acompanhamentos"

        headers = [
            "Protocolo", "Cliente", "Tipo Serviço", "Origem", "Destino",
            "Responsável", "Agente", "Placa Agente",
            "Motorista", "Placa Motorista",
            "Data Solicitada", "Hora Solicitada",
            "Data Inicial", "Hora Inicial",
            "Data Final", "Hora Final",
            "Hora Total", "Hora Excedente",
            "KM Início", "KM Final", "KM Total", "KM Excedente",
            "Pedágio", "Franquia",
            "Valor Agente(s)",
            "TOTAL",
            "Ocorrência", "Feito Por"
        ]

        sheet.append(headers)

        total_geral = Decimal("0.00")

        for item in queryset:
            agentes = item.agentes.all()
            agentes_validos = agentes_nao_carona(agentes)
            total_geral += item.total_valor_agentes or Decimal("0.00")

            sheet.append([
                item.id,
                str(item.cliente),
                str(item.tipo_servico),
                item.origem,
                item.destino,

                join_values(a.responsavel_agente.nome for a in agentes_validos),
                join_values(str(a.agente) for a in agentes),
                join_values(a.placa_agente for a in agentes_validos),

                join_values(a.motorista for a in agentes_validos),
                join_values(a.placa_motorista for a in agentes_validos),

                join_values(a.data_solicitada.strftime("%d/%m/%Y") for a in agentes_validos if a.data_solicitada),
                join_values(a.horario_solicitado.strftime("%H:%M") for a in agentes_validos if a.horario_solicitado),

                join_values(a.data_inicio.strftime("%d/%m/%Y") for a in agentes_validos if a.data_inicio),
                join_values(a.horario_inicio.strftime("%H:%M") for a in agentes_validos if a.horario_inicio),

                join_values(a.data_finalizacao.strftime("%d/%m/%Y") for a in agentes_validos if a.data_finalizacao),
                join_values(a.horario_finalizacao.strftime("%H:%M") for a in agentes_validos if a.horario_finalizacao),

                join_values(str(a.horario_total) for a in agentes_validos if a.horario_total),
                join_values(str(a.horario_excedente) for a in agentes_validos if a.horario_excedente),

                join_values(str(a.km_inicio) for a in agentes_validos if a.km_inicio is not None),
                join_values(str(a.km_final) for a in agentes_validos if a.km_final is not None),
                join_values(str(a.km_total) for a in agentes_validos if a.km_total is not None),
                join_values(str(a.km_excedente) for a in agentes_validos if a.km_excedente),

                join_values(moeda(a.pedagio) for a in agentes_validos if a.pedagio),
                join_values(a.franquia.nome if a.franquia else "" for a in agentes_validos),

                join_values(moeda(a.valor_agente) for a in agentes if a.valor_agente),
                moeda(item.total_valor_agentes),

                item.ocorrencia,
                item.nome_user,
            ])

            row_number = sheet.max_row

            multi_columns = [
                6, 7, 8, 9, 10,
                11, 12, 13, 14, 15, 16,
                17, 18,
                19, 20, 21, 22,
                23, 24, 25
            ]

            max_lines = 1
            for col in multi_columns:
                cell = sheet.cell(row=row_number, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.value:
                    max_lines = max(max_lines, cell.value.count("\n") + 1)

            sheet.row_dimensions[row_number].height = 15 * max_lines

        sheet.append([])

        total_row = [""] * len(headers)
        total_row[25] = f"TOTAL GERAL: {moeda(total_geral)}"
        sheet.append(total_row)

        last_row = sheet.max_row
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D3D3D3",
                end_color="D3D3D3",
                fill_type="solid"
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="acompanhamentos.xlsx"'
        workbook.save(response)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["franquias"] = FranquiaAgente.objects.all().order_by("nome")

        context["pendentes_count"] = self._apply_common_filters(self.get_pendentes_queryset()).distinct().count()

        cancelados_base = self._apply_common_filters(
            registroacompanhamento.objects.all()
        )
        context["cancelados_count"] = self._is_cancelado(cancelados_base).distinct().count()
        context["mostrando_canceladas"] = self.request.GET.get("canceladas") == "true"

        context["data"] = self.request.GET.get("data", "")
        context["data2"] = self.request.GET.get("data2", "")
        context["agente"] = self.request.GET.get("agente", "")
        context["cliente"] = self.request.GET.get("cliente", "")
        context["tipo_cadastro"] = self.request.GET.get("tipo_cadastro", "")
        context["tipos_cadastro"] = self.TIPOS_CADASTRO

        return context

class RegistroAcompanhamentoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = registroacompanhamento
    form_class = RegistroAcompanhamentoEditForm
    template_name = "acompanhamento2_create.html"
    success_url = reverse_lazy("acompanhamentosList")
    permission_required = "acompanhamentos.change_registroacompanhamento"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        if self.request.method == "POST":
            ctx["agentes_formset"] = RegistroAcompanhamentoAgenteEditFormSet(
                self.request.POST,
                instance=self.object,
            )
        else:
            ctx["agentes_formset"] = RegistroAcompanhamentoAgenteEditFormSet(
                instance=self.object,
            )

        ctx["is_edit"] = True
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        agentes_formset = ctx["agentes_formset"]

        if agentes_formset.is_valid():
            self.object = form.save()
            agentes_formset.save()

            self.object.refresh_from_db()

            if self.object.supabase_mission_id:
                try:
                    success, mission_id, error = sync_acompanhamento_to_supabase(
                        self.object
                    )
                    if success:
                        messages.success(
                            self.request,
                            "Acompanhamento atualizado com sucesso! "
                            "Supabase sincronizado."
                        )
                    else:
                        messages.warning(
                            self.request,
                            f"Acompanhamento atualizado, mas falhou ao "
                            f"sincronizar no Supabase: {error}"
                        )
                except Exception as e:
                    logger.exception("Erro ao sincronizar edição no Supabase")
                    messages.warning(
                        self.request,
                        f"Acompanhamento atualizado, mas erro no Supabase: {e}"
                    )
            else:
                messages.success(
                    self.request,
                    "Acompanhamento atualizado com sucesso!"
                )

            return redirect(self.success_url)
        else:
            messages.error(self.request, "Corrija os erros nos agentes.")
            return self.render_to_response(ctx)

class AcompanhamentoFaturamentoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = registroacompanhamento
    template_name = "acompanhamento_faturamento_list.html"
    context_object_name = "itens"
    permission_required = "acompanhamentos.view_listacompanhamento"

    TIPOS_CADASTRO = [
        ("", "Todos"),
        ("acompanhamento", "Acompanhamento"),
        ("pronta_resposta", "Pronta Resposta"),
    ]

    def get_pendentes_queryset(self):
        return (
            registroacompanhamento.objects
            .filter(
                status_acompanhamento="concluido",
                status__in=["pendente", "aguardando"]
            )
            .exclude(is_reuso=True)
            .distinct()
        )

    def get_queryset(self):
        qs = (
            registroacompanhamento.objects
            .prefetch_related(
                "agentes",
                "agentes__agente",
                "agentes__franquia"
            )
            .select_related("cliente")
            .filter(
                status_acompanhamento="verificado",
                is_reuso=False,  # <- garante que não vem reuso
            )
        )

        data = self.request.GET.get("data")
        data2 = self.request.GET.get("data2")
        agente = self.request.GET.get("agente")
        cliente = self.request.GET.get("cliente")
        status = self.request.GET.get("status")
        tipo_cadastro = self.request.GET.get("tipo_cadastro")

        if data and data2:
            if data > data2:
                data, data2 = data2, data
            qs = qs.filter(agentes__data_solicitada__range=[data, data2])
        elif data:
            qs = qs.filter(agentes__data_solicitada=data)
        elif data2:
            qs = qs.filter(agentes__data_solicitada=data2)

        if agente:
            qs = qs.filter(agentes__agente__nome__icontains=agente)

        if cliente:
            qs = qs.filter(cliente__nome__icontains=cliente)

        if status:
            qs = qs.filter(status=status)

        if tipo_cadastro in {"acompanhamento", "pronta_resposta"}:
            qs = qs.filter(tipo_servico__tipo_cadastro=tipo_cadastro)

        return qs.distinct().order_by("-criado_em")

    def render_to_response(self, context, **response_kwargs):
        queryset = context["itens"]
        export = self.request.GET.get("export")

        if export == "faturamento_excel":
            return self.exportar_excel(queryset)
        
        if export == "faturamento_excel_cliente":
            return self.exportar_excel_cliente(queryset)

        if export == "faturamento_pdf":
            return self.exportar_pdf(queryset)

        return super().render_to_response(context, **response_kwargs)

    def exportar_pdf(self, queryset):
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=(2000, 900),
            topMargin=20,
            leftMargin=20,
            rightMargin=20,
            bottomMargin=20
        )

        elements = []

        styles = getSampleStyleSheet()
        title_style = styles["Title"]

        cell_style = ParagraphStyle(
            name="CellStyle",
            fontName="Helvetica",
            fontSize=7,
            leading=9,
            wordWrap="CJK",
            alignment=0
        )

        title = Paragraph("Relatório de Acompanhamentos", title_style)
        title.alignment = 1
        elements.append(title)
        elements.append(Paragraph("<br/>", styles["Normal"]))

        headers = [
            "Protocolo", "Cliente", "Tipo Serviço", "Origem", "Destino",
            "Responsável", "Agente", "Placa Agente",
            "Motorista", "Placa Motorista",
            "Data Solicitada", "Hora Solicitada",
            "Data Inicial", "Hora Inicial",
            "Data Final", "Hora Final",
            "Hora Total", "Hora Excedente",
            "KM Início", "KM Final", "KM Total", "KM Excedente",
            "Pedágio", "Franquia",
            "Valor Agente(s)",
            "Total Agentes",
            "Ocorrência",
            "Usuário",
            "Valor Contrato Cliente",
            "Lucro",
            "Pagamento",
            "Status",
            "NF",
        ]


        data = [headers]

        total_geral = Decimal("0.00")
        total_contrato = Decimal("0.00")
        total_lucro = Decimal("0.00")

        for item in queryset:
            agentes = item.agentes.all()
            agentes_validos = agentes_nao_carona(agentes)

            total_geral += item.total_valor_agentes
            total_contrato += item.valor_contrato or Decimal("0.00")
            total_lucro += item.lucro_total or Decimal("0.00")

            row = [
                str(item.id),
                Paragraph(str(item.cliente), cell_style),
                Paragraph(str(item.tipo_servico), cell_style),
                Paragraph(item.origem or "", cell_style),
                Paragraph(item.destino or "", cell_style),

                Paragraph(join_values(a.responsavel_agente.nome for a in agentes_validos), cell_style),
                Paragraph(join_values(str(a.agente) for a in agentes), cell_style),
                Paragraph(join_values(a.placa_agente for a in agentes_validos), cell_style),

                Paragraph(join_values(a.motorista for a in agentes_validos), cell_style),
                Paragraph(join_values(a.placa_motorista for a in agentes_validos), cell_style),

                join_values_nao_carona(a.data_solicitada.strftime("%d/%m/%Y") for a in agentes_validos if a.data_solicitada),
                join_values_nao_carona(a.horario_solicitado.strftime("%H:%M") for a in agentes_validos if a.horario_solicitado),

                join_values(a.data_inicio.strftime("%d/%m/%Y") for a in agentes_validos if a.data_inicio),
                join_values(a.horario_inicio.strftime("%H:%M") for a in agentes_validos if a.horario_inicio),

                join_values(a.data_finalizacao.strftime("%d/%m/%Y") for a in agentes_validos if a.data_finalizacao),
                join_values(a.horario_finalizacao.strftime("%H:%M") for a in agentes_validos if a.horario_finalizacao),

                join_values_nao_carona(str(a.horario_total) for a in agentes_validos if a.horario_total),
                join_values(str(a.horario_excedente) for a in agentes_validos if a.horario_excedente),

                join_values_nao_carona(str(a.km_inicio) for a in agentes_validos if a.km_inicio is not None),
                join_values_nao_carona(str(a.km_final) for a in agentes_validos if a.km_final is not None),
                join_values_nao_carona(str(a.km_total) for a in agentes_validos if a.km_total is not None),
                join_values(str(a.km_excedente) for a in agentes_validos if a.km_excedente),

                join_values(moeda(a.pedagio) for a in agentes_validos if a.pedagio),
                join_values(a.franquia.nome if a.franquia else "" for a in agentes_validos),

                join_values(moeda(a.valor_agente) for a in agentes if a.valor_agente),
                moeda(item.total_valor_agentes),

                Paragraph(item.ocorrencia or "", cell_style),
                Paragraph(item.nome_user or "", cell_style),

                moeda(item.valor_contrato) if item.valor_contrato else "",
                moeda(item.lucro_total) if item.lucro_total is not None else "",

                "Pago" if item.validar_pagamento else "",

                item.get_status_display(),

                item.nf or "",

            ]

            data.append(row)

        total_row = [""] * len(headers)
        total_row[25] = Paragraph(f"Total Geral Agentes: {moeda(total_geral)}", cell_style)
        total_row[28] = Paragraph(f"Total Contrato: {moeda(total_contrato)}",cell_style)
        total_row[29] = Paragraph(f"Total Lucro: {moeda(total_lucro)}",cell_style)
        data.append(total_row)

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

            ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="faturamento_acompanhamentos.pdf"'
        return response

    def exportar_excel(self, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Acompanhamentos"

        headers = [
            "Protocolo", "Cliente", "Tipo Serviço", "Origem", "Destino",
            "Responsável", "Agente", "Placa Agente",
            "Motorista", "Placa Motorista",
            "Data Solicitada", "Hora Solicitada",
            "Data Inicial", "Hora Inicial",
            "Data Final", "Hora Final",
            "Hora Total", "Hora Excedente",
            "KM Início", "KM Final", "KM Total", "KM Excedente",
            "Pedágio", "Franquia",
            "Valor Agente(s)",
            "Total Agentes",
            "Ocorrência",
            "Usuário",
            "Valor Contrato Cliente",
            "Lucro",
            "Pagamento",
            "Status",
            "NF",

        ]


        sheet.append(headers)

        total_geral = Decimal("0.00")
        total_contrato = Decimal("0.00")
        total_lucro = Decimal("0.00")

        for item in queryset:
            agentes = item.agentes.all()
            agentes_validos = agentes_nao_carona(agentes)

            total_geral += item.total_valor_agentes or Decimal("0.00")
            total_contrato += item.valor_contrato or Decimal("0.00")
            total_lucro += item.lucro_total or Decimal("0.00")

            sheet.append([
                item.id,
                str(item.cliente),
                str(item.tipo_servico),
                item.origem,
                item.destino,

                join_values(a.responsavel_agente.nome for a in agentes_validos),
                join_values(str(a.agente) for a in agentes),
                join_values(a.placa_agente for a in agentes_validos),

                join_values(a.motorista for a in agentes_validos),
                join_values(a.placa_motorista for a in agentes_validos),

                join_values(a.data_solicitada.strftime("%d/%m/%Y") for a in agentes_validos if a.data_solicitada),
                join_values(a.horario_solicitado.strftime("%H:%M") for a in agentes_validos if a.horario_solicitado),

                join_values(a.data_inicio.strftime("%d/%m/%Y") for a in agentes_validos if a.data_inicio),
                join_values(a.horario_inicio.strftime("%H:%M") for a in agentes_validos if a.horario_inicio),

                join_values(a.data_finalizacao.strftime("%d/%m/%Y") for a in agentes_validos if a.data_finalizacao),
                join_values(a.horario_finalizacao.strftime("%H:%M") for a in agentes_validos if a.horario_finalizacao),

                join_values(str(a.horario_total) for a in agentes_validos if a.horario_total),
                join_values(str(a.horario_excedente) for a in agentes_validos if a.horario_excedente),

                join_values(str(a.km_inicio) for a in agentes_validos if a.km_inicio is not None),
                join_values(str(a.km_final) for a in agentes_validos if a.km_final is not None),
                join_values(str(a.km_total) for a in agentes_validos if a.km_total is not None),
                join_values(str(a.km_excedente) for a in agentes_validos if a.km_excedente),

                join_values(moeda(a.pedagio) for a in agentes_validos if a.pedagio),
                join_values(a.franquia.nome if a.franquia else "" for a in agentes_validos),

                join_values(moeda(a.valor_agente) for a in agentes if a.valor_agente),
                moeda(item.total_valor_agentes),

                item.ocorrencia,
                item.nome_user,

                moeda(item.valor_contrato) if item.valor_contrato else "",
                moeda(item.lucro_total) if item.lucro_total is not None else "",

                "Pago" if item.validar_pagamento else "",

                item.get_status_display(),

                item.nf or "",
            ])

            row_number = sheet.max_row

            multi_columns = [
                6, 7, 8, 9, 10,
                11, 12, 13, 14, 15, 16,
                17, 18,
                19, 20, 21, 22,
                23, 24, 25
            ]

            max_lines = 1
            for col in multi_columns:
                cell = sheet.cell(row=row_number, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.value:
                    max_lines = max(max_lines, cell.value.count("\n") + 1)

            sheet.row_dimensions[row_number].height = 15 * max_lines

        sheet.append([])

        total_row = [""] * len(headers)
        total_row[25] = f"Total Geral Agentes: {moeda(total_geral)}"
        total_row[28] = f"Total Contrato: {moeda(total_contrato)}"
        total_row[29] = f"Total Lucro: {moeda(total_lucro)}"
        sheet.append(total_row)


        last_row = sheet.max_row
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D3D3D3",
                end_color="D3D3D3",
                fill_type="solid"
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="faturamento_acompanhamentos.xlsx"'
        workbook.save(response)
        return response
    
    def exportar_excel_cliente(self, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Acompanhamentos"

        titulo_custom = next(
            (
                i.campo_personalizado_titulo.strip()
                for i in queryset
                if i.campo_personalizado_titulo and i.campo_personalizado_titulo.strip()
            ),
            " "
        )

        headers = [
            "Protocolo",
            "Cliente",
            titulo_custom,
            "Origem",
            "Destino",

            "Agente",
            "Placa Agente",
            "Motorista",
            "Placa Motorista",

            "Data Solicitada",
            "Hora Solicitada",
            "Data Inicial",
            "Hora Inicial",
            "Data Final",
            "Hora Final",

            "KM Início",
            "KM Final",
            "KM Total",

            "Franquia de Horas",
            "Hora Total",
            "Hora Excedente",
            "Valor Hora Excedente",
            "Valor Hora Excedente Total",

            "Franquia de KM",
            "KM Excedente",
            "Valor KM Excedente",
            "Valor KM Excedente Total",

            "Valor Diária",
            "Pedágio",
            "Valor Total",

            "Tipo Serviço"
        ]

        sheet.append(headers)

        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        sheet.freeze_panes = "A2"

        total_contrato = Decimal("0.00")

        def join_values_excel(values):
            return "\n\n".join(str(v) for v in values if v)

        for item in queryset:
            agentes = item.agentes.all()

            agentes_visuais = agentes              # TODOS (inclui carona)
            agentes_validos = agentes_nao_carona(agentes)  # só principais

            total_contrato += item.valor_contrato or Decimal("0.00")

            sheet.append([
                item.id,
                str(item.cliente),
                item.campo_personalizado_valor or "",
                item.origem,
                item.destino,

                # --- AGENTES ---
                join_values_excel(str(a.agente) for a in agentes_visuais),
                join_values(a.placa_agente for a in agentes_validos),
                join_values(a.motorista for a in agentes_validos),
                join_values(a.placa_motorista for a in agentes_validos),

                join_values(a.data_solicitada.strftime("%d/%m/%Y") for a in agentes_validos if a.data_solicitada),
                join_values(a.horario_solicitado.strftime("%H:%M") for a in agentes_validos if a.horario_solicitado),

                join_values(a.data_inicio.strftime("%d/%m/%Y") for a in agentes_validos if a.data_inicio),
                join_values(a.horario_inicio.strftime("%H:%M") for a in agentes_validos if a.horario_inicio),

                join_values(a.data_finalizacao.strftime("%d/%m/%Y") for a in agentes_validos if a.data_finalizacao),
                join_values(a.horario_finalizacao.strftime("%H:%M") for a in agentes_validos if a.horario_finalizacao),

                join_values(str(a.km_inicio) for a in agentes_validos if a.km_inicio is not None),
                join_values(str(a.km_final) for a in agentes_validos if a.km_final is not None),
                join_values(str(a.km_total) for a in agentes_validos if a.km_total is not None),

                str(item.cliente.franquia_horas) if item.cliente and item.cliente.franquia_horas is not None else "",
                join_values(str(a.horario_total) for a in agentes_validos if a.horario_total),
                join_values(str(a.horario_excedente) for a in agentes_validos if a.horario_excedente),
                moeda(item.cliente.valor_horas_excedente) if item.cliente and item.cliente.valor_horas_excedente else "",
                join_values(
                    moeda(
                        (Decimal(a.horario_excedente.total_seconds()) / Decimal(3600))
                        * item.cliente.valor_horas_excedente
                    )
                    for a in agentes_validos
                    if (
                        a.horario_excedente
                        and item.cliente
                        and item.cliente.valor_horas_excedente
                    )
                ),

                str(item.cliente.franquia_km) if item.cliente and item.cliente.franquia_km is not None else "",
                join_values(str(a.km_excedente) for a in agentes_validos if a.km_excedente),
                moeda(item.cliente.valor_km_excedente) if item.cliente and item.cliente.valor_km_excedente else "",
                join_values(
                    moeda(
                        Decimal(a.km_excedente or 0) * item.cliente.valor_km_excedente
                    )
                    for a in agentes_validos
                    if item.cliente and item.cliente.valor_km_excedente and a.km_excedente
                ),

                moeda(item.cliente.valor_acionamento) if item.cliente and item.cliente.valor_acionamento else "",
                join_values(moeda(a.pedagio) for a in agentes_validos if a.pedagio),
                moeda(item.valor_contrato),

                str(item.tipo_servico),
            ])

            row_number = sheet.max_row

            multi_columns = [
                6,
                7, 8, 9, 10, 11,
                12, 13, 14, 15, 16, 17,
                18, 19,
                20, 21, 22, 23,
                24, 25, 26
            ]

            max_lines = 1
            for col in multi_columns:
                cell = sheet.cell(row=row_number, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.value:
                    max_lines = max(max_lines, cell.value.count("\n") + 1)

            sheet.row_dimensions[row_number].height = 15 * max_lines

        sheet.append([])

        total_row = [""] * len(headers)
        total_row[29] = f"TOTAL GERAL (CONTRATOS): {moeda(total_contrato)}"
        sheet.append(total_row)

        last_row = sheet.max_row
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D3D3D3",
                end_color="D3D3D3",
                fill_type="solid"
            )

        for col in sheet.columns:
            max_length = 0
            column_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="faturamento_acompanhamentos.xlsx"'
        workbook.save(response)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["franquias"] = FranquiaAgente.objects.all().order_by("nome")

        context["pendentes_count"] = self.get_pendentes_queryset().count()
        
        context["status"] = self.request.GET.get("status", "")
        context["data"] = self.request.GET.get("data", "")
        context["data2"] = self.request.GET.get("data2", "")
        context["agente"] = self.request.GET.get("agente", "")
        context["cliente"] = self.request.GET.get("cliente", "")
        context["tipo_cadastro"] = self.request.GET.get("tipo_cadastro", "")
        context["tipos_cadastro"] = self.TIPOS_CADASTRO
        
        return context

# ACOMPANHAMENTO PANICO

class AcompanhamentoPanicoListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = registroacompanhamento
    template_name = "acompanhamento_panico.html"
    context_object_name = "itens"
    permission_required = "acompanhamentos.view_registroacompanhamento"

    TIPOS_CADASTRO = [
        ("", "Todos"),
        ("acompanhamento", "Acompanhamento"),
        ("pronta_resposta", "Pronta Resposta"),
    ]

    def get_pendentes_queryset(self):
        return registroacompanhamento.objects.filter(
            Q(tipo_servico__isnull=True) |
            Q(agentes__responsavel_agente__isnull=True) |
            Q(agentes__agente__isnull=True) |
            Q(agentes__data_solicitada__isnull=True) |
            Q(agentes__horario_solicitado__isnull=True)
        ).distinct()

    def get_queryset(self):
        qs = (
            registroacompanhamento.objects
            .prefetch_related(
                "agentes",
                "agentes__agente",
                "agentes__franquia"
            )
            .select_related("cliente")
        )

        pendente = self.request.GET.get("pendente")
        if pendente == "true":
            return self.get_pendentes_queryset().annotate(
                status_rank=Case(
                    When(status_acompanhamento="missao_aceita", then=Value(0)),
                    When(status_acompanhamento="agendada", then=Value(1)),
                    When(status_acompanhamento="pendente", then=Value(2)),
                    When(status_acompanhamento="em_deslocamento", then=Value(3)),
                    When(status_acompanhamento="no_local", then=Value(4)),
                    When(status_acompanhamento="em_andamento", then=Value(5)),
                    When(status_acompanhamento="sem_sinal", then=Value(6)),
                    When(status_acompanhamento="concluido", then=Value(99)),
                    default=Value(50),
                    output_field=IntegerField(),
                )
            ).order_by("status_rank", "-criado_em")

        data = self.request.GET.get("data")
        data2 = self.request.GET.get("data2")
        agente = self.request.GET.get("agente")
        cliente = self.request.GET.get("cliente")
        responsavel = self.request.GET.get("responsavel")
        tipo_cadastro = self.request.GET.get("tipo_cadastro")

        if data and data2:
            if data > data2:
                data, data2 = data2, data
            qs = qs.filter(agentes__data_solicitada__range=[data, data2])
        elif data:
            qs = qs.filter(agentes__data_solicitada=data)
        elif data2:
            qs = qs.filter(agentes__data_solicitada=data2)

        if agente:
            qs = qs.filter(agentes__agente__nome__icontains=agente)

        if cliente:
            qs = qs.filter(cliente__nome__icontains=cliente)

        if responsavel:
            qs = qs.filter(
                agentes__responsavel_agente_id=responsavel
            )

        if tipo_cadastro in {"acompanhamento", "pronta_resposta"}:
            qs = qs.filter(tipo_servico__tipo_cadastro=tipo_cadastro)

        qs = qs.annotate(
            status_rank=Case(
                When(status_acompanhamento="missao_aceita", then=Value(0)),
                When(status_acompanhamento="agendada", then=Value(1)),
                When(status_acompanhamento="pendente", then=Value(2)),
                # depois vem o resto
                When(status_acompanhamento="em_deslocamento", then=Value(3)),
                When(status_acompanhamento="no_local", then=Value(4)),
                When(status_acompanhamento="em_andamento", then=Value(5)),
                When(status_acompanhamento="sem_sinal", then=Value(6)),
                # ✅ concluídos por último
                When(status_acompanhamento="concluido", then=Value(99)),
                default=Value(50),
                output_field=IntegerField(),
            )
        )

        return qs.distinct().order_by("status_rank", "-criado_em")

    def render_to_response(self, context, **response_kwargs):
        queryset = context["itens"]

        if self.request.GET.get("export") == "excel_panico":
            return self.exportar_excel_panico(queryset)
        
        if self.request.GET.get("export") == "pdf_panico":
            return self.exportar_pdf_panico(queryset)

        return super().render_to_response(context, **response_kwargs)

    def exportar_pdf_panico(self, queryset):
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=(2000, 900),
            topMargin=20,
            leftMargin=20,
            rightMargin=20,
            bottomMargin=20
        )

        elements = []

        styles = getSampleStyleSheet()
        title_style = styles["Title"]

        cell_style = ParagraphStyle(
            name="CellStyle",
            fontName="Helvetica",
            fontSize=7,
            leading=9,
            wordWrap="CJK",
            alignment=0
        )

        title = Paragraph("Relatório de Acompanhamentos - Pânico", title_style)
        title.alignment = 1
        elements.append(title)
        elements.append(Paragraph("<br/>", styles["Normal"]))

        headers = [
            "Protocolo", "Status", "Cliente", "Tipo Serviço", "Origem", "Destino",
            "Responsável Agente", "Agente", "Placa Agente",
            "Motorista", "Placa Motorista",
            "Data Solicitada", "Horário Solicitado"
        ]

        data = [headers]

        for item in queryset:
            agentes = item.agentes.all()
            agentes_validos = agentes_nao_carona(agentes)

            row = [
                str(item.id),
                item.get_status_acompanhamento_display() if hasattr(item, 'get_status_acompanhamento_display') else item.status,
                Paragraph(str(item.cliente), cell_style),
                Paragraph(str(item.tipo_servico) if item.tipo_servico else "", cell_style),
                Paragraph(item.origem or "", cell_style),
                Paragraph(item.destino or "", cell_style),

                Paragraph(
                    join_values(a.responsavel_agente.nome for a in agentes_validos if a.responsavel_agente),
                    cell_style
                ),
                Paragraph(join_values(str(a.agente) for a in agentes), cell_style),
                Paragraph(join_values(a.placa_agente for a in agentes_validos if a.placa_agente), cell_style),

                Paragraph(join_values(a.motorista for a in agentes_validos if a.motorista), cell_style),
                Paragraph(join_values(a.placa_motorista for a in agentes_validos if a.placa_motorista), cell_style),

                join_values_nao_carona(a.data_solicitada.strftime("%d/%m/%Y") for a in agentes_validos if a.data_solicitada),
                join_values_nao_carona(a.horario_solicitado.strftime("%H:%M") for a in agentes_validos if a.horario_solicitado),
            ]

            data.append(row)

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="acompanhamentos_panico.pdf"'
        return response

    def exportar_excel_panico(self, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Acompanhamentos Panico"

        headers = [
            "Protocolo", "Status", "Cliente", "Tipo Serviço", "Origem", "Destino",
            "Responsável Agente", "Agente", "Placa Agente",
            "Motorista", "Placa Motorista",
            "Data Solicitada", "Horário Solicitado"
        ]

        sheet.append(headers)

        for item in queryset:
            agentes = item.agentes.all()
            agentes_validos = agentes_nao_carona(agentes)

            sheet.append([
                item.id,
                item.get_status_acompanhamento_display() if hasattr(item, 'get_status_acompanhamento_display') else item.status,
                str(item.cliente),
                str(item.tipo_servico) if item.tipo_servico else "",
                item.origem,
                item.destino,

                join_values(a.responsavel_agente.nome for a in agentes_validos if a.responsavel_agente),
                join_values(str(a.agente) for a in agentes),
                join_values(a.placa_agente for a in agentes_validos if a.placa_agente),

                join_values(a.motorista for a in agentes_validos if a.motorista),
                join_values(a.placa_motorista for a in agentes_validos if a.placa_motorista),

                join_values(a.data_solicitada.strftime("%d/%m/%Y") for a in agentes_validos if a.data_solicitada),
                join_values(a.horario_solicitado.strftime("%H:%M") for a in agentes_validos if a.horario_solicitado),
            ])

            row_number = sheet.max_row

            multi_columns = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]

            max_lines = 1
            for col in multi_columns:
                cell = sheet.cell(row=row_number, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.value:
                    max_lines = max(max_lines, cell.value.count("\n") + 1)

            sheet.row_dimensions[row_number].height = 15 * max_lines

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="acompanhamentos_panico.xlsx"'
        workbook.save(response)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        itens = context["itens"]
        for item in itens:
            item.link_app = gerar_link_app_missao(item, request=self.request)

        context["nao_concluidos_count"] = registroacompanhamento.objects.exclude(
            status_acompanhamento="concluido"
        ).count()

        context["tipo_cadastro"] = self.request.GET.get("tipo_cadastro", "")
        context["tipos_cadastro"] = self.TIPOS_CADASTRO

        return context


class MonitoramentoAoVivoView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = "acompanhamento_monitoramento_ao_vivo.html"
    permission_required = "acompanhamentos.view_registroacompanhamento"

@login_required
def acompanhamento_missao(request, pk):
    acompanhamento = get_object_or_404(registroacompanhamento, pk=pk)

    return render(
        request,
        "acompanhamento_missao.html",
        {
            "acompanhamento": acompanhamento
        }
    )

@login_required
def acompanhamento_mapa(request, pk):
    acompanhamento = get_object_or_404(registroacompanhamento, pk=pk)

    # ✅ Descobre o UUID do Supabase salvo no seu model (ajuste se necessário)
    supabase_mission_id = (
        getattr(acompanhamento, "supabase_mission_id", None)
        or getattr(acompanhamento, "mission_uuid", None)
        or getattr(acompanhamento, "supabase_id", None)
        or getattr(acompanhamento, "mission_id", None)
    )
    supabase_mission_id = str(supabase_mission_id or "")

    # Nome do agente
    agente_vinculo = acompanhamento.agentes.filter(tipo_agente="principal").first() or acompanhamento.agentes.first()
    nome_agente = agente_vinculo.agente.nome if agente_vinculo and agente_vinculo.agente else "Não atribuído"

    # Pontos iniciais do Supabase (para render inicial sem depender do JS)
    localizacoes = []
    if supabase_mission_id:
        sb = get_supabase()
        tracking_res = (
            sb.table("mission_tracking")
            .select("id,lat,lng,timestamp,created_at")
            .eq("mission_id", supabase_mission_id)
            .order("timestamp", desc=False)
            .limit(5000)
            .execute()
        )
        tracking_rows = tracking_res.data or []

        localizacoes = [
            {
                "id": r.get("id"),
                "latitude": float(r["lat"]) if r.get("lat") is not None else None,
                "longitude": float(r["lng"]) if r.get("lng") is not None else None,
                "criado_em": r.get("timestamp") or r.get("created_at"),
                "origem": getattr(acompanhamento, "origem", "") or "",
            }
            for r in tracking_rows
            if r.get("lat") is not None and r.get("lng") is not None
        ]

    return render(
        request,
        "acompanhamento_mapa.html",
        {
            "acompanhamento": acompanhamento,
            "nome_agente": nome_agente,
            "origem": getattr(acompanhamento, "origem", "") or "",
            "supabase_mission_id": supabase_mission_id,
            "localizacoes_json": json.dumps(localizacoes, cls=DjangoJSONEncoder),
        }
    )

@login_required
def acompanhamento_mapa_supabase(request, mission_id):
    """
    Renderiza o mapa usando APENAS o UUID do Supabase.
    Não depende do registroacompanhamento do Django.
    """

    sb = get_supabase()
    mission_uuid = str(mission_id)

    # 1) Buscar a missão no Supabase (pra mostrar origem/status/agente)
    mission_res = (
        sb.table("missions_control")
        .select("*")
        .eq("id", mission_uuid)
        .maybe_single()
        .execute()
    )
    mission = mission_res.data or {}

    # 2) Buscar tracking inicial (pra desenhar no load)
    tracking_res = (
        sb.table("mission_tracking")
        .select("id,lat,lng,timestamp,created_at")
        .eq("mission_id", mission_uuid)
        .order("timestamp", desc=False)
        .limit(5000)
        .execute()
    )
    tracking_rows = tracking_res.data or []

    localizacoes = [
        {
            "id": r.get("id"),
            "latitude": float(r["lat"]) if r.get("lat") is not None else None,
            "longitude": float(r["lng"]) if r.get("lng") is not None else None,
            "criado_em": r.get("timestamp") or r.get("created_at"),
            "origem": (mission.get("origem") or ""),
        }
        for r in tracking_rows
        if r.get("lat") is not None and r.get("lng") is not None
    ]

    return render(
        request,
        "acompanhamento_mapa.html",
        {
            "mission_id": mission_uuid,
            "supabase_mission_id": mission_uuid,
            "mission": mission,                         # pra aparecer no header do template
            "origem": mission.get("origem") or "-",                         # pra aparecer no header do template
            "status": mission.get("status") or "-",                         # pra aparecer no header do template
            "nome_agente": mission.get("agente") or "-",# idem
            "localizacoes_json": json.dumps(localizacoes, cls=DjangoJSONEncoder),
        }
    )

def validar_pagamento(request, id):
    acompanhamento = get_object_or_404(registroacompanhamento, id=id)

    acompanhamento.validar_pagamento = True
    acompanhamento.save(update_fields=["validar_pagamento"])

    return redirect("acompanhamentosListFaturamento")

def validar_pagamento(request, id):
    acompanhamento = get_object_or_404(registroacompanhamento, id=id)

    acompanhamento.recalcular_financeiro(commit=False)

    acompanhamento.validar_pagamento = '1'

    acompanhamento.save(update_fields=[
        "validar_pagamento",
        "valor_contrato",
        "lucro_total",
    ])

    return redirect('acompanhamentosListFaturamento')

def usuario_eh_faturamento_master(user):
    return user.groups.filter(name="FaturamentoMaster").exists()

@require_POST
def atualizar_status_acompanhamento(request):

    if not request.user.groups.filter(name="FaturamentoMaster").exists():
        return JsonResponse({
            "success": False,
            "error": "Sem permissão para alterar status."
        }, status=403)

    try:
        data = json.loads(request.body)

        acompanhamento = get_object_or_404(
            registroacompanhamento,
            id=data.get("acompanhamento_id")
        )

        status = data.get("status")

        if status not in dict(registroacompanhamento.STATUS_CHOICES):
            return JsonResponse({"success": False, "error": "Status inválido."})

        acompanhamento.status = status

        if status != "faturado":
            acompanhamento.nf = None

        acompanhamento.save(update_fields=["status", "nf"])

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

@require_POST
def atualizar_nf_acompanhamento(request):

    if not request.user.groups.filter(name="FaturamentoMaster").exists():
        return JsonResponse({
            "success": False,
            "error": "Sem permissão para alterar NF."
        }, status=403)

    try:
        data = json.loads(request.body)

        acompanhamento = get_object_or_404(
            registroacompanhamento,
            id=data.get("acompanhamento_id")
        )

        if acompanhamento.status != "faturado":
            return JsonResponse({
                "success": False,
                "error": "NF só pode ser informada quando status for faturado."
            })

        acompanhamento.nf = data.get("nf")
        acompanhamento.save(update_fields=["nf"])

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

# @login_required
# @require_POST
# def atualizar_valor_contrato_cliente(request):
#     try:
#         data = json.loads(request.body)

#         acompanhamento = get_object_or_404(
#             registroacompanhamento,
#             id=data.get("acompanhamento_id")
#         )

#         valor_raw = data.get("valor_contrato_cliente")

#         if not valor_raw:
#             acompanhamento.valor_contrato = None
#             acompanhamento.lucro_total = None
#         else:
#             try:
#                 valor_contrato = Decimal(
#                     str(valor_raw)
#                     .replace(' ', '')
#                     .replace('.', '')
#                     .replace(',', '.')
#                 )
#             except InvalidOperation:
#                 return JsonResponse(
#                     {"success": False, "error": "Valor monetário inválido"},
#                     status=400
#                 )

#             acompanhamento.valor_contrato = valor_contrato

#             total_agentes = acompanhamento.total_valor_agentes or Decimal("0.00")
#             acompanhamento.lucro_total = valor_contrato - total_agentes

#         acompanhamento.save(update_fields=["valor_contrato", "lucro_total"])

#         return JsonResponse({
#             "success": True,
#             "lucro_total": f"{acompanhamento.lucro_total:.2f}" if acompanhamento.lucro_total is not None else ""
#         })

#     except Exception as e:
#         return JsonResponse(
#             {"success": False, "error": str(e)},
#             status=400
#         )

@login_required
@permission_required("acompanhamentos.change_registroacompanhamento", raise_exception=False)
def atualizar_franquia_acompanhamento(request):

    if request.method != "POST":
        return JsonResponse({"success": False}, status=405)

    data = json.loads(request.body)

    acompanhamento = get_object_or_404(
        registroacompanhamento,
        id=data.get("acompanhamento_id")
    )

    franquia = None
    franquia_id = data.get("franquia_id")

    if franquia_id:
        franquia = get_object_or_404(FranquiaAgente, id=franquia_id)

    agentes_response = []

    for agente in acompanhamento.agentes.all():
        agente.franquia = franquia
        agente.save()  # dispara save() → recalcula tudo

        agentes_response.append({
            "horario_excedente": str(agente.horario_excedente) if agente.horario_excedente else "—",
            "km_excedente": agente.km_excedente if agente.km_excedente and agente.km_excedente > 0 else "—",
            "valor_agente": f"R$ {agente.valor_agente:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                if agente.valor_agente else "—",
        })

    # Só recalcula financeiro se já estiver validado
    if acompanhamento.status_acompanhamento == "verificado":
        acompanhamento.recalcular_financeiro()
    acompanhamento.nome_user = request.user.get_full_name() or request.user.username
    acompanhamento.save(update_fields=["nome_user"])

    return JsonResponse({
        "success": True,
        "agentes": agentes_response,
        "total": acompanhamento.total_valor_agentes_formatado,
        "usuario": acompanhamento.nome_user,
    })

def format_decimal(value):
    if value is None:
        return ""
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ------------------------------------------------------
#               ACOMPANHAMENTO DASHBOARD
# ------------------------------------------------------
class AcompanhamentoDashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = "acompanhamento_dashboard.html"
    permission_required = "acompanhamentos.view_listacompanhamento"

@login_required
def acompanhamento_dashboard_data(request):

    periodo = request.GET.get('periodo', 'mensal')
    hoje = timezone.now()

    if periodo == 'semanal':
        data_inicio = hoje - timedelta(days=7)
    elif periodo == 'quinzenal':
        data_inicio = hoje - timedelta(days=15)
    else:
        data_inicio = hoje - timedelta(days=30)

    dados = (
        registrodeacompanhamento.objects
        .filter(criado_em__gte=data_inicio)
        .exclude(cliente__isnull=True)
        .exclude(cliente="")
        .exclude(cliente__icontains="teste")
        .values('cliente')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )

    return JsonResponse({
        'labels': [d['cliente'] for d in dados],
        'valores': [d['total'] for d in dados],
    })
# =================================================================
# SUBSTITUIR no views.py do Sistema INT
# =================================================================

from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from .models import RequisicaoSolicitacao, Cliente
from collections import OrderedDict
import json


class RequisicaoSolicitacaoListView(LoginRequiredMixin, ListView):
    model = RequisicaoSolicitacao
    template_name = "requisicoes_solicitacao_list.html"
    context_object_name = "requisicoes_raw"
    paginate_by = None

    def _is_cancelado(self, queryset):
        return queryset.filter(
            Q(status_requisicao__iexact="cancelado") |
            Q(status_requisicao__iexact="cancelada") |
            Q(status_requisicao__iexact="canceladas")
        )

    def _apply_common_filters(self, queryset):
        cliente = self.request.GET.get("cliente")
        placa = self.request.GET.get("placa")
        motorista = self.request.GET.get("motorista")
        data_inicio = self.request.GET.get("data_inicio")
        data_fim = self.request.GET.get("data_fim")
        busca = self.request.GET.get("busca")
        tipo_cadastro = self.request.GET.get("tipo_cadastro")

        if cliente:
            queryset = queryset.filter(cliente__nome__icontains=cliente)

        if placa:
            queryset = queryset.filter(placa__icontains=placa)

        if motorista:
            queryset = queryset.filter(motorista__icontains=motorista)

        if data_inicio and data_fim:
            queryset = queryset.filter(data_agendamento__range=[data_inicio, data_fim])
        elif data_inicio:
            queryset = queryset.filter(data_agendamento=data_inicio)

        if busca:
            queryset = queryset.filter(
                Q(origem__icontains=busca) |
                Q(destino__icontains=busca) |
                Q(nome_user__icontains=busca)
            )

        if tipo_cadastro in {"acompanhamento", "pronta_resposta"}:
            queryset = queryset.filter(tipo_servico__tipo_cadastro=tipo_cadastro)

        return queryset

    def _count_grupos(self, queryset):
        total = 0
        vistos = set()

        for req in queryset:
            key = req.requisicao_id_externo or f"solo_{req.pk}"
            if key in vistos:
                continue
            vistos.add(key)
            total += 1

        return total

    def get_queryset(self):
        base_qs = (
            RequisicaoSolicitacao.objects
            .select_related("cliente", "tipo_servico", "missao")
            .filter(solicitado=False)
        )

        queryset = self._apply_common_filters(base_qs)

        mostrar_canceladas = str(self.request.GET.get("canceladas", "")).strip().lower() in {
            "1", "true", "t", "yes", "y", "sim"
        }

        if mostrar_canceladas:
            queryset = self._is_cancelado(queryset)
        else:
            queryset = queryset.exclude(
                Q(status_requisicao__iexact="cancelado") |
                Q(status_requisicao__iexact="cancelada") |
                Q(status_requisicao__iexact="canceladas")
            )

        return queryset.order_by("-criado_em")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filtros"] = self.request.GET

        context["tipos_cadastro"] = [
            ("", "Todos"),
            ("acompanhamento", "Acompanhamento"),
            ("pronta_resposta", "Pronta Resposta"),
        ]

        qs = context["requisicoes_raw"]
        agrupado = OrderedDict()

        for req in qs:
            key = req.requisicao_id_externo or f"solo_{req.pk}"

            if key not in agrupado:
                agrupado[key] = {
                    "primeiro": req,
                    "servicos": [],
                    "requisicao_id_externo": req.requisicao_id_externo,
                    "tem_multiplos": False,
                }

            agrupado[key]["servicos"].append(req)

        for grupo in agrupado.values():
            grupo["tem_multiplos"] = len(grupo["servicos"]) > 1

            if grupo["tem_multiplos"]:
                grupo["servicos_json"] = json.dumps([
                    {
                        "pk": s.pk,
                        "id_externo": s.id_externo,
                        "tipo_servico": str(s.tipo_servico),
                        "motorista": s.motorista,
                        "placa": s.placa,
                        "agente": s.agente or "-",
                        "placa_agente": s.placa_agente or "-",
                        "data_agendamento": s.data_agendamento.strftime("%d/%m/%Y") if s.data_agendamento else "-",
                        "horario_agendamento": s.horario_agendamento.strftime("%H:%M") if s.horario_agendamento else "-",
                        "url_criar": reverse("RequisicaoCreate", args=[s.pk]),
                        "comboio": bool(getattr(s, "comboio", False)),
                        "quantidade_veiculos_comboio": (
                            int(s.quantidade_veiculos_comboio)
                            if getattr(s, "quantidade_veiculos_comboio", None)
                            else None
                        ),
                        "is_reuso": bool(getattr(s, "is_reuso", False)),
                        "agente_nome_reuso": getattr(s, "agente_nome_reuso", "") or "",
                        "agente_placa_reuso": getattr(s, "agente_placa_reuso", "") or "",
                    }
                    for s in grupo["servicos"]
                ])

        context["grupos"] = list(agrupado.values())
        context["total_grupos"] = len(agrupado)

        base_qs = self._apply_common_filters(
            RequisicaoSolicitacao.objects
            .select_related("cliente", "tipo_servico", "missao")
            .filter(solicitado=False)
        ).order_by("-criado_em")

        pendentes_qs = base_qs.exclude(
            Q(status_requisicao__iexact="cancelado") |
            Q(status_requisicao__iexact="cancelada") |
            Q(status_requisicao__iexact="canceladas")
        )
        canceladas_qs = self._is_cancelado(base_qs)

        context["total_grupos_pendentes"] = self._count_grupos(pendentes_qs)
        context["total_grupos_cancelados"] = self._count_grupos(canceladas_qs)
        context["mostrando_canceladas"] = str(self.request.GET.get("canceladas", "")).strip().lower() in {
            "1", "true", "t", "yes", "y", "sim"
        }

        return context

from django.views.generic import CreateView
from django.urls import reverse_lazy, reverse
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

from .models import registroacompanhamento, registroacompanhamentoagente, RequisicaoSolicitacao
from .forms import RegistroAcompanhamentoFromRequisicaoForm, RegistroAcompanhamentoAgenteCreateFormSet

class AcompanhamentoFromRequisicaoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = registroacompanhamento
    form_class = RegistroAcompanhamentoFromRequisicaoForm
    template_name = "acompanhamento2_create.html"
    success_url = reverse_lazy("requisicao_solicitacao_list")
    permission_required = "acompanhamentos.add_registroacompanhamento"

    def get_requisicao(self):
        """Busca a requisição usada como base para preencher o acompanhamento."""
        pk = self.kwargs.get("pk")
        return get_object_or_404(RequisicaoSolicitacao, pk=pk)

    def get_form_kwargs(self):
        """Envia a requisição para dentro do form (caso o form use essa info para validação)."""
        kwargs = super().get_form_kwargs()
        kwargs["requisicao"] = self.get_requisicao()
        return kwargs

    def get_context_data(self, **kwargs):
        """Monta o contexto do template + formset dos agentes."""
        context = super().get_context_data(**kwargs)
        requisicao = self.get_requisicao()

        # Requisição no template
        context["requisicao"] = requisicao

        # Lista de agentes cadastrados (para selects/autocomplete no template)
        context["agentes_cadastrados"] = (
            registrodeagenteacompanhamento.objects.all().order_by("nome")
        )

        # Formset: se POST, mantém dados enviados; se GET, cria vazio e injeta initial no 1º form
        if self.request.POST:
            context["agentes_formset"] = RegistroAcompanhamentoAgenteCreateFormSet(
                self.request.POST,
                prefix="agentes",
                form_kwargs={
                    "cliente_id": requisicao.cliente_id,
                    "tipo_servico_id": requisicao.tipo_servico_id,
                },
            )
        else:
            formset = RegistroAcompanhamentoAgenteCreateFormSet(
                prefix="agentes",
                initial=[{
                    "tipo_servico": requisicao.tipo_servico.pk if requisicao.tipo_servico else None,
                    "motorista": requisicao.motorista,
                    "placa_motorista": requisicao.placa,
                    "numero_motorista": requisicao.numero_motorista,
                    "data_solicitada": requisicao.data_agendamento,
                    "horario_solicitado": requisicao.horario_agendamento,
                }],
                form_kwargs={
                    "cliente_id": requisicao.cliente_id,
                    "tipo_servico_id": requisicao.tipo_servico_id,
                },
            )

            # Pré-preenche o primeiro form do formset usando dados da requisição
            if formset.forms:
                if requisicao.is_reuso:
                    # tenta achar por nome exato (você pode trocar por icontains se precisar)
                    agente_obj = (
                        registrodeagenteacompanhamento.objects
                        .filter(nome__iexact=(requisicao.agente_nome_reuso or "").strip())
                        .first()
                    )
                    if agente_obj:
                        formset.forms[0].initial["agente"] = agente_obj.pk

                    # placa do agente no card
                    if requisicao.agente_placa_reuso:
                        formset.forms[0].initial["placa_agente"] = requisicao.agente_placa_reuso

                formset.forms[0].initial = formset.forms[0].initial

            context["agentes_formset"] = formset

        return context

    def get_initial(self):
        """Pré-preenche o form principal com os dados da requisição."""
        initial = super().get_initial()
        requisicao = self.get_requisicao()

        initial.update({
            "cliente": requisicao.cliente,
            "tipo_servico": requisicao.tipo_servico,
            "origem": requisicao.origem,
            "origem_2": requisicao.origem_2 or "",
            "origem_3": requisicao.origem_3 or "",
            "destino": requisicao.destino or "",
            "destino_2": requisicao.destino_2 or "",
            "destino_3": requisicao.destino_3 or "",
            "latitude_origem": requisicao.latitude_origem,
            "longitude_origem": requisicao.longitude_origem,
            "latitude_origem2": requisicao.latitude_origem_2,
            "longitude_origem2": requisicao.longitude_origem_2,
            "latitude_origem3": requisicao.latitude_origem_3,
            "longitude_origem3": requisicao.longitude_origem_3,
            "latitude_destino": requisicao.latitude_destino,
            "longitude_destino": requisicao.longitude_destino,
            "latitude_destino_2": requisicao.latitude_destino_2,
            "longitude_destino_2": requisicao.longitude_destino_2,
            "latitude_destino_3": requisicao.latitude_destino_3,
            "longitude_destino_3": requisicao.longitude_destino_3,
            "campo_personalizado_titulo": requisicao.campo_personalizado_titulo or "",
            "campo_personalizado_valor": requisicao.campo_personalizado_valor or "",
            "ocorrencia": requisicao.ocorrencia or "",
        })

        return initial

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        agentes_formset = context["agentes_formset"]
        requisicao = self.get_requisicao()

        if not agentes_formset.is_valid():
            return self.render_to_response(context)

        user = self.request.user
        nome_user = user.get_full_name() or user.username

        base = form.save(commit=False)
        if not getattr(base, "status", None):
            base.status = "pendente"

        # força dados da requisição
        base.cliente = requisicao.cliente
        base.tipo_servico = requisicao.tipo_servico
        km_carro_post, km_moto_post = _resolve_km_estimado_from_post(self.request, requisicao.tipo_servico)
        base.km_estimado_carro = km_carro_post if km_carro_post is not None else requisicao.km_estimado_carro
        base.km_estimado_moto = km_moto_post if km_moto_post is not None else requisicao.km_estimado_moto
        base.nome_user = nome_user
        # ✅ Copia flag is_reuso da requisição
        base.is_reuso = requisicao.is_reuso

        agentes_principais = _extract_principais_from_formset(agentes_formset)
        if not agentes_principais:
            messages.warning(self.request, "Nenhum agente/serviço foi preenchido.")
            return self.render_to_response(context)

        criados = []
        for agente_inst in agentes_principais:
            tipo_serv = agente_inst.tipo_servico or requisicao.tipo_servico

            novo_acomp = _build_clone_acompanhamento(base, tipo_servico=tipo_serv, nome_user=nome_user)

            novo_acomp.requisicao_solicitacao = requisicao
            novo_acomp.km_estimado_carro = base.km_estimado_carro
            novo_acomp.km_estimado_moto = base.km_estimado_moto
            novo_acomp.status_requisicao = requisicao.status_requisicao
            novo_acomp.taxa_cancelamento_percentual = requisicao.taxa_cancelamento_percentual
            novo_acomp.valor_cancelamento = requisicao.valor_cancelamento
            novo_acomp.aceitou_termos_cancelamento = requisicao.aceitou_termos_cancelamento
            novo_acomp.usuario_aceitou_termos_cancelamento = requisicao.usuario_aceitou_termos_cancelamento
            novo_acomp.justificativa_cancelamento = requisicao.justificativa_cancelamento
            novo_acomp.motivos_cancelamento = requisicao.motivos_cancelamento
            novo_acomp.cancelado_em = requisicao.cancelado_em

            novo_acomp.save()

            agente_inst.acompanhamento = novo_acomp
            agente_inst.tipo_servico = agente_inst.tipo_servico or tipo_serv
            agente_inst.save()

            criados.append((novo_acomp, agente_inst))

        # caronas: aplica só no primeiro acompanhamento criado
        caronas = self.request.POST.getlist("carona_agente[]")
        if caronas:
            alvo_acomp = criados[0][0]
            principal = alvo_acomp.agentes.filter(tipo_agente="principal").first()
            _criar_caronas(alvo_acomp, principal, caronas)

            if len(criados) > 1:
                messages.warning(self.request, "Caronas aplicadas apenas no 1º serviço.")

        # marca requisição como solicitada
        requisicao.solicitado = True
        requisicao.save(update_fields=["solicitado"])

        # sync supabase (1 por acompanhamento)
        ok = 0
        for (acomp, _) in criados:
            success, mission_id, error = _sync_supabase_safe(self.request, acomp)
            if success:
                ok += 1
            else:
                messages.warning(self.request, f"Acompanhamento #{acomp.id} criado, mas falhou Supabase: {error}")

        messages.success(
            self.request,
            f"Foram criados {len(criados)} acompanhamentos para esta requisição. "
            f"Supabase ok: {ok}/{len(criados)}."
        )

        for (acomp, agente_inst) in criados:
            try:
                ag_principal = acomp.agentes.filter(
                    tipo_agente="principal"
                ).select_related("agente").first()
                gs_acionamento.notificar_acompanhamento_criado(requisicao, ag_principal)
            except Exception as e:
                logger.warning(f"Erro ao notificar GS para acomp #{acomp.id}: {e}")

        return redirect(self.success_url)

    def form_invalid(self, form):
        """Renderiza a tela novamente exibindo os erros do form."""
        return self.render_to_response(self.get_context_data(form=form))

class FranquiaListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = FranquiaAgente
    template_name = "franquia_list.html"
    context_object_name = "franquias"
    permission_required = "acompanhamentos.view_listfranquia"

    def get_queryset(self):
        queryset = super().get_queryset().select_related('cliente', 'tipo_servico__cliente')

        nome = self.request.GET.get("nome")
        cliente = self.request.GET.get("cliente")

        if nome:
            queryset = queryset.filter(nome__icontains=nome)

        if cliente:
            queryset = queryset.filter(
                Q(cliente__nome__icontains=cliente) |
                Q(tipo_servico__cliente__nome__icontains=cliente)
            )

        return queryset.order_by("-criado_em")

class ClientListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Cliente
    template_name = "cliente_acompanhamento_list.html"
    context_object_name = "clientes"
    permission_required = "view_registroacompanhamento"

    def get_queryset(self):
        queryset = super().get_queryset()

        nome = self.request.GET.get("nome")

        if nome:
            queryset = queryset.filter(nome__icontains=nome)

        return queryset.order_by("-criado_em")

# =================================================================
# ADICIONAR NO views.py DO SISTEMA INT
# (logo abaixo da AcompanhamentoFromRequisicaoCreateView existente)
# =================================================================

class AcompanhamentoFromGrupoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = registroacompanhamento
    form_class = RegistroAcompanhamentoFromRequisicaoForm
    template_name = "acompanhamento2_create.html"
    success_url = reverse_lazy("requisicao_solicitacao_list")
    permission_required = "acompanhamentos.add_registroacompanhamento"

    def get_requisicoes(self):
        """Busca todas as RequisicaoSolicitacao do grupo (mesmo requisicao_id_externo)."""
        req_id_externo = self.kwargs.get("req_id_externo")
        reqs = (
            RequisicaoSolicitacao.objects
            .filter(requisicao_id_externo=req_id_externo, solicitado=False)
            .select_related("cliente", "tipo_servico", "missao")
            .order_by("pk")
        )
        if not reqs.exists():
            from django.http import Http404
            raise Http404("Nenhuma requisição pendente encontrada para este grupo.")
        return reqs

    def get_requisicao_principal(self):
        """Primeira do grupo — usada para dados comuns (cliente, origem, etc.)"""
        return self.get_requisicoes().first()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["requisicao"] = self.get_requisicao_principal()
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        req = self.get_requisicao_principal()

        initial["cliente"] = req.cliente
        initial["tipo_servico"] = req.tipo_servico
        initial["origem"] = req.origem
        initial["origem_2"] = req.origem_2 or ""
        initial["origem_3"] = req.origem_3 or ""
        initial["destino"] = req.destino or ""
        initial["destino_2"] = req.destino_2 or ""
        initial["destino_3"] = req.destino_3 or ""
        initial["latitude_origem"] = req.latitude_origem
        initial["longitude_origem"] = req.longitude_origem
        initial["latitude_origem2"] = req.latitude_origem_2
        initial["longitude_origem2"] = req.longitude_origem_2
        initial["latitude_origem3"] = req.latitude_origem_3
        initial["longitude_origem3"] = req.longitude_origem_3
        initial["latitude_destino"] = req.latitude_destino
        initial["longitude_destino"] = req.longitude_destino
        initial["latitude_destino_2"] = req.latitude_destino_2
        initial["longitude_destino_2"] = req.longitude_destino_2
        initial["latitude_destino_3"] = req.latitude_destino_3
        initial["longitude_destino_3"] = req.longitude_destino_3
        initial["campo_personalizado_titulo"] = req.campo_personalizado_titulo or ""
        initial["campo_personalizado_valor"] = req.campo_personalizado_valor or ""
        initial["ocorrencia"] = req.ocorrencia or ""
        initial["numero_motorista"] = req.numero_motorista or ""

        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        requisicoes = list(self.get_requisicoes())
        req_principal = requisicoes[0]

        context["requisicao"] = req_principal
        context["requisicoes_grupo"] = requisicoes
        context["agentes_cadastrados"] = (
            registrodeagenteacompanhamento.objects.all().order_by("nome")
        )

        if self.request.POST:
            context["agentes_formset"] = RegistroAcompanhamentoAgenteCreateFormSet(
                self.request.POST
            )
        else:
            # Cria formset com N forms = N serviços do grupo
            GrupoFormSet = inlineformset_factory(
                registroacompanhamento,
                registroacompanhamentoagente,
                form=RegistroAcompanhamentoAgenteForm,
                formset=BaseAgenteObrigatorioFormSet,
                extra=len(requisicoes),
                can_delete=True
            )

            if self.request.POST:
                context["agentes_formset"] = GrupoFormSet(self.request.POST)
            else:
                formset = GrupoFormSet(
                    form_kwargs={"cliente_id": req_principal.cliente_id}
                )
                for i, req in enumerate(requisicoes):
                    if i < len(formset.forms):
                        formset.forms[i].initial = {
                            "tipo_servico": req.tipo_servico.pk if req.tipo_servico else None,
                            "motorista": req.motorista,
                            "placa_motorista": req.placa,
                            "placa_agente": req.placa_agente or "",
                            "data_solicitada": req.data_agendamento,
                            "horario_solicitado": req.horario_agendamento,
                        }
                context["agentes_formset"] = formset

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        agentes_formset = context["agentes_formset"]
        requisicoes = list(self.get_requisicoes())

        if not agentes_formset.is_valid():
            return self.render_to_response(context)

        user = self.request.user
        nome_user = user.get_full_name() or user.username

        # base (não salva)
        base = form.save(commit=False)
        if not getattr(base, "status", None):
            base.status = "pendente"

        # dados comuns (do “principal”)
        req_principal = requisicoes[0]
        base.cliente = req_principal.cliente
        base.km_estimado_carro = req_principal.km_estimado_carro
        base.km_estimado_moto = req_principal.km_estimado_moto
        base.nome_user = nome_user
        # ✅ Copia is_reuso da requisição principal
        base.is_reuso = req_principal.is_reuso

        criados = []
        reqs_para_marcar = []

        # cada form corresponde a um serviço (mesma ordem do get_context_data inicial)
        for i, f in enumerate(agentes_formset.forms):
            if not getattr(f, "cleaned_data", None):
                continue
            if f.cleaned_data.get("DELETE"):
                continue
            if not f.has_changed():
                continue
            if i >= len(requisicoes):
                continue

            req = requisicoes[i]
            agente_inst = f.save(commit=False)
            if getattr(agente_inst, "tipo_agente", None) == "carona":
                continue

            tipo_serv = agente_inst.tipo_servico or req.tipo_servico

            novo_acomp = _build_clone_acompanhamento(base, tipo_servico=tipo_serv, nome_user=nome_user)

            novo_acomp.requisicao_solicitacao = req
            novo_acomp.km_estimado_carro = req.km_estimado_carro
            novo_acomp.km_estimado_moto = req.km_estimado_moto
            novo_acomp.status_requisicao = req.status_requisicao
            novo_acomp.taxa_cancelamento_percentual = req.taxa_cancelamento_percentual
            novo_acomp.valor_cancelamento = req.valor_cancelamento
            novo_acomp.aceitou_termos_cancelamento = req.aceitou_termos_cancelamento
            novo_acomp.usuario_aceitou_termos_cancelamento = req.usuario_aceitou_termos_cancelamento
            novo_acomp.justificativa_cancelamento = req.justificativa_cancelamento
            novo_acomp.motivos_cancelamento = req.motivos_cancelamento
            novo_acomp.cancelado_em = req.cancelado_em

            novo_acomp.save()

            agente_inst.acompanhamento = novo_acomp
            agente_inst.tipo_servico = agente_inst.tipo_servico or tipo_serv
            agente_inst.save()

            criados.append((novo_acomp, agente_inst))
            reqs_para_marcar.append(req.pk)

        if not criados:
            messages.warning(self.request, "Nenhum serviço foi confirmado no grupo.")
            return self.render_to_response(context)

        # marca apenas os serviços que você realmente criou
        RequisicaoSolicitacao.objects.filter(pk__in=reqs_para_marcar).update(solicitado=True)

        # caronas: aplica só no primeiro acompanhamento criado
        caronas = self.request.POST.getlist("carona_agente[]")
        if caronas:
            alvo_acomp = criados[0][0]
            principal = alvo_acomp.agentes.filter(tipo_agente="principal").first()
            _criar_caronas(alvo_acomp, principal, caronas)

            if len(criados) > 1:
                messages.warning(self.request, "Caronas aplicadas apenas no 1º serviço do grupo.")

        # sync supabase
        ok = 0
        for (acomp, _) in criados:
            success, mission_id, error = _sync_supabase_safe(self.request, acomp)
            if success:
                ok += 1
            else:
                messages.warning(self.request, f"Acompanhamento #{acomp.id} criado, mas falhou Supabase: {error}")

        messages.success(
            self.request,
            f"Grupo desmembrado em {len(criados)} acompanhamentos (1 por serviço). "
            f"Supabase ok: {ok}/{len(criados)}."
        )

        for (acomp, agente_inst) in criados:
            try:
                # Busca a requisição correspondente pelo id_externo do agente
                req_correspondente = RequisicaoSolicitacao.objects.filter(
                    id_externo=agente_inst.id_externo  # se guardou
                ).first()

                # Fallback: pega pelo tipo_servico
                if not req_correspondente:
                    for req in requisicoes:
                        if req.tipo_servico == agente_inst.tipo_servico:
                            req_correspondente = req
                            break

                if req_correspondente:
                    ag_principal = acomp.agentes.filter(
                        tipo_agente="principal"
                    ).select_related("agente").first()
                    gs_acionamento.notificar_acompanhamento_criado(req_correspondente, ag_principal)
            except Exception as e:
                logger.warning(f"Erro ao notificar GS para acomp #{acomp.id}: {e}")

        return redirect(self.success_url)

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

@login_required
@require_POST
def recusar_reuso(request, pk):
    req = get_object_or_404(RequisicaoSolicitacao, pk=pk, is_reuso=True, solicitado=False)

    req.solicitado = True
    req.save(update_fields=['solicitado'])

    if req.id_externo:
        try:
            gs_acionamento.notificar_status_requisicao(
                req,
                status_requisicao='reuso_recusado',
                agente_principal=None,
                extra_payload={
                    'reuso_recusado': True,
                    'motivo': 'Agente não pode atender. Escalar outro agente.',
                },
            )
        except Exception as e:
            logger.warning(f"Erro ao notificar GS sobre recusa de reuso: {e}")

    messages.info(
        request,
        f"Reuso recusado para req #{req.pk}. "
        f"Agente: {req.agente_nome_reuso or req.agente}. "
        f"O GS foi notificado para escalar outro agente."
    )
    return redirect('requisicao_solicitacao_list')

@login_required
def api_alertas_agendamento(request):
    agora = tz.localtime(tz.now())
    limite = agora + timedelta(hours=1)

    from .models import registroacompanhamentoagente, registroacompanhamento

    agentes_qs = (
        registroacompanhamentoagente.objects
        .filter(
            tipo_agente="principal",
            acompanhamento__status_acompanhamento__in=["missao_aceita", "agendada"],
            data_solicitada__isnull=False,
            horario_solicitado__isnull=False,
        )
        .select_related(
            "acompanhamento",
            "acompanhamento__cliente",
            "agente",
            "responsavel_agente",
        )
    )

    alertas = []
    for ag in agentes_qs:
        # Combinar data + hora
        dt_agendado = tz.make_aware(
            datetime.combine(ag.data_solicitada, ag.horario_solicitado)
        )

        # Só alertar se está entre agora e 1h no futuro
        # (não alertar missões que já passaram do horário)
        if dt_agendado < agora - timedelta(minutes=5):
            continue  # já passou há mais de 5min, ignorar

        if dt_agendado > limite:
            continue  # falta mais de 1h, ainda não é hora

        # Calcular minutos restantes
        diff = dt_agendado - agora
        minutos_restantes = max(0, int(diff.total_seconds() / 60))

        # Se já passou do horário mas está dentro da tolerância de 5min
        if diff.total_seconds() < 0:
            minutos_restantes = 0

        acomp = ag.acompanhamento
        alertas.append({
            "acompanhamento_id": acomp.pk,
            "cliente": acomp.cliente.nome if acomp.cliente else "-",
            "origem": acomp.origem or "-",
            "destino": acomp.destino or "-",
            "status": acomp.get_status_acompanhamento_display(),
            "agente_nome": ag.agente.nome if ag.agente else "-",
            "agente_placa": ag.placa_agente or "-",
            "responsavel": (
                ag.responsavel_agente.nome 
                if ag.responsavel_agente else "-"
            ),
            "data_agendada": ag.data_solicitada.strftime("%d/%m/%Y"),
            "horario_agendado": ag.horario_solicitado.strftime("%H:%M"),
            "minutos_restantes": minutos_restantes,
            "urgencia": (
                "critico" if minutos_restantes <= 15
                else "alerta" if minutos_restantes <= 30
                else "aviso"
            ),
            "supabase_mission_id": str(acomp.supabase_mission_id or ""),
            "link_mapa": (
                f"/acompanhamentos/missao/{acomp.supabase_mission_id}/mapa/"
                if acomp.supabase_mission_id
                else f"/acompanhamentos/missao/{acomp.pk}/mapa/"
            ),
        })

    # Ordenar: mais urgentes primeiro
    alertas.sort(key=lambda x: x["minutos_restantes"])

    return JsonResponse({
        "alertas": alertas,
        "total": len(alertas),
        "verificado_em": agora.strftime("%H:%M:%S"),
    })

@login_required
@permission_required("acompanhamentos.change_registroacompanhamento", raise_exception=True)
@require_POST
def atualizar_pedagio_acompanhamento(request, pk):
    acompanhamento = get_object_or_404(registroacompanhamento, pk=pk)

    raw = (request.POST.get("pedagio") or "").strip()

    if not raw:
        valor = Decimal("0.00")
    else:
        raw = raw.replace(".", "").replace(",", ".") if raw.count(",") == 1 else raw.replace(",", ".")
        try:
            valor = Decimal(raw)
        except (InvalidOperation, ValueError):
            return JsonResponse({"success": False, "error": "Valor de pedágio inválido."}, status=400)

        if valor < 0:
            return JsonResponse({"success": False, "error": "Pedágio não pode ser negativo."}, status=400)

    qs = (
        registroacompanhamentoagente.objects
        .filter(acompanhamento=acompanhamento)
        .exclude(tipo_agente="carona")
    )

    updated = qs.update(pedagio=valor)

    for ag in qs.select_related("franquia"):
            ag.recalcular_franquia_e_valores()
            ag.save(update_fields=["km_excedente", "horario_excedente", "valor_agente"])

    acompanhamento.refresh_from_db()
    
    if acompanhamento.status_acompanhamento == "verificado":
        acompanhamento.recalcular_financeiro()

    if acompanhamento.status_acompanhamento != "verificado":
        acompanhamento.status_acompanhamento = "verificado"
        acompanhamento.save(update_fields=["status_acompanhamento"])

        try:
            notificar_gs_verificado(acompanhamento.pk)
        except Exception as e:
            logger.warning(
                f"Erro ao notificar GS no verificado (acomp #{acompanhamento.pk}): {e}"
            )

    return JsonResponse({
        "success": True,
        "updated": updated,
        "pedagio": str(valor),
        "status_acompanhamento": acompanhamento.status_acompanhamento,
    })

def validar_acompanhamento(request, id):
    acompanhamento = get_object_or_404(registroacompanhamento, id=id)

    acompanhamento.validar_acompanhamento = True

    acompanhamento.nome_user = request.user.get_full_name() or request.user.username
    acompanhamento.save(update_fields=["validar_acompanhamento", "nome_user"])

    return redirect("acompanhamentosListFaturamento")