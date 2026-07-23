"""Regras de negócio do controle_acionamentos.

A CalculadoraValorAgente e os validadores de documentos são funções puras (sem
models/banco, testáveis sem DB); os serviços de orquestração (ex.:
vincular_franquia_em_lote) tocam a persistência por natureza.
"""

from dataclasses import dataclass
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation

from django.core.exceptions import FieldDoesNotExist, ValidationError
from django.db import transaction
from django.utils import timezone
from django.utils.text import capfirst


def _so_digitos(documento: str) -> list[int]:
    """Extrai apenas os dígitos de um documento, como lista de inteiros.

    Base comum dos validadores de CPF, CNPJ e CNH — ignora pontos,
    barra, hífen e espaços. O underscore inicial sinaliza "uso interno
    do módulo".
    """
    return [int(d) for d in documento if d.isdigit()]


def validar_cpf(cpf: str) -> bool:
    """Valida um CPF pelo algoritmo dos dígitos verificadores.

    Aceita CPF com ou sem máscara (pontos e hífen são ignorados).
    Retorna True se o CPF for válido, False caso contrário.
    """
    # Mantém só os dígitos, descartando pontos, hífen e espaços.
    numeros = _so_digitos(cpf)

    # CPF tem exatamente 11 dígitos.
    if len(numeros) != 11:
        return False

    # CPFs com todos os dígitos iguais (ex.: 111.111.111-11) são inválidos
    # por convenção (RN-01), embora passem no cálculo do dígito verificador.
    if len(set(numeros)) == 1:
        return False

    # Calcula um dígito verificador a partir dos dígitos anteriores.
    def calcular_digito(parciais: list[int]) -> int:
        # O peso começa em (quantidade de dígitos + 1) e decresce.
        peso = len(parciais) + 1
        soma = 0
        for numero in parciais:
            soma += numero * peso
            peso -= 1
        resto = (soma * 10) % 11
        # Resto 10 é tratado como 0 (regra do algoritmo).
        return resto if resto < 10 else 0

    # Primeiro dígito verificador: calculado sobre os 9 primeiros números.
    primeiro = calcular_digito(numeros[:9])
    # Segundo dígito verificador: calculado sobre os 10 primeiros números.
    segundo = calcular_digito(numeros[:10])

    # O CPF é válido se os dois dígitos calculados batem com os informados.
    return numeros[9] == primeiro and numeros[10] == segundo


def validar_cnpj(cnpj: str) -> bool:
    """Valida um CNPJ pelo algoritmo dos dígitos verificadores.

    Aceita CNPJ com ou sem máscara (pontos, barra e hífen são ignorados).
    Retorna True se o CNPJ for válido, False caso contrário.
    """
    # Mantém só os dígitos, descartando pontos, barra, hífen e espaços.
    numeros = _so_digitos(cnpj)

    # CNPJ tem exatamente 14 dígitos.
    if len(numeros) != 14:
        return False

    # CNPJs com todos os dígitos iguais (ex.: 00.000.000/0000-00) são inválidos
    # por convenção (RN-02), embora passem no cálculo do dígito verificador.
    if len(set(numeros)) == 1:
        return False

    # Calcula um dígito verificador a partir dos dígitos e seus pesos.
    # Diferente do CPF, os pesos do CNPJ não seguem uma sequência simples
    # (saltam de 2 para 9 no meio), então vêm de fora, explícitos.
    def calcular_digito(parciais: list[int], pesos: list[int]) -> int:
        soma = sum(numero * peso for numero, peso in zip(parciais, pesos))
        resto = (soma * 10) % 11
        # Resto 10 é tratado como 0 (mesma regra do validar_cpf).
        return resto if resto < 10 else 0

    # Pesos oficiais do CNPJ para cada dígito verificador.
    pesos_primeiro = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    pesos_segundo = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]

    # Primeiro dígito verificador: calculado sobre os 12 primeiros números.
    primeiro = calcular_digito(numeros[:12], pesos_primeiro)
    # Segundo dígito verificador: calculado sobre os 13 primeiros números.
    segundo = calcular_digito(numeros[:13], pesos_segundo)

    # O CNPJ é válido se os dois dígitos calculados batem com os informados.
    return numeros[12] == primeiro and numeros[13] == segundo

def validar_cnh(cnh: str) -> bool:
    """Valida uma CNH pelo algoritmo dos dígitos verificadores.

    Aceita CNH com ou sem máscara (espaços e pontuação são ignorados).
    Retorna True se a CNH for válida, False caso contrário.
    """
    numeros = _so_digitos(cnh)

    # CNH tem exatamente 11 dígitos.
    if len(numeros) != 11:
        return False

    # CNHs com todos os dígitos iguais são inválidas, embora passem na conta.
    if len(set(numeros)) == 1:
        return False

    # Primeiro dígito verificador: pesos 9..1 sobre os 9 primeiros números.
    soma = sum(numeros[i] * (9 - i) for i in range(9))
    dv1 = soma % 11
    # Pegadinha da CNH: se o cálculo "estoura" (>= 10), o dígito vira 0 e
    # marca um desconto que será aplicado no segundo dígito.
    desconto = 0
    if dv1 >= 10:
        dv1 = 0
        desconto = 2

    # Segundo dígito verificador: pesos 1..9 sobre os 9 primeiros números.
    # A ORDEM dos ajustes importa: subtrai o desconto -> corrige se ficar
    # negativo (+11) -> e só então trata o "estouro" (>= 10) como 0.
    soma = sum(numeros[i] * (i + 1) for i in range(9))
    dv2 = soma % 11
    dv2 = dv2 - desconto
    if dv2 < 0:
        dv2 += 11
    if dv2 >= 10:
        dv2 = 0

    # A CNH é válida se os dois dígitos calculados batem com os informados.
    return numeros[9] == dv1 and numeros[10] == dv2


# ---------------------------------------------------------------------------
# CalculadoraValorAgente — cálculo do valor pago ao agente (§8 e §11 do PRD)
# ---------------------------------------------------------------------------

# Dinheiro e horas são sempre apresentados com 2 casas decimais. Centralizamos
# o quantum num lugar só para a regra de arredondamento não se espalhar.
_DOIS_CASAS = Decimal("0.01")


def _quantizar(valor: Decimal) -> Decimal:
    """Arredonda um Decimal para 2 casas, meio-para-cima (ROUND_HALF_UP).

    É o arredondamento contábil esperado em valor monetário. O default do
    Python para Decimal é ROUND_HALF_EVEN ("banker's rounding"), que
    arredondaria 2,005 para 2,00 — não é o que queremos aqui.
    """
    return valor.quantize(_DOIS_CASAS, rounding=ROUND_HALF_UP)


@dataclass(frozen=True)
class EntradaCalculoAgente:
    """Entrada do cálculo do valor do agente — os 9 campos-fonte do contrato.

    Dado puro, sem Django/model: a calculadora recebe ``km_total`` e
    ``horas_total`` JÁ resolvidos (a derivação do §8.2 é feita por quem chama).
    Campos de exibição como ``nome_servico`` não entram aqui — não participam de
    nenhuma conta. ``frozen=True`` porque a entrada de um cálculo é imutável:
    montou, não muda mais.
    """

    valor_acionamento: Decimal
    franquia_km: int
    franquia_horas: Decimal
    valor_km_excedente: Decimal
    valor_hora_excedente: Decimal
    escalonamento_ativo: bool
    km_total: int
    horas_total: Decimal
    pedagio: Decimal


@dataclass(frozen=True)
class ResultadoCalculoAgente:
    """Resultado do cálculo — todos os campos derivados do contrato.

    Além do ``valor_agente`` final, expõe os passos intermediários do
    escalonamento (``blocos`` e os ``*_ajustada``) para a auditoria do PRD
    (§8.6/§8.7) poder conferir cada etapa, não só o número final. ``frozen=True``
    para o resultado de um cálculo não ser alterado depois de produzido.
    """

    blocos: int
    franquia_km_ajustada: int
    franquia_horas_ajustada: Decimal
    valor_acionamento_ajustado: Decimal
    km_excedente: int
    hora_excedente: Decimal
    valor_agente: Decimal


def calcular_valor_agente(entrada: EntradaCalculoAgente) -> ResultadoCalculoAgente:
    """Calcula o valor a pagar ao agente a partir dos campos-fonte (§8.4/§8.5).

    Escopo atual (Green do Cenário 1): SEM escalonamento — os valores ajustados
    são iguais aos de base e ``blocos == 0``. A lógica do §8.3 (escalonamento)
    entra nos Cenários 3 e 4 e só vai alterar o bloco de ajuste abaixo; a soma
    final do §8.5 já está na forma definitiva (usa ``valor_acionamento_ajustado``).
    """
    # Guard defensivo (cinto-e-suspensório do §11.1 C9): o escalonamento do §8.3
    # divide por franquia_km. O model FranquiaAgente já proíbe salvar
    # franquia_km=0 com escalonamento, mas a calculadora pura não pode confiar só
    # nisso — recusa explicitamente ANTES da divisão, com erro claro. ValueError
    # puro do Python (não ValidationError): este módulo não importa Django.
    if entrada.escalonamento_ativo and entrada.franquia_km == 0:
        raise ValueError(
            "Escalonamento ativo exige franquia_km maior que zero "
            "(divisão por zero no §8.3)."
        )

    # §8.3 — escalonamento: a cada 40 km acima da franquia base soma-se um
    # "bloco" à franquia (km e horas) e o valor é escalado na MESMA proporção do
    # km. A razão é calculada em Decimal/Decimal para nunca passar por float —
    # dinheiro não tolera o erro binário do float.
    if entrada.escalonamento_ativo and entrada.km_total > entrada.franquia_km:
        blocos = (entrada.km_total - entrada.franquia_km) // 40
        franquia_km_ajustada = entrada.franquia_km + blocos * 40
        franquia_horas_ajustada = entrada.franquia_horas + blocos
        razao = Decimal(franquia_km_ajustada) / Decimal(entrada.franquia_km)
        valor_acionamento_ajustado = _quantizar(entrada.valor_acionamento * razao)
    else:
        # Sem escalonamento (ou km dentro da franquia): o "ajustado" é o de base.
        blocos = 0
        franquia_km_ajustada = entrada.franquia_km
        franquia_horas_ajustada = entrada.franquia_horas
        valor_acionamento_ajustado = entrada.valor_acionamento

    # §8.4 — excedente é o que passou da franquia ajustada (nunca negativo).
    km_excedente = max(0, entrada.km_total - franquia_km_ajustada)
    hora_excedente = _quantizar(
        max(Decimal("0"), entrada.horas_total - franquia_horas_ajustada)
    )

    # §8.5 — valor final = base ajustada + excedentes às tarifas + pedágio.
    valor_agente = _quantizar(
        valor_acionamento_ajustado
        + (km_excedente * entrada.valor_km_excedente)
        + (hora_excedente * entrada.valor_hora_excedente)
        + entrada.pedagio
    )

    return ResultadoCalculoAgente(
        blocos=blocos,
        franquia_km_ajustada=franquia_km_ajustada,
        franquia_horas_ajustada=franquia_horas_ajustada,
        valor_acionamento_ajustado=valor_acionamento_ajustado,
        km_excedente=km_excedente,
        hora_excedente=hora_excedente,
        valor_agente=valor_agente,
    )


# ---------------------------------------------------------------------------
# calcular_valor_cliente — valor COBRADO DO CLIENTE (DD-068/ST1)
# ---------------------------------------------------------------------------


def calcular_valor_cliente(
    *,
    valor_acionamento: Decimal,
    franquia_km: int,
    franquia_horas: Decimal,
    valor_km_excedente: Decimal,
    valor_hora_excedente: Decimal,
    km_total: int,
    horas_total: Decimal,
) -> Decimal:
    """Valor cobrado do CLIENTE a partir dos valores do serviço (DD-068/ST1).

    Base do serviço mais os excedentes de km e de hora às respectivas tarifas.
    Diferente do valor do agente, aqui NÃO entram pedágio nem escalonamento — são
    exclusivos do cálculo do agente. Função pura (sem models); Decimal em tudo,
    arredondado a 2 casas pelo _quantizar da casa (ROUND_HALF_UP).
    """
    km_excedente = max(0, km_total - franquia_km)
    hora_excedente = _quantizar(max(Decimal("0"), horas_total - franquia_horas))
    return _quantizar(
        valor_acionamento
        + (km_excedente * valor_km_excedente)
        + (hora_excedente * valor_hora_excedente)
    )


# ---------------------------------------------------------------------------
# Regras de suspeita do saneamento de legados (DD-070 ST1)
# ---------------------------------------------------------------------------
# Funções PURAS (sem models): recebem os valores já extraídos do registro e
# devolvem uma lista de motivos de suspeita — lista vazia = registro limpo.
# O script de saneamento só RELATA; nenhuma função daqui altera dado algum.

# Teto acordado na ST1 da DD-070 — borda INCLUSIVA: exatamente 120 km/h é limpo.
VELOCIDADE_MAXIMA_KMH = Decimal("120")

# Faixas informadas pelos operadores do administrativo em 22/07/2026 —
# bordas INCLUSIVAS nos dois extremos (piso e teto ainda são valores válidos).
FAIXA_TARIFA_KM = (Decimal("1.00"), Decimal("6.00"))
FAIXA_TARIFA_HORA = (Decimal("5.00"), Decimal("60.00"))
FAIXA_VALOR_BASE = (Decimal("200.00"), Decimal("20000.00"))


def avaliar_suspeita_velocidade(
    km_inicial: Decimal, km_final: Decimal, minutos: int | None
) -> list[str]:
    """Suspeita por velocidade média implausível (saneamento DD-070).

    Apenas RELATA motivos de suspeita — nunca altera o registro. Lista
    vazia significa registro limpo.
    """
    motivos: list[str] = []
    km_percorrido = km_final - km_inicial

    # Hodômetro andando para trás: registro inconsistente por si só —
    # não faz sentido calcular velocidade sobre um percurso negativo.
    if km_percorrido < 0:
        motivos.append(
            f"km final ({km_final}) menor que km inicial ({km_inicial})"
        )
        return motivos

    # Houve percurso, mas não há tempo apurável (None ou <= 0): impossível
    # ter velocidade — também não calcula.
    if km_percorrido > 0 and (minutos is None or minutos <= 0):
        motivos.append(
            f"tempo impossível ({minutos} min) para {km_percorrido} km percorridos"
        )
        return motivos

    # Percurso zero com tempo válido é limpo; com percurso, a velocidade
    # média é km / (min/60), tudo em Decimal — nunca float (§18 do CLAUDE.md).
    if km_percorrido > 0:
        velocidade_kmh = km_percorrido / (Decimal(minutos) / Decimal("60"))
        if velocidade_kmh > VELOCIDADE_MAXIMA_KMH:
            motivos.append(
                f"velocidade média de {_quantizar(velocidade_kmh)} km/h "
                f"acima do teto de {VELOCIDADE_MAXIMA_KMH} km/h"
            )

    return motivos


def avaliar_suspeita_faixas(
    valor_base: Decimal, tarifa_km: Decimal, tarifa_hora: Decimal
) -> list[str]:
    """Suspeita por valor fora da faixa operacional (saneamento DD-070).

    Apenas RELATA motivos de suspeita — nunca altera o registro. Cada campo
    fora da sua faixa (bordas inclusivas) gera um motivo independente; lista
    vazia significa registro limpo.
    """
    motivos: list[str] = []

    piso, teto = FAIXA_VALOR_BASE
    if not piso <= valor_base <= teto:
        motivos.append(
            f"valor base ({valor_base}) fora da faixa de {piso} a {teto}"
        )

    piso, teto = FAIXA_TARIFA_KM
    if not piso <= tarifa_km <= teto:
        motivos.append(
            f"tarifa de km ({tarifa_km}) fora da faixa de {piso} a {teto}"
        )

    piso, teto = FAIXA_TARIFA_HORA
    if not piso <= tarifa_hora <= teto:
        motivos.append(
            f"tarifa de hora ({tarifa_hora}) fora da faixa de {piso} a {teto}"
        )

    return motivos


def avaliar_suspeita_valores_identicos(
    valor_base: Decimal, tarifa_km: Decimal, tarifa_hora: Decimal
) -> list[str]:
    """Suspeita por valores idênticos entre campos (saneamento DD-070).

    Apenas RELATA motivos de suspeita — nunca altera o registro. Um par só é
    suspeito se os dois valores forem iguais E ambos maiores que zero (zero a
    zero não é digitação repetida); lista vazia significa registro limpo.
    """
    motivos: list[str] = []

    if valor_base == tarifa_km and valor_base > 0:
        motivos.append("valor base e tarifa de km idênticos")

    if valor_base == tarifa_hora and valor_base > 0:
        motivos.append("valor base e tarifa de hora idênticos")

    if tarifa_km == tarifa_hora and tarifa_km > 0:
        motivos.append("tarifa de km e tarifa de hora idênticas")

    return motivos


# ---------------------------------------------------------------------------
# calcular_valor_agente_por_franquia — valor do AGENTE pela franquia (DD-068/ST2)
# ---------------------------------------------------------------------------


def _entrada_da_franquia(
    franquia_agente,
    km_total: int,
    horas_total: Decimal,
    pedagio: Decimal,
) -> EntradaCalculoAgente:
    """Monta a EntradaCalculoAgente a partir dos valores da FRANQUIA (DD-068).

    Fonte ÚNICA do mapeamento franquia → entrada da calculadora: valores-base,
    tarifas e escalonamento vêm todos da franquia (única tradução de nome:
    ``escalonamento_automatico`` → ``escalonamento_ativo``); km/horas totais e
    pedágio vêm da jornada. Compartilhado por ``calcular_valor_agente_por_franquia``,
    ``recalcular_valor_agente`` e ``compor_valor_agente``, para que todos leiam a
    MESMA fonte e nunca divirjam. ``franquia_agente`` é duck-typed (qualquer
    objeto com os 6 atributos) — o módulo segue sem importar models.
    """
    return EntradaCalculoAgente(
        valor_acionamento=franquia_agente.valor_acionamento,
        franquia_km=franquia_agente.franquia_km,
        franquia_horas=franquia_agente.franquia_horas,
        valor_km_excedente=franquia_agente.valor_km_excedente,
        valor_hora_excedente=franquia_agente.valor_hora_excedente,
        escalonamento_ativo=franquia_agente.escalonamento_automatico,
        km_total=km_total,
        horas_total=horas_total,
        pedagio=pedagio,
    )


def calcular_valor_agente_por_franquia(
    *,
    franquia_agente,
    km_total: int,
    horas_total: Decimal,
    pedagio: Decimal,
):
    """Resolve o valor a pagar ao AGENTE pela franquia vinculada (DD-068/ST2).

    Sem franquia (None) não há o que calcular: retorna None — estado PENDENTE
    (o vínculo pode vir depois, inclusive em lote). Com franquia, a entrada é
    montada EXCLUSIVAMENTE com os valores dela (nunca os inline do acionamento)
    pelo ``_entrada_da_franquia``, e a matemática é toda delegada ao
    calcular_valor_agente — escalonamento e pedágio já moram lá, nada é
    duplicado aqui. Função pura, sem importar models.
    """
    if franquia_agente is None:
        return None

    entrada = _entrada_da_franquia(
        franquia_agente,
        km_total=km_total,
        horas_total=horas_total,
        pedagio=pedagio,
    )
    return calcular_valor_agente(entrada).valor_agente


def recalcular_valor_agente(acionamento) -> None:
    """Preenche os campos calculados de um Acionamento (RN-07, §8 + DD-068/ST3).

    Ponte entre o model e a calculadora pura: deriva km/horas totais (§8.2 —
    fatos da jornada, sempre) e resolve o trio financeiro pela FRANQUIA
    vinculada, ÚNICA fonte do valor do agente desde a DD-068/ST3 (o RN-08 e o
    cálculo pelo inline morreram): sem franquia, km_excedente/hora_excedente/
    valor_agente ficam PENDENTES (None) — nunca calculados pelo inline, nunca
    zero. Lê a instância por duck-typing (NÃO importa models) — este módulo
    segue puro e testável sem banco. Não persiste: quem chama
    (``Acionamento.save``) é que faz o ``super().save()``.
    """
    # §8.2 — derivação dos totais. Tudo em Decimal puro (sem float): o timedelta
    # já vem decomposto em days/seconds/microseconds, que montamos em segundos e
    # convertemos para horas com o mesmo arredondamento contábil dos valores.
    acionamento.km_total = acionamento.km_final - acionamento.km_inicio
    delta = acionamento.data_hora_final - acionamento.data_hora_inicio
    segundos = (
        Decimal(delta.days) * 86400
        + Decimal(delta.seconds)
        + Decimal(delta.microseconds) / Decimal(1_000_000)
    )
    acionamento.horas_total = _quantizar(segundos / Decimal(3600))

    # DD-068/ST3 — sem franquia vinculada a conta do agente não roda: o trio
    # (saídas da MESMA conta) fica pendente em bloco. Os totais acima permanecem.
    if not acionamento.franquia_agente:
        acionamento.km_excedente = None
        acionamento.hora_excedente = None
        acionamento.valor_agente = None
        return

    entrada = _entrada_da_franquia(
        acionamento.franquia_agente,
        km_total=acionamento.km_total,
        horas_total=acionamento.horas_total,
        pedagio=acionamento.pedagio,
    )
    resultado = calcular_valor_agente(entrada)

    # km_total/horas_total já vieram do passo §8.2 — aqui só os derivados da conta.
    acionamento.km_excedente = resultado.km_excedente
    acionamento.hora_excedente = resultado.hora_excedente
    acionamento.valor_agente = resultado.valor_agente


def _entrada_do_snapshot(acionamento) -> dict:
    """Monta os kwargs do calcular_valor_cliente a partir do SNAPSHOT INLINE do
    acionamento (DD-070/ST4) — espelho do _entrada_da_franquia no lado do
    cliente. Fonte ÚNICA do mapeamento snapshot → entrada, compartilhada por
    ``recalcular_valor_cliente`` e ``compor_valor_cliente`` para nunca
    divergirem. Requer km_total/horas_total já derivados. Duck-typing — o
    módulo segue sem importar models.
    """
    return dict(
        valor_acionamento=acionamento.valor_acionamento,
        franquia_km=acionamento.franquia_km,
        franquia_horas=acionamento.franquia_horas,
        valor_km_excedente=acionamento.valor_km_excedente,
        valor_hora_excedente=acionamento.valor_hora_excedente,
        km_total=acionamento.km_total,
        horas_total=acionamento.horas_total,
    )


def recalcular_valor_cliente(acionamento) -> None:
    """Preenche ``acionamento.valor_cliente`` (DD-070/ST4, §8.9 do PRD v1.5).

    Ponte entre o model e a calculadora pura, espelho do contrato do
    recalcular_valor_agente: a fonte é o SNAPSHOT INLINE do serviço gravado no
    acionamento (nunca a franquia — lado agente), sobre os km_total/horas_total
    JÁ derivados pelo recalcular_valor_agente no mesmo save (por isso ela roda
    depois dele). A matemática mora toda no calcular_valor_cliente (fonte
    única, DD-068/ST1) — sem pedágio e sem escalonamento, por regra de negócio.
    Não persiste: quem chama (``Acionamento.save``) faz o ``super().save()``.
    """
    acionamento.valor_cliente = calcular_valor_cliente(
        **_entrada_do_snapshot(acionamento)
    )


def aplicar_servico_ao_acionamento(acionamento, servico):
    """DD-067/ST1 — aplica um ServicoCliente a um Acionamento, em momento
    EXPLÍCITO (nunca no save()).

    Congelamento forte: a FK `servico_cliente` é só a REFERÊNCIA de qual serviço
    foi escolhido; o SNAPSHOT copia os 5 valores do serviço para os campos
    inline do acionamento. Desde a DD-068/ST3 o inline NÃO alimenta mais o
    valor do AGENTE (que vem só da franquia) — o snapshot permanece como
    registro congelado dos valores do serviço (auditoria e base do valor do
    CLIENTE, DD-068/ST1): re-salvar o acionamento nunca relê o catálogo, e
    editar o catálogo depois não altera acionamentos já registrados.

    O `nome_servico` do acionamento também DERIVA do serviço (DD-067/ST2): passa a
    ser o label humano do choice escolhido (get_nome_display), não mais um texto
    digitado — o catálogo é a fonte única do nome exibido.

    NÃO salva: quem chama decide o momento do save (padrão commit=False das views).
    """
    acionamento.servico_cliente = servico
    acionamento.nome_servico = servico.get_nome_display()
    acionamento.valor_acionamento = servico.valor_acionamento
    acionamento.franquia_km = servico.franquia_km
    acionamento.franquia_horas = servico.franquia_horas
    acionamento.valor_km_excedente = servico.valor_km_excedente
    acionamento.valor_hora_excedente = servico.valor_hora_excedente


@dataclass(frozen=True)
class ComposicaoValorAgente:
    """Extrato de parcelas do valor do agente, para exibição no detalhe (DD-032/ST5).

    Recompõe as parcelas que formam o total, a partir da FRANQUIA vinculada
    (única fonte do valor do agente desde a DD-068/ST3): a base
    pós-escalonamento, os subtotais de excedente às tarifas da franquia e o
    pedágio. ``fonte_franquia`` (sempre True quando o extrato existe) e
    ``blocos`` alimentam a anotação da 1ª linha ("da franquia" / "escalonado ·
    N blocos"). ``frozen=True`` — extrato produzido não muda.

    Invariante: ``valor_acionamento_ajustado + subtotal_km + subtotal_hora +
    pedagio == valor_agente``.
    """

    valor_acionamento_ajustado: Decimal
    blocos: int
    fonte_franquia: bool
    km_excedente: int
    valor_unitario_km: Decimal
    subtotal_km: Decimal
    hora_excedente: Decimal
    valor_unitario_hora: Decimal
    subtotal_hora: Decimal
    pedagio: Decimal
    valor_agente: Decimal


def compor_valor_agente(acionamento):
    """Monta o extrato de parcelas do detalhe (DD-032/ST5) — exibição PURA.

    DD-068/ST3: SEM franquia vinculada não há composição — retorna None (estado
    PENDENTE; o card do detalhe não renderiza). Com franquia, nada é persistido:
    monta a entrada pelo mesmo ``_entrada_da_franquia`` do recálculo, roda a
    calculadora e recompõe as parcelas em R$ (base ajustada + subtotais de
    excedente às tarifas da FRANQUIA + pedágio). Os subtotais usam o mesmo
    ``_quantizar`` do módulo, para o arredondamento contábil bater com o do
    cálculo. Invariante (garantido pelos testes): a soma das parcelas ==
    ``valor_agente``. Requer km_total/horas_total já derivados (acionamento salvo).
    """
    if not acionamento.franquia_agente:
        return None

    entrada = _entrada_da_franquia(
        acionamento.franquia_agente,
        km_total=acionamento.km_total,
        horas_total=acionamento.horas_total,
        pedagio=acionamento.pedagio,
    )
    resultado = calcular_valor_agente(entrada)

    # Subtotais em R$: quantidade de excedente × tarifa da franquia.
    subtotal_km = _quantizar(resultado.km_excedente * entrada.valor_km_excedente)
    subtotal_hora = _quantizar(resultado.hora_excedente * entrada.valor_hora_excedente)

    return ComposicaoValorAgente(
        valor_acionamento_ajustado=resultado.valor_acionamento_ajustado,
        blocos=resultado.blocos,
        fonte_franquia=True,
        km_excedente=resultado.km_excedente,
        valor_unitario_km=entrada.valor_km_excedente,
        subtotal_km=subtotal_km,
        hora_excedente=resultado.hora_excedente,
        valor_unitario_hora=entrada.valor_hora_excedente,
        subtotal_hora=subtotal_hora,
        pedagio=entrada.pedagio,
        valor_agente=resultado.valor_agente,
    )


@dataclass(frozen=True)
class ComposicaoValorCliente:
    """Extrato de parcelas do valor do CLIENTE, para o detalhe (DD-069/ST1).

    Espelho do ComposicaoValorAgente, mas com a fonte nos valores do SERVIÇO
    gravados no acionamento (inline/snapshot) — nunca a franquia, que é lado
    agente. Por regra de negócio (DD-068/ST1) o cliente NÃO paga pedágio nem
    escalonamento, então essas parcelas simplesmente não existem aqui.

    Invariante: ``valor_acionamento + subtotal_km + subtotal_hora ==
    valor_cliente``.
    """

    valor_acionamento: Decimal
    km_excedente: int
    valor_unitario_km: Decimal
    subtotal_km: Decimal
    hora_excedente: Decimal
    valor_unitario_hora: Decimal
    subtotal_hora: Decimal
    valor_cliente: Decimal


def compor_valor_cliente(acionamento):
    """Monta o extrato de parcelas do valor do CLIENTE (DD-069/ST1) — exibição
    PURA, espelho de compor_valor_agente.

    Não depende de franquia (existe mesmo no estado pendente do agente): a
    entrada são os campos de serviço persistidos no acionamento. O total é
    delegado ao calcular_valor_cliente (fonte única da matemática, DD-068/ST1);
    aqui só se recompõem as quantidades e subtotais com o mesmo ``_quantizar``
    da casa. Requer km_total/horas_total já derivados (acionamento salvo).
    """
    km_excedente = max(0, acionamento.km_total - acionamento.franquia_km)
    hora_excedente = _quantizar(
        max(Decimal("0"), acionamento.horas_total - acionamento.franquia_horas)
    )
    subtotal_km = _quantizar(km_excedente * acionamento.valor_km_excedente)
    subtotal_hora = _quantizar(hora_excedente * acionamento.valor_hora_excedente)

    valor_cliente = calcular_valor_cliente(**_entrada_do_snapshot(acionamento))

    return ComposicaoValorCliente(
        valor_acionamento=acionamento.valor_acionamento,
        km_excedente=km_excedente,
        valor_unitario_km=acionamento.valor_km_excedente,
        subtotal_km=subtotal_km,
        hora_excedente=hora_excedente,
        valor_unitario_hora=acionamento.valor_hora_excedente,
        subtotal_hora=subtotal_hora,
        valor_cliente=valor_cliente,
    )


def acionamentos_em_conflito_de_franquia(pks, franquia):
    """DD-051/ST2 (AC-06.5) — FONTE ÚNICA da regra de conflito de sobrescrita:
    acionamentos do lote (`pks`) que já têm franquia vinculada E DIFERENTE da
    `franquia` selecionada. Franquia IDÊNTICA não é conflito (re-vincular a mesma
    recalcula igual), por isso o .exclude(franquia_agente=franquia).

    Retorna a queryset (lazy): o service usa .exists() para recusar o lote; a view
    encadeia select_related para listar os itens na página de confirmação. A regra
    mora aqui, num lugar só, para service e view nunca divergirem.

    Import local de Acionamento: evita import circular (models importa deste módulo).
    """
    from controle_acionamentos.models import Acionamento

    return (
        Acionamento.objects.filter(pk__in=pks, franquia_agente__isnull=False)
        .exclude(franquia_agente=franquia)
    )


def vincular_franquia_em_lote(pks, franquia, sobrescrever=False, *, editado_por):
    """DD-015/M4 (AC-06.4) — vincula `franquia` a todos os acionamentos de
    `pks` e recalcula os campos derivados de cada um, em transação atômica:
    falha em qualquer item desfaz o lote inteiro (AC-06.6). Retorna a
    contagem de acionamentos atualizados.

    `sobrescrever` (default False = caminho seguro): com False, se algum item do
    lote já tiver franquia DIFERENTE da selecionada, o lote é recusado (AC-06.5,
    DD-051/ST2) — franquia IDÊNTICA não é conflito e passa normalmente; com True,
    os itens com outra franquia são re-vinculados e recalculados.

    `editado_por` (DD-070/ST3) — autor da trilha de auditoria, keyword-only e
    SEM default de propósito: não pode existir caminho de gravação em lote sem
    autor (um default que pulasse o registro recriaria o furo que este ciclo
    corrige). Para cada item, a foto do estado persistido é tirada ANTES da
    mutação (o mesmo padrão antigo/novo da view acionamento_update) e, após o
    save, registrar_edicao_acionamento grava as linhas de histórico das
    mudanças (franquia_agente, valor_agente e demais CAMPOS_AUDITADOS que
    mudarem). O registro fica DENTRO do atomic de propósito: se a trilha
    falhar, o lote inteiro desfaz — dado e auditoria nunca divergem.

    Import local de Acionamento: evita import circular (models importa
    recalcular_valor_agente deste módulo).
    """
    from controle_acionamentos.models import Acionamento

    with transaction.atomic():
        if not sobrescrever:
            if acionamentos_em_conflito_de_franquia(pks, franquia).exists():
                raise ValidationError(
                    "Há acionamentos com OUTRA franquia já vinculada no lote; "
                    "a sobrescrita exige confirmação explícita (AC-06.5)."
                )

        atualizados = 0
        for acionamento in Acionamento.objects.filter(pk__in=pks):
            # Foto do estado persistido ANTES da mutação (padrão antigo/novo
            # da view acionamento_update) — é o "antes" das linhas de trilha.
            antigo = Acionamento.objects.get(pk=acionamento.pk)
            acionamento.franquia_agente = franquia
            acionamento.full_clean()
            acionamento.save()
            registrar_edicao_acionamento(antigo, acionamento, editado_por)
            atualizados += 1
        return atualizados


def sincronizar_catalogo_do_cliente(cliente, formset):
    """DD-066/ST2 — persiste o catálogo de serviços de um cliente a partir do
    formset já validado. Retorna a quantidade de registros gravados.

      * linha vazia SEM registro existente → ignora;
      * linha preenchida → cria ou atualiza o ServicoCliente (cliente, nome),
        inclusive o campo `ativo`;
      * NUNCA deleta registro (o catálogo desativa, não remove);
      * roda full_clean() de cada registro antes do save.

    A linha vazia cujo serviço JÁ TEM registro nem chega aqui: o clean do formset
    (BaseCatalogoServicosFormSet) já a barra. Import local de ServicoCliente para
    não acoplar o módulo de services ao de models no topo (mesmo padrão do
    vincular_franquia_em_lote).
    """
    from controle_acionamentos.models import ServicoCliente

    existentes = {s.nome: s for s in ServicoCliente.objects.filter(cliente=cliente)}
    persistidos = 0
    for form in formset.forms:
        if form.linha_vazia:
            continue
        nome = form.cleaned_data["nome"]
        servico = existentes.get(nome) or ServicoCliente(cliente=cliente, nome=nome)
        servico.ativo = bool(form.cleaned_data.get("ativo"))
        servico.valor_acionamento = form.cleaned_data.get("valor_acionamento")
        servico.franquia_km = form.cleaned_data.get("franquia_km")
        servico.franquia_horas = form.cleaned_data.get("franquia_horas")
        servico.valor_km_excedente = form.cleaned_data.get("valor_km_excedente")
        servico.valor_hora_excedente = form.cleaned_data.get("valor_hora_excedente")
        servico.full_clean()
        servico.save()
        persistidos += 1
    return persistidos


# Caminho B — o que a trilha de edição (DD-049) audita: os campos EDITÁVEIS do
# AcionamentoForm + os CALCULADOS FINANCEIROS do model (excedentes, valor_agente
# e valor_cliente), para a auditoria responder "o que — e quanto de dinheiro —
# mudou". Os totais km_total/horas_total ficam de fora: são distância/tempo,
# não valor financeiro.
CAMPOS_AUDITADOS = [
    # editáveis do AcionamentoForm
    "cliente",
    "nome_servico",
    "valor_acionamento",
    "franquia_km",
    "franquia_horas",
    "valor_km_excedente",
    "valor_hora_excedente",
    "origem",
    "destino",
    "responsavel_agente",
    "agente",
    "placa_agente",
    "motorista",
    "placa_motorista",
    "numero_motorista",
    "data_hora_solicitado",
    "data_hora_inicio",
    "data_hora_final",
    "km_inicio",
    "km_final",
    "pedagio",
    "franquia_agente",
    # calculados financeiros persistidos no model
    "km_excedente",
    "hora_excedente",
    "valor_agente",
    "valor_cliente",
]


def _serializar_valor(valor):
    """None -> "" ; senão str(valor). Datetimes aware são normalizados ao fuso
    local antes do str() — o mesmo instante lido do banco (UTC) e parseado do
    form (fuso local) deve serializar idêntico (e a trilha fica legível em
    horário local). Decimal segue como string (sem float)."""
    if valor is None:
        return ""
    if isinstance(valor, datetime) and timezone.is_aware(valor):
        valor = timezone.localtime(valor)
    return str(valor)


def registrar_edicao_acionamento(antigo, novo, editado_por):
    """Grava a trilha de auditoria da edição de um Acionamento (DD-049/ST3).

    `antigo` = foto da instância recarregada do banco ANTES do save; `novo` = a
    instância já salva com as mudanças. Compara CAMPOS_AUDITADOS campo a campo e
    cria um AcionamentoHistorico por campo cujo valor SERIALIZADO tenha mudado.
    Retorna a lista dos registros criados (vazia se nada mudou).

    Orquestração pura: SEM transaction.atomic aqui — a atomicidade (save do
    acionamento + trilha num único bloco) é responsabilidade do chamador (a view
    acionamento_update, na etapa de integração).

    Import local de AcionamentoHistorico: evita import circular (models importa
    deste módulo).
    """
    from controle_acionamentos.models import AcionamentoHistorico

    registros = []
    for nome in CAMPOS_AUDITADOS:
        # attname resolve o atributo cru: FK -> "<nome>_id" (compara valor sem
        # query), demais -> o próprio nome. O `campo` gravado é sempre `nome`.
        attname = novo._meta.get_field(nome).attname
        anterior = _serializar_valor(getattr(antigo, attname))
        atual = _serializar_valor(getattr(novo, attname))
        if anterior != atual:
            registros.append(
                AcionamentoHistorico.objects.create(
                    acionamento=novo,
                    editado_por=editado_por,
                    campo=nome,
                    valor_anterior=anterior,
                    valor_novo=atual,
                )
            )
    return registros


# --- Backlog DD-070 1c: formatação de APRESENTAÇÃO da trilha ---
# O texto gravado em AcionamentoHistorico não muda (histórico imutável); estas
# funções só traduzem o texto cru para exibição no detalhe. Grupos de campos
# por tipo de formatação (subconjuntos de CAMPOS_AUDITADOS):
_CAMPOS_TRILHA_MOEDA = {
    "valor_cliente",
    "valor_agente",
    "pedagio",
    "valor_acionamento",
    "valor_km_excedente",
    "valor_hora_excedente",
}
_CAMPOS_TRILHA_DATA = {"data_hora_solicitado", "data_hora_inicio", "data_hora_final"}
_CAMPOS_TRILHA_KM = {"km_inicio", "km_final", "km_excedente", "franquia_km"}
_CAMPOS_TRILHA_HORAS = {"hora_excedente", "franquia_horas"}


def rotulo_campo_trilha(campo):
    """Rótulo humano de um campo da trilha: o verbose_name do campo no model
    Acionamento, com inicial maiúscula.

    Campo que não existe no model (histórico legado, dado sintético) volta CRU,
    inalterado — a trilha nunca pode quebrar por causa de um rótulo.

    Import local do model: services não importa models no topo (evita ciclo —
    models importa deste módulo — e mantém as funções puras testáveis sem DB).
    """
    from controle_acionamentos.models import Acionamento

    try:
        field = Acionamento._meta.get_field(campo)
    except FieldDoesNotExist:
        return campo
    return capfirst(field.verbose_name)


def _moeda_brasileira(valor: Decimal) -> str:
    """Máscara monetária brasileira sem depender de locale do sistema: quantiza
    a 2 casas e troca os separadores do format spec dos EUA (1,943.00) pelos
    nossos (1.943,00) via caractere sentinela."""
    quantizado = valor.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{quantizado:,.2f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def formatar_valor_trilha(campo, texto, nomes_franquias):
    """Valor de apresentação de uma linha da trilha, formatado pelo TIPO do
    campo (moeda R$, data d/m/Y H:i, km, horas, nome de franquia via mapa
    pk -> nome vindo do selector — o service segue puro, sem query).

    100% defensiva: string vazia volta vazia e QUALQUER parse que falhar
    devolve o texto cru — a função jamais levanta exceção, porque apresentação
    nunca pode derrubar o detalhe (a trilha é histórico imutável, inclusive
    com texto legado fora do formato atual).
    """
    if not texto:
        return texto
    try:
        if campo in _CAMPOS_TRILHA_MOEDA:
            return "R$ " + _moeda_brasileira(Decimal(texto))
        if campo in _CAMPOS_TRILHA_DATA:
            return datetime.fromisoformat(texto).strftime("%d/%m/%Y %H:%M")
        if campo in _CAMPOS_TRILHA_KM:
            return f"{int(texto)} km"
        if campo in _CAMPOS_TRILHA_HORAS:
            quantizado = Decimal(texto).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            return f"{quantizado:.2f}".replace(".", ",") + " h"
        if campo == "franquia_agente":
            return nomes_franquias.get(int(texto), texto)
    except (InvalidOperation, ValueError, TypeError):
        return texto
    return texto


# --- Backlog pós-DD-070 item 3: exportação Excel dos pagamentos de agentes ---
# Layout fixo da planilha de pagamentos (30 colunas, ordem do contrato de
# testes). SQ, EVENTO, H. CHEGADA e OBS são preenchimento manual da operação.
COLUNAS_EXPORTACAO_PAGAMENTOS = [
    "DATA", "SQ", "EVENTO", "CLIENTE", "ORIGEM", "DESTINO", "AGENTE",
    "PLACA AGENTE", "MOTORISTA", "PLACA MOT.", "H. SOLIC", "H. CHEGADA",
    "H. INICIAL", "H. FINAL", "KM INICIAL", "KM FINAL", "KM TOTAL",
    "FRANQUIA DE HORAS", "H. TRABALHADA", "H. EXTRA", "VALOR DA HORA EXTRA",
    "VALOR H.E", "FRANQUIA DE KM", "KM EXCEDENTE", "VLR KM EXTRA",
    "VALOR K.H", "VALOR DA DIÁRIA", "PEDÁGIO", "VALOR TOTAL", "OBS",
]


def _datetime_local_naive(valor):
    """openpyxl não aceita datetime aware: normaliza ao fuso local e tira o
    tzinfo — a célula fica com o horário que a operação enxerga."""
    return timezone.localtime(valor).replace(tzinfo=None)


def _linha_exportacao(acionamento, compor):
    """Uma linha da planilha, na ordem de COLUNAS_EXPORTACAO_PAGAMENTOS.

    O lado do AGENTE (parâmetros da franquia, excedentes, parcelas da
    composição e VALOR TOTAL) só é preenchido com franquia vinculada; sem ela
    o acionamento está PENDENTE (DD-068) e essas células saem None. PEDÁGIO é
    sempre do registro. Colunas manuais (SQ, EVENTO, H. CHEGADA, OBS) = None.
    """
    franquia = acionamento.franquia_agente
    if franquia:
        composicao = compor(acionamento)
        franquia_horas = franquia.franquia_horas
        hora_extra = acionamento.hora_excedente
        valor_hora_extra = franquia.valor_hora_excedente
        valor_he = composicao.subtotal_hora
        franquia_km = franquia.franquia_km
        km_excedente = acionamento.km_excedente
        vlr_km_extra = franquia.valor_km_excedente
        valor_kh = composicao.subtotal_km
        valor_diaria = composicao.valor_acionamento_ajustado
        valor_total = acionamento.valor_agente
    else:
        franquia_horas = hora_extra = valor_hora_extra = valor_he = None
        franquia_km = km_excedente = vlr_km_extra = valor_kh = None
        valor_diaria = valor_total = None
    return [
        _datetime_local_naive(acionamento.data_hora_solicitado).date(),  # DATA
        None,                                     # SQ (manual)
        None,                                     # EVENTO (manual)
        str(acionamento.cliente),                 # CLIENTE
        acionamento.origem,                       # ORIGEM
        acionamento.destino,                      # DESTINO
        str(acionamento.agente),                  # AGENTE
        acionamento.placa_agente,                 # PLACA AGENTE
        acionamento.motorista,                    # MOTORISTA
        acionamento.placa_motorista,              # PLACA MOT.
        _datetime_local_naive(acionamento.data_hora_solicitado),  # H. SOLIC
        None,                                     # H. CHEGADA (manual)
        _datetime_local_naive(acionamento.data_hora_inicio),      # H. INICIAL
        _datetime_local_naive(acionamento.data_hora_final),       # H. FINAL
        acionamento.km_inicio,                    # KM INICIAL
        acionamento.km_final,                     # KM FINAL
        acionamento.km_total,                     # KM TOTAL
        franquia_horas,                           # FRANQUIA DE HORAS
        acionamento.horas_total,                  # H. TRABALHADA
        hora_extra,                               # H. EXTRA
        valor_hora_extra,                         # VALOR DA HORA EXTRA
        valor_he,                                 # VALOR H.E
        franquia_km,                              # FRANQUIA DE KM
        km_excedente,                             # KM EXCEDENTE
        vlr_km_extra,                             # VLR KM EXTRA
        valor_kh,                                 # VALOR K.H
        valor_diaria,                             # VALOR DA DIÁRIA
        acionamento.pedagio,                      # PEDÁGIO
        valor_total,                              # VALOR TOTAL
        None,                                     # OBS (manual)
    ]


# Acabamento visual do arquivo (só estilo — o conteúdo das células não muda).
# Cores/formatos seguem a planilha de referência da operação.
_COR_CABECALHO_EXPORTACAO = "FFFFCE29"  # aRGB: FF de alpha + FFCE29
_FORMATO_DATA_EXPORTACAO = "DD/MM/YYYY"
_FORMATO_DATA_HORA_EXPORTACAO = "DD/MM/YYYY HH:MM"
_FORMATO_MOEDA_EXPORTACAO = '"R$" #,##0.00'
_FORMATO_DECIMAL_EXPORTACAO = "0.00"

# number_format por título de coluna (coluna fora do dict fica sem formato).
_FORMATOS_EXPORTACAO = {
    "DATA": _FORMATO_DATA_EXPORTACAO,
    **{t: _FORMATO_DATA_HORA_EXPORTACAO
       for t in ("H. SOLIC", "H. INICIAL", "H. FINAL")},
    **{t: _FORMATO_MOEDA_EXPORTACAO
       for t in ("VALOR DA HORA EXTRA", "VALOR H.E", "VLR KM EXTRA",
                 "VALOR K.H", "VALOR DA DIÁRIA", "PEDÁGIO", "VALOR TOTAL")},
    **{t: _FORMATO_DECIMAL_EXPORTACAO
       for t in ("H. TRABALHADA", "H. EXTRA", "FRANQUIA DE HORAS")},
}

# Largura por título de coluna (unidade do Excel; default 13 corta datetime).
_LARGURAS_EXPORTACAO = {
    "DATA": 12, "SQ": 8, "EVENTO": 14, "CLIENTE": 22, "ORIGEM": 25,
    "DESTINO": 25, "AGENTE": 22, "PLACA AGENTE": 14, "MOTORISTA": 18,
    "PLACA MOT.": 14, "H. SOLIC": 17, "H. CHEGADA": 17, "H. INICIAL": 17,
    "H. FINAL": 17, "KM INICIAL": 12, "KM FINAL": 12, "KM TOTAL": 12,
    "FRANQUIA DE HORAS": 16, "H. TRABALHADA": 14, "H. EXTRA": 12,
    "VALOR DA HORA EXTRA": 16, "VALOR H.E": 14, "FRANQUIA DE KM": 14,
    "KM EXCEDENTE": 14, "VLR KM EXTRA": 14, "VALOR K.H": 14,
    "VALOR DA DIÁRIA": 15, "PEDÁGIO": 12, "VALOR TOTAL": 15, "OBS": 20,
}


def _aplicar_formatos_na_linha(ws, numero):
    """number_format por coluna numa linha de dados — APENAS em célula com
    valor: célula None fica sem formato e o lado pendente (sem franquia)
    segue visualmente vazio."""
    for indice, titulo in enumerate(COLUNAS_EXPORTACAO_PAGAMENTOS, start=1):
        formato = _FORMATOS_EXPORTACAO.get(titulo)
        if formato is None:
            continue
        celula = ws.cell(row=numero, column=indice)
        if celula.value is not None:
            celula.number_format = formato


def montar_workbook_pagamentos(acionamentos_por_cliente, compor):
    """Monta o xlsx de pagamentos dos agentes e devolve os BYTES do arquivo.

    `acionamentos_por_cliente`: dict ORDENADO nome do cliente -> lista de
    acionamentos JÁ na ordem certa (quem filtra, ordena e agrupa é a view via
    selector — aqui não há query). `compor`: função que devolve a composição
    do valor do agente (a view injeta compor_valor_agente; o parâmetro existe
    para o service seguir puro e testável com um stub).

    Uma aba por cliente, na ordem de chegada do dict, nome truncado em 31
    caracteres (limite do Excel para título de aba). Cada aba: cabeçalho fixo
    (COLUNAS_EXPORTACAO_PAGAMENTOS), uma linha por acionamento e a linha final
    "Total" com a SOMA dos valor_agente não nulos como número estático —
    nunca fórmula (o arquivo é conferido fora do sistema).

    Import local do openpyxl: dependência exclusiva da exportação, não precisa
    carregar junto com o módulo inteiro de services.
    """
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    negrito = Font(bold=True)
    preenchimento_cabecalho = PatternFill(
        start_color=_COR_CABECALHO_EXPORTACAO,
        end_color=_COR_CABECALHO_EXPORTACAO,
        fill_type="solid",
    )

    wb = Workbook()
    wb.remove(wb.active)  # aba default vazia do openpyxl
    for nome_cliente, acionamentos in acionamentos_por_cliente.items():
        ws = wb.create_sheet(title=nome_cliente[:31])
        ws.append(COLUNAS_EXPORTACAO_PAGAMENTOS)
        for indice, titulo in enumerate(COLUNAS_EXPORTACAO_PAGAMENTOS, start=1):
            celula = ws.cell(row=1, column=indice)
            celula.font = negrito
            celula.fill = preenchimento_cabecalho
            ws.column_dimensions[get_column_letter(indice)].width = (
                _LARGURAS_EXPORTACAO[titulo]
            )
        total = Decimal("0.00")
        for acionamento in acionamentos:
            ws.append(_linha_exportacao(acionamento, compor))
            _aplicar_formatos_na_linha(ws, ws.max_row)
            if acionamento.valor_agente is not None:
                total += acionamento.valor_agente
        linha_total = [None] * len(COLUNAS_EXPORTACAO_PAGAMENTOS)
        linha_total[0] = "Total"
        linha_total[COLUNAS_EXPORTACAO_PAGAMENTOS.index("VALOR TOTAL")] = total
        ws.append(linha_total)
        ws.cell(row=ws.max_row, column=1).font = negrito
        celula_soma = ws.cell(
            row=ws.max_row,
            column=COLUNAS_EXPORTACAO_PAGAMENTOS.index("VALOR TOTAL") + 1,
        )
        celula_soma.number_format = _FORMATO_MOEDA_EXPORTACAO
    if not wb.worksheets:
        # xlsx exige ao menos uma aba; sem acionamento nenhum, sai só o cabeçalho
        wb.create_sheet(title="Acionamentos").append(COLUNAS_EXPORTACAO_PAGAMENTOS)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()