from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
import zoneinfo

from .models import (
    ExecucaoTarefaAdministrativa,
    ComentarioTarefa,
    FuncionarioAdministrativo,
    StatusExecucao,
    TarefaDivisao,
)
from .selectors import (
    get_semana_atual,
    get_dia_atual,
    get_execucoes_da_semana,
    get_blocos_da_semana,
    get_funcionarios_ativos,
    get_resumo_semana,
)
from .services import (
    gerar_execucoes_semana,
    gerar_blocos_semana,
    marcar_atrasadas,
    criar_tarefa_avulsa,
    criar_tarefa_recorrente,
    marcar_comentarios_lidos,
    excluir_execucao,
    adicionar_item_bloco,
    toggle_item_bloco,
    excluir_item_bloco,
    adicionar_comentario_item_bloco,
    atualizar_item_bloco,
    adicionar_tarefa_divisao,
    editar_tarefa_divisao,
    excluir_tarefa_divisao,
)

def _pode_editar(request):
    """Retorna True se o usuário pode editar — admin ou operador."""
    if request.user.is_superuser:
        return True
    try:
        func = FuncionarioAdministrativo.objects.get(usuario=request.user)
        return func.perfil == 'operador'
    except FuncionarioAdministrativo.DoesNotExist:
        return False
    
def _semana_atual_check(semana_iso, ano):
    """Retorna True se a semana informada é a semana atual."""
    from .selectors import get_semana_atual
    semana_atual, ano_atual = get_semana_atual()
    return semana_iso == semana_atual and ano == ano_atual

@login_required
def painel(request, semana_iso=None, ano=None):
    semana_atual, ano_atual = get_semana_atual()

    if semana_iso is None or ano is None:
        semana_iso = semana_atual
        ano        = ano_atual

    semana_iso = int(semana_iso)
    ano        = int(ano)

    marcar_atrasadas(semana_atual, ano_atual)
    gerar_execucoes_semana(semana_iso, ano)
    gerar_blocos_semana(semana_iso, ano, request.user)

    execucoes    = get_execucoes_da_semana(semana_iso, ano, user=request.user)
    blocos       = get_blocos_da_semana(semana_iso, ano)
    funcionarios = get_funcionarios_ativos()
    resumo       = get_resumo_semana(semana_iso, ano)
    dia_atual    = get_dia_atual()

    import datetime
    MESES_PT = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    data_ref        = datetime.date.fromisocalendar(ano, semana_iso, 1)
    data_fim        = data_ref + datetime.timedelta(days=4)
    data_anterior   = data_ref - datetime.timedelta(weeks=1)
    data_proxima    = data_ref + datetime.timedelta(weeks=1)
    semana_anterior = data_anterior.isocalendar()[1]
    ano_anterior    = data_anterior.isocalendar()[0]
    semana_proxima  = data_proxima.isocalendar()[1]
    ano_proxima     = data_proxima.isocalendar()[0]

    if data_ref.month == data_fim.month:
        intervalo_semana = f"{data_ref.day} a {data_fim.day} de {MESES_PT[data_ref.month]} de {ano}"
    else:
        intervalo_semana = f"{data_ref.day} de {MESES_PT[data_ref.month]} a {data_fim.day} de {MESES_PT[data_fim.month]} de {ano}"

    eh_semana_atual = (semana_iso == semana_atual and ano == ano_atual)

    # Tarefas da divisão por funcionário — lista de dicts para o template
    funcionarios_com_tarefas = []
    for func in funcionarios:
        tarefas = list(TarefaDivisao.objects.filter(
            funcionario=func,
            semana_iso=semana_iso,
            ano=ano,
        ).order_by('ordem', 'criado_em'))
        funcionarios_com_tarefas.append({
            'funcionario': func,
            'tarefas_divisao': tarefas,
        })

    # Define perfil do usuário logado
    perfil_usuario = 'admin'
    if not request.user.is_superuser:
        try:
            func = FuncionarioAdministrativo.objects.get(usuario=request.user)
            perfil_usuario = func.perfil
        except FuncionarioAdministrativo.DoesNotExist:
            perfil_usuario = 'visitante'

    context = {
        'execucoes':                execucoes,
        'blocos':                   blocos,
        'funcionarios':             funcionarios,
        'funcionarios_com_tarefas': funcionarios_com_tarefas,
        'resumo':                   resumo,
        'dia_atual':                dia_atual,
        'semana_iso':               semana_iso,
        'ano':                      ano,
        'perfil_usuario':           perfil_usuario,
        'intervalo_semana':         intervalo_semana,
        'eh_semana_atual':          eh_semana_atual,
        'semana_anterior':          semana_anterior,
        'ano_anterior':             ano_anterior,
        'semana_proxima':           semana_proxima,
        'ano_proxima':              ano_proxima,
        'dias': [
            ('segunda', 'Segunda-feira'),
            ('terca',   'Terça-feira'),
            ('quarta',  'Quarta-feira'),
            ('quinta',  'Quinta-feira'),
            ('sexta',   'Sexta-feira'),
        ],
    }
    return render(request, 'controle_administrativo/painel.html', context)


@login_required
@require_POST
def toggle_execucao(request, execucao_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    execucao = get_object_or_404(ExecucaoTarefaAdministrativa, id=execucao_id)
    if not request.user.is_superuser and not _semana_atual_check(execucao.semana_iso, execucao.ano):
        return JsonResponse({'success': False, 'error': 'Semanas passadas não podem ser editadas.'}, status=403)

    execucao.is_done = not execucao.is_done

    if execucao.is_done:
        execucao.status        = StatusExecucao.CONCLUIDA
        execucao.concluido_por = request.user
        execucao.concluido_em  = timezone.now()
    else:
        execucao.status        = StatusExecucao.PENDENTE
        execucao.concluido_por = None
        execucao.concluido_em  = None

    execucao.atualizado_por = request.user
    execucao.save()

    semana_iso, ano = get_semana_atual()
    resumo = get_resumo_semana(semana_iso, ano)

    return JsonResponse({
        'success':    True,
        'is_done':    execucao.is_done,
        'status':     execucao.status,
        'percentual': resumo['percentual'],
        'concluidas': resumo['concluidas'],
        'total':      resumo['total'],
    })


@login_required
@require_POST
def adicionar_comentario(request, execucao_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    execucao = get_object_or_404(ExecucaoTarefaAdministrativa, id=execucao_id)
    if not request.user.is_superuser and not _semana_atual_check(execucao.semana_iso, execucao.ano):
        return JsonResponse({'success': False, 'error': 'Semanas passadas não podem ser editadas.'}, status=403)
    conteudo = request.POST.get('conteudo', '').strip()

    if not conteudo:
        return JsonResponse({'success': False, 'error': 'Comentário não pode ser vazio.'})

    comentario = ComentarioTarefa.objects.create(
        execucao=execucao,
        autor=request.user,
        conteudo=conteudo,
    )

    return JsonResponse({
        'success':   True,
        'id':        comentario.id,
        'autor':     comentario.autor.get_full_name() or comentario.autor.username,
        'conteudo':  comentario.conteudo,
        'criado_em': timezone.localtime(comentario.criado_em).strftime('%d/%m/%Y às %H:%M'),
    })


@login_required
def detalhe_execucao(request, execucao_id):
    execucao = get_object_or_404(
        ExecucaoTarefaAdministrativa.objects.select_related(
            'tarefa_modelo',
            'tarefa_modelo__responsavel',
            'tarefa_modelo__categoria',
            'responsavel_avulso',
            'concluido_por',
            'atualizado_por',
        ).prefetch_related('comentarios', 'comentarios__autor'),
        id=execucao_id
    )

    marcar_comentarios_lidos(execucao, request.user)

    comentarios = [
        {
            'id':        c.id,
            'autor':     c.autor.get_full_name() or c.autor.username if c.autor else 'Desconhecido',
            'conteudo':  c.conteudo,
            'criado_em': timezone.localtime(c.criado_em).strftime('%d/%m/%Y às %H:%M'),
            'is_done':   c.is_done,
        }
        for c in execucao.comentarios.all()
    ]

    return JsonResponse({
        'success':        True,
        'id':             execucao.id,
        'titulo':         execucao.titulo_display,
        'descricao':      execucao.descricao_display,
        'responsavel':    execucao.responsavel_display,
        'dia':            execucao.dia_display,
        'periodo':        execucao.periodo_display,
        'status':         execucao.get_status_display(),
        'is_done':        execucao.is_done,
        'is_avulsa':      execucao.is_avulsa,
        'prazo':          execucao.prazo.strftime('%d/%m/%Y') if execucao.prazo else None,
        'concluido_por':  execucao.concluido_por.get_full_name() or execucao.concluido_por.username if execucao.concluido_por else None,
        'concluido_em':   timezone.localtime(execucao.concluido_em).strftime('%d/%m/%Y às %H:%M') if execucao.concluido_em else None,
        'atualizado_por': execucao.atualizado_por.get_full_name() or execucao.atualizado_por.username if execucao.atualizado_por else None,
        'atualizado_em':  timezone.localtime(execucao.atualizado_em).strftime('%d/%m/%Y às %H:%M') if execucao.atualizado_em else None,
        'comentarios':    comentarios,
    })


@login_required
@require_POST
def criar_tarefa(request):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    titulo         = request.POST.get('titulo', '').strip()
    dia            = request.POST.get('dia', '').strip()
    periodo        = request.POST.get('periodo', '').strip()
    responsavel_id = request.POST.get('responsavel_id', '').strip()
    descricao      = request.POST.get('descricao', '').strip()
    tipo           = request.POST.get('tipo', 'avulsa').strip()

    if not all([titulo, dia, periodo, responsavel_id]):
        return JsonResponse({'success': False, 'error': 'Preencha todos os campos obrigatórios.'})

    semana_iso, ano = get_semana_atual()

    try:
        if tipo == 'recorrente':
            criar_tarefa_recorrente(
                titulo=titulo, dia=dia, periodo=periodo,
                responsavel_id=responsavel_id, descricao=descricao,
                semana_iso=semana_iso, ano=ano,
            )
        else:
            criar_tarefa_avulsa(
                semana_iso=semana_iso, ano=ano, titulo=titulo,
                dia=dia, periodo=periodo, responsavel_id=responsavel_id,
                descricao=descricao,
            )
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': True})


@login_required
@require_POST
def excluir_tarefa(request, execucao_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    execucao = get_object_or_404(ExecucaoTarefaAdministrativa, id=execucao_id)
    if not request.user.is_superuser and not _semana_atual_check(execucao.semana_iso, execucao.ano):
        return JsonResponse({'success': False, 'error': 'Semanas passadas não podem ser editadas.'}, status=403)
    try:
        excluir_execucao(execucao_id, request.user)
        semana_iso, ano = get_semana_atual()
        resumo = get_resumo_semana(semana_iso, ano)
        return JsonResponse({
            'success':    True,
            'percentual': resumo['percentual'],
            'concluidas': resumo['concluidas'],
            'total':      resumo['total'],
        })
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def adicionar_item_bloco_view(request, bloco_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    conteudo = request.POST.get('conteudo', '').strip()
    is_fixo  = request.POST.get('is_fixo', 'false') == 'true'
    try:
        item = adicionar_item_bloco(bloco_id, conteudo, is_fixo, user=request.user)
        return JsonResponse({
            'success':  True,
            'id':       item.id,
            'conteudo': item.conteudo,
            'is_fixo':  item.is_fixo,
            'is_done':  item.is_done,
        })
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def toggle_item_bloco_view(request, item_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    try:
        item = toggle_item_bloco(item_id)
        return JsonResponse({'success': True, 'is_done': item.is_done})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def excluir_item_bloco_view(request, item_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    try:
        excluir_item_bloco(item_id)
        return JsonResponse({'success': True})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def detalhe_item_bloco(request, item_id):
    from .models import ItemBlocoSemanal
    try:
        item = ItemBlocoSemanal.objects.select_related(
            'responsavel', 'criado_por'
        ).prefetch_related('comentarios', 'comentarios__autor').get(id=item_id)
    except ItemBlocoSemanal.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Item não encontrado.'})

    comentarios = [
        {
            'id':        c.id,
            'autor':     c.autor.get_full_name() or c.autor.username if c.autor else 'Desconhecido',
            'conteudo':  c.conteudo,
            'criado_em': timezone.localtime(c.criado_em).strftime('%d/%m/%Y às %H:%M'),
        }
        for c in item.comentarios.all()
    ]

    return JsonResponse({
        'success':          True,
        'id':               item.id,
        'conteudo':         item.conteudo,
        'is_fixo':          item.is_fixo,
        'is_done':          item.is_done,
        'prazo':            item.prazo.strftime('%Y-%m-%d') if item.prazo else '',
        'prazo_fmt':        item.prazo.strftime('%d/%m/%Y') if item.prazo else '—',
        'responsavel_id':   item.responsavel.id if item.responsavel else '',
        'responsavel_nome': item.responsavel.nome if item.responsavel else '—',
        'criado_por':       item.criado_por.get_full_name() or item.criado_por.username if item.criado_por else '—',
        'criado_em':        timezone.localtime(item.criado_em).strftime('%d/%m/%Y às %H:%M'),
        'comentarios':      comentarios,
    })


@login_required
@require_POST
def atualizar_item_bloco_view(request, item_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    conteudo       = request.POST.get('conteudo', '').strip() or None
    responsavel_id = request.POST.get('responsavel_id', None)
    prazo          = request.POST.get('prazo', None)
    is_fixo_str    = request.POST.get('is_fixo', None)
    is_fixo        = (is_fixo_str == 'true') if is_fixo_str is not None else None

    try:
        item = atualizar_item_bloco(item_id, conteudo, responsavel_id, prazo, is_fixo)
        return JsonResponse({'success': True, 'conteudo': item.conteudo, 'is_fixo': item.is_fixo})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def adicionar_comentario_item_bloco_view(request, item_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    conteudo = request.POST.get('conteudo', '').strip()
    try:
        comentario = adicionar_comentario_item_bloco(item_id, conteudo, request.user)
        return JsonResponse({
            'success':   True,
            'id':        comentario.id,
            'autor':     comentario.autor.get_full_name() or comentario.autor.username,
            'conteudo':  comentario.conteudo,
            'criado_em': timezone.localtime(comentario.criado_em).strftime('%d/%m/%Y às %H:%M'),
        })
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def adicionar_tarefa_divisao_view(request, funcionario_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    conteudo  = request.POST.get('conteudo', '').strip()
    is_fixo   = request.POST.get('is_fixo', 'false') == 'true'
    prazo_str = request.POST.get('prazo', '').strip()
    prazo     = prazo_str if prazo_str else None
    semana_iso, ano = get_semana_atual()
    try:
        tarefa = adicionar_tarefa_divisao(funcionario_id, semana_iso, ano, conteudo, is_fixo=is_fixo, prazo=prazo)
        return JsonResponse({
            'success':  True,
            'id':       tarefa.id,
            'conteudo': tarefa.conteudo,
            'is_fixo':  tarefa.is_fixo,
            'prazo':    tarefa.prazo.strftime('%d/%m/%Y') if tarefa.prazo else '—',
        })
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def editar_tarefa_divisao_view(request, tarefa_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    conteudo = request.POST.get('conteudo', '').strip()
    try:
        tarefa = editar_tarefa_divisao(tarefa_id, conteudo)
        return JsonResponse({'success': True, 'conteudo': tarefa.conteudo})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def excluir_tarefa_divisao_view(request, tarefa_id):
    if not _pode_editar(request):
        return JsonResponse({'success': False, 'error': 'Sem permissão.'}, status=403)
    try:
        excluir_tarefa_divisao(tarefa_id)
        return JsonResponse({'success': True})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def historico(request, semana_iso=None, ano=None):
    """Visualização somente leitura de qualquer semana."""
    semana_atual, ano_atual = get_semana_atual()

    # Se não foi passada semana, usa a atual
    if semana_iso is None or ano is None:
        semana_iso = semana_atual
        ano        = ano_atual

    semana_iso = int(semana_iso)
    ano        = int(ano)

    # Calcula semana anterior e próxima para navegação
    import datetime
    MESES_PT = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }
    data_ref        = datetime.date.fromisocalendar(ano, semana_iso, 1)
    data_fim        = data_ref + datetime.timedelta(days=4)
    data_anterior   = data_ref - datetime.timedelta(weeks=1)
    data_proxima    = data_ref + datetime.timedelta(weeks=1)
    semana_anterior = data_anterior.isocalendar()[1]
    ano_anterior    = data_anterior.isocalendar()[0]
    semana_proxima  = data_proxima.isocalendar()[1]
    ano_proxima     = data_proxima.isocalendar()[0]

    if data_ref.month == data_fim.month:
        intervalo_semana = f"{data_ref.day} a {data_fim.day} de {MESES_PT[data_ref.month]} de {ano}"
    else:
        intervalo_semana = f"{data_ref.day} de {MESES_PT[data_ref.month]} a {data_fim.day} de {MESES_PT[data_fim.month]} de {ano}"

    execucoes    = get_execucoes_da_semana(semana_iso, ano, user=request.user)
    blocos       = get_blocos_da_semana(semana_iso, ano)
    funcionarios = get_funcionarios_ativos()
    resumo       = get_resumo_semana(semana_iso, ano)

    # Tarefas da divisão por funcionário
    funcionarios_com_tarefas = []
    for func in funcionarios:
        tarefas = list(TarefaDivisao.objects.filter(
            funcionario=func,
            semana_iso=semana_iso,
            ano=ano,
        ).order_by('ordem', 'criado_em'))
        funcionarios_com_tarefas.append({
            'funcionario': func,
            'tarefas_divisao': tarefas,
        })

    eh_semana_atual = (semana_iso == semana_atual and ano == ano_atual)

    context = {
        'execucoes':                execucoes,
        'blocos':                   blocos,
        'funcionarios':             funcionarios,
        'funcionarios_com_tarefas': funcionarios_com_tarefas,
        'resumo':                   resumo,
        'semana_iso':               semana_iso,
        'ano':                      ano,
        'eh_semana_atual':          eh_semana_atual,
        'semana_anterior':          semana_anterior,
        'ano_anterior':             ano_anterior,
        'semana_proxima':           semana_proxima,
        'ano_proxima':              ano_proxima,
        'intervalo_semana':         intervalo_semana,
        'dias': [
            ('segunda', 'Segunda-feira'),
            ('terca',   'Terça-feira'),
            ('quarta',  'Quarta-feira'),
            ('quinta',  'Quinta-feira'),
            ('sexta',   'Sexta-feira'),
        ],
    }
    return render(request, 'controle_administrativo/historico.html', context)

@login_required
def exportar_excel(request, semana_iso, ano):
    """Exporta as execuções da semana em formato Excel com abas por funcionário."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    semana_iso = int(semana_iso)
    ano        = int(ano)

    from .models import FuncionarioAdministrativo, ExecucaoTarefaAdministrativa

    funcionarios = FuncionarioAdministrativo.objects.filter(
        perfil='operador', ativo=True
    ).order_by('nome')

    execucoes = ExecucaoTarefaAdministrativa.objects.filter(
        semana_iso=semana_iso, ano=ano
    ).exclude(status='cancelada').select_related(
        'tarefa_modelo', 'tarefa_modelo__responsavel',
        'responsavel_avulso', 'concluido_por',
    ).prefetch_related('comentarios').order_by(
        'tarefa_modelo__dia_da_semana', 'tarefa_modelo__periodo'
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Cores ──
    AZUL_ESCURO  = '1F3864'
    AZUL_CLARO   = 'D9E1F2'
    VERDE_CLARO  = 'E2EFDA'
    VERMELHO_CL  = 'FCE4D6'
    AMARELO_CL   = 'FFF2CC'
    CINZA_CLARO  = 'F2F2F2'
    BRANCO       = 'FFFFFF'
    VERDE_TEXTO  = '375623'
    VERMELHO_TX  = '9C0006'
    AMARELO_TX   = '7D6608'

    # ── Bordas ──
    borda = Border(
        left   = Side(style='thin', color='BFBFBF'),
        right  = Side(style='thin', color='BFBFBF'),
        top    = Side(style='thin', color='BFBFBF'),
        bottom = Side(style='thin', color='BFBFBF'),
    )

    DIAS_ORDEM  = ['segunda', 'terca', 'quarta', 'quinta', 'sexta']
    DIAS_PT     = {'segunda': 'Segunda-feira', 'terca': 'Terça-feira', 'quarta': 'Quarta-feira', 'quinta': 'Quinta-feira', 'sexta': 'Sexta-feira'}
    PERIODOS_PT = {'manha': 'Manhã', 'tarde': 'Tarde'}
    STATUS_PT   = {'pendente': 'Pendente', 'em_andamento': 'Em Andamento', 'concluida': 'Concluída', 'atrasada': 'Atrasada'}

    for func in funcionarios:
        ws = wb.create_sheet(title=func.nome[:31])
        ws.sheet_view.showGridLines = False

        # ── Título ──
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Controle Administrativo — {func.nome} — Semana {semana_iso}/{ano}'
        ws['A1'].font      = Font(bold=True, size=13, color=BRANCO)
        ws['A1'].fill      = PatternFill('solid', fgColor=AZUL_ESCURO)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # ── Cabeçalho ──
        headers = ['Dia', 'Período', 'Tarefa', 'Status', 'Comentários', 'Concluído por']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = Font(bold=True, size=10, color=AZUL_ESCURO)
            cell.fill      = PatternFill('solid', fgColor=AZUL_CLARO)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = borda
        ws.row_dimensions[2].height = 18

        # ── Dados ──
        execucoes_func = [
            e for e in execucoes
            if (e.tarefa_modelo and e.tarefa_modelo.responsavel_id == func.id)
            or (e.is_avulsa and e.responsavel_avulso_id == func.id)
        ]

        execucoes_func.sort(key=lambda e: (
            DIAS_ORDEM.index(e.dia_key) if e.dia_key in DIAS_ORDEM else 99,
            e.periodo_key or ''
        ))

        row = 3
        for i, e in enumerate(execucoes_func):
            comentarios_texto = ' | '.join([c.conteudo for c in e.comentarios.all()])
            concluido_por = ''
            if e.concluido_por:
                concluido_por = e.concluido_por.get_full_name() or e.concluido_por.username

            # Cor por status
            if e.status == 'concluida':
                bg, fg = VERDE_CLARO, VERDE_TEXTO
            elif e.status == 'atrasada':
                bg, fg = VERMELHO_CL, VERMELHO_TX
            elif e.status == 'em_andamento':
                bg, fg = AMARELO_CL, AMARELO_TX
            else:
                bg = CINZA_CLARO if i % 2 == 0 else BRANCO
                fg = '000000'

            valores = [
                DIAS_PT.get(e.dia_key, e.dia_key or '—'),
                PERIODOS_PT.get(e.periodo_key, e.periodo_key or '—'),
                e.titulo_display,
                STATUS_PT.get(e.status, e.status),
                comentarios_texto or '—',
                concluido_por or '—',
            ]

            for col, val in enumerate(valores, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill      = PatternFill('solid', fgColor=bg)
                cell.font      = Font(size=10, color=fg)
                cell.border    = borda
                cell.alignment = Alignment(
                    horizontal='center' if col <= 2 else 'left',
                    vertical='center',
                    wrap_text=True
                )

            ws.row_dimensions[row].height = 16
            row += 1

        # ── Larguras ──
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="controle_adm_semana_{semana_iso}_{ano}.xlsx"'
    wb.save(response)
    return response