import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db.models import Case, When, IntegerField
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from functools import wraps

from kanban_inteligencia.models import TarefaInteligencia
from .models import TarefaTI


# ===== INBOX DE TICKETS DESTINADOS AO T.I =====

# Mapeia a prioridade da TarefaInteligencia (que possui 'avaliar') para os
# valores aceitos pela TarefaTI (baixa/media/alta).
PRIORIDADE_TICKET_PARA_TI = {
    'avaliar': 'media',
    'baixa': 'baixa',
    'media': 'media',
    'alta': 'alta',
}


def tickets_inbox_ti():
    """Tickets destinados ao T.I aguardando triagem (ainda não enviados ao board)."""
    return (
        TarefaInteligencia.objects
        .filter(destinado='TI', enviado_kanban_ti=False)
        .select_related('usuario')
        .order_by('-data_criacao', '-id')
    )


# ===== UTILITÁRIOS =====

def order_by_prioridade(queryset):
    """Ordena queryset por prioridade (alta > média > baixa)"""
    return queryset.annotate(
        prioridade_order=Case(
            When(prioridade='alta', then=3),
            When(prioridade='media', then=2),
            When(prioridade='baixa', then=1),
            default=0,
            output_field=IntegerField()
        )
    ).order_by('-prioridade_order')


def aplicar_conclusao(tarefa, status):
    """
    Aplica lógica de conclusão/reabertura de tarefa.
    - Se concluído: define data_conclusao e zera prazo
    - Se reaberto: limpa data_conclusao
    """
    if status == 'concluido':
        if not tarefa.data_conclusao:
            tarefa.data_conclusao = timezone.now().date()
        tarefa.data_limite = None
    else:
        tarefa.data_conclusao = None


def sincronizar_ticket_origem(tarefa):
    """
    Reflete a conclusão/reabertura de uma TarefaTI no ticket original
    (TarefaInteligencia destinado ao T.I), para que a central de tickets some
    o chamado quando a tarefa é movida para "Concluído" — e o traga de volta
    caso seja reaberta.
    """
    origem = tarefa.ticket_origem
    if not origem:
        return

    if tarefa.status == 'concluido':
        origem.status = 'concluido'
        if not origem.data_conclusao:
            origem.data_conclusao = tarefa.data_conclusao or timezone.now().date()
        origem.data_limite = None
        origem.mensagem_status = 'Ticket concluído.'
        origem.save(update_fields=['status', 'data_conclusao', 'data_limite', 'mensagem_status'])
    elif tarefa.status in ('em_progresso', 'validacao'):
        # Em execução no board TI → "Em andamento" (amarelo) na central,
        # refletindo o prazo definido na edição do card.
        origem.status = 'em_progresso'
        origem.data_conclusao = None
        origem.data_limite = tarefa.data_limite
        origem.mensagem_status = 'Ticket em andamento'
        origem.save(update_fields=['status', 'data_conclusao', 'data_limite', 'mensagem_status'])
    elif origem.status == 'concluido':
        # Tarefa reaberta no board → o ticket volta a aparecer na central.
        origem.status = 'a_fazer'
        origem.data_conclusao = None
        origem.mensagem_status = 'Ticket avaliado'
        origem.save(update_fields=['status', 'data_conclusao', 'mensagem_status'])


def json_error(message, status=400):
    """Retorna JsonResponse de erro padronizado"""
    return JsonResponse({'success': False, 'error': message}, status=status)


def json_success(**kwargs):
    """Retorna JsonResponse de sucesso padronizado"""
    return JsonResponse({'success': True, **kwargs})


def ti_or_superuser_required(view_func):
    """Permite acesso apenas para superusers ou usuários do grupo 'TI'."""
    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        user = request.user
        if user.is_superuser or user.groups.filter(name='TI').exists():
            return view_func(request, *args, **kwargs)
        return HttpResponseForbidden('Acesso negado')

    return _wrapped


# ===== VIEWS =====

@ti_or_superuser_required
def home(request):
    tarefas = TarefaTI.objects.all()

    destinado_filter = request.GET.get('destinado')
    if destinado_filter and destinado_filter != 'all':
        tarefas = tarefas.filter(destinado=destinado_filter)

    a_fazer = order_by_prioridade(tarefas.filter(status='a_fazer'))
    em_progresso = order_by_prioridade(tarefas.filter(status='em_progresso'))
    validacao = order_by_prioridade(tarefas.filter(status='validacao'))
    concluido = order_by_prioridade(tarefas.filter(status='concluido'))

    total_tarefas = tarefas.count()
    total_concluidas = concluido.count()
    taxa_conclusao = round((total_concluidas / total_tarefas * 100) if total_tarefas > 0 else 0)

    context = {
        'a_fazer': a_fazer,
        'em_progresso': em_progresso,
        'validacao': validacao,
        'concluido': concluido,
        'responsavel_choices': TarefaTI.RESPONSAVEL_CHOICES,
        'stats': {
            'total_tarefas': total_tarefas,
            'concluidas': total_concluidas,
            'em_progresso': em_progresso.count(),
            'atrasadas': sum(1 for t in tarefas if t.esta_atrasada),
            'taxa_conclusao': taxa_conclusao,
        },
        'destinado_filter': destinado_filter or 'all',
        'tickets_inbox_count': tickets_inbox_ti().count(),
    }

    return render(request, 'kanban_TI/home.html', context)


@ti_or_superuser_required
@require_POST
def adicionar_tarefa(request):
    try:
        TarefaTI.objects.create(
            titulo=request.POST.get('titulo'),
            descricao=request.POST.get('descricao', ''),
            data_limite=request.POST.get('data_limite') or None,
            responsavel=request.POST.get('responsavel') or None,
            responsavel_cor=request.POST.get('responsavel_cor', 'azul'),
            cor=request.POST.get('cor', 'azul'),
            prioridade=request.POST.get('prioridade', 'media'),
            imagem=request.FILES.get('imagem'),
        )
        return json_success()

    except Exception as e:
        return json_error(str(e))


@ti_or_superuser_required
def obter_tarefa(request, tarefa_id):
    try:
        tarefa = TarefaTI.objects.get(id=tarefa_id)

        return json_success(tarefa={
            'id': tarefa.id,
            'titulo': tarefa.titulo,
            'descricao': tarefa.descricao,
            'status': tarefa.status,
            'data_limite': tarefa.data_limite.strftime('%Y-%m-%d') if tarefa.data_limite else '',
            'data_criacao': tarefa.data_criacao.strftime('%d/%m/%Y'),
            'data_conclusao': tarefa.data_conclusao.strftime('%d/%m/%Y') if tarefa.data_conclusao else None,
            'esta_atrasada': tarefa.esta_atrasada,
            'imagem_url': tarefa.imagem.url if tarefa.imagem else None,
            'responsavel': tarefa.responsavel or '',
            'responsavel_label': tarefa.get_responsavel_display() if tarefa.responsavel else '',
            'responsavel_cor': tarefa.responsavel_cor,
            'mensagem_status': tarefa.ticket_origem.mensagem_status if tarefa.ticket_origem else '',
            'criador': (tarefa.usuario.get_full_name() or tarefa.usuario.username) if tarefa.usuario else '',
            'cor': tarefa.cor,
            'prioridade': tarefa.prioridade,
        })

    except TarefaTI.DoesNotExist:
        return json_error('Tarefa não encontrada', status=404)


@ti_or_superuser_required
@require_POST
def atualizar_tarefa(request, tarefa_id):
    try:
        tarefa = TarefaTI.objects.get(id=tarefa_id)

        tarefa.titulo = request.POST.get('titulo', tarefa.titulo)
        tarefa.descricao = request.POST.get('descricao', tarefa.descricao)
        tarefa.status = request.POST.get('status', tarefa.status)
        tarefa.prioridade = request.POST.get('prioridade', tarefa.prioridade)
        tarefa.responsavel = request.POST.get('responsavel') or None
        tarefa.responsavel_cor = request.POST.get('responsavel_cor', tarefa.responsavel_cor)
        tarefa.cor = request.POST.get('cor', tarefa.cor)

        data_limite = request.POST.get('data_limite')
        if tarefa.status != 'concluido' and data_limite:
            tarefa.data_limite = data_limite

        aplicar_conclusao(tarefa, tarefa.status)

        if request.FILES.get('imagem'):
            tarefa.imagem = request.FILES.get('imagem')

        tarefa.save()
        sincronizar_ticket_origem(tarefa)
        return json_success()

    except TarefaTI.DoesNotExist:
        return json_error('Tarefa não encontrada', status=404)

    except Exception as e:
        return json_error(str(e))


@ti_or_superuser_required
@require_POST
def atualizar_status(request, tarefa_id):
    try:
        tarefa = TarefaTI.objects.get(id=tarefa_id)
        novo_status = request.POST.get('status')

        tarefa.status = novo_status
        # Prazo definido no popup ao mover para "Em progresso"/"Validação".
        if novo_status in ('em_progresso', 'validacao'):
            prazo_data = request.POST.get('data_limite')
            if prazo_data:
                tarefa.data_limite = prazo_data
        aplicar_conclusao(tarefa, novo_status)
        tarefa.save()
        sincronizar_ticket_origem(tarefa)

        sla_dias = None
        if tarefa.data_conclusao:
            sla_dias = (tarefa.data_conclusao - tarefa.data_criacao).days

        return json_success(
            status=tarefa.status,
            data_conclusao=tarefa.data_conclusao.strftime('%d/%m/%Y') if tarefa.data_conclusao else None,
            sla_dias=sla_dias
        )

    except TarefaTI.DoesNotExist:
        return json_error('Tarefa não encontrada', status=404)

    except Exception as e:
        return json_error(str(e))


@ti_or_superuser_required
@require_POST
def deletar_tarefa(request, tarefa_id):
    try:
        tarefa = TarefaTI.objects.get(id=tarefa_id)
        tarefa.delete()
        return json_success()

    except TarefaTI.DoesNotExist:
        return json_error('Tarefa não encontrada', status=404)

    except Exception as e:
        return json_error(str(e))


# ===== VIEWS DA INBOX (TICKETS DESTINADOS AO T.I) =====

# Máximo de tickets exibidos por página dentro do modal de triagem.
TICKETS_POR_PAGINA_MODAL = 5


def _contexto_triagem(extra=None):
    """Choices reutilizadas pelos formulários da triagem (definir como TarefaTI)."""
    ctx = {
        'responsaveis': TarefaTI.RESPONSAVEL_CHOICES,
        'cores': TarefaTI.COR_CHOICES,
        'prioridades': TarefaTI.PRIORIDADE_CHOICES,
        'status_choices': TarefaTI.STATUS_CHOICES,
    }
    if extra:
        ctx.update(extra)
    return ctx


@ti_or_superuser_required
def inbox_tickets_partial(request):
    """Tickets aguardando triagem (corpo do modal), paginados — 5 por página."""
    try:
        pagina_num = int(request.GET.get('pagina', 1))
    except (TypeError, ValueError):
        pagina_num = 1

    paginator = Paginator(tickets_inbox_ti(), TICKETS_POR_PAGINA_MODAL)
    page_obj = paginator.get_page(pagina_num)

    return render(request, 'kanban_TI/partials/inbox_tickets.html', _contexto_triagem({
        'tickets_inbox': page_obj.object_list,
        'page_obj': page_obj,
    }))


@ti_or_superuser_required
def inbox_badge(request):
    """Retorna apenas o badge com a contagem atualizada da inbox."""
    return render(request, 'kanban_TI/partials/inbox_badge.html', {
        'tickets_inbox_count': tickets_inbox_ti().count(),
    })


@ti_or_superuser_required
def tickets_novos(request):
    """
    Retorna tickets destinados ao T.I criados após o timestamp informado.

    Consumido por polling leve (a cada 5s) pelo JS do board para exibir um toast
    de "Novo ticket" no canto inferior direito. Endpoint enxuto — devolve apenas
    o necessário para montar a notificação.
    """
    desde_ts = request.GET.get('desde')

    try:
        desde = timezone.datetime.fromtimestamp(
            int(desde_ts) / 1000,
            tz=timezone.get_current_timezone(),
        )
    except (TypeError, ValueError):
        desde = timezone.now() - timedelta(seconds=10)

    tickets = (
        TarefaInteligencia.objects
        .filter(destinado='TI', criado_em__gt=desde)
        .select_related('usuario')
        .order_by('criado_em')
    )

    return JsonResponse({
        'tickets': [
            {
                'id': str(t.pk),
                'titulo': t.titulo,
                'criado_por': (
                    (t.usuario.get_full_name() or t.usuario.username)
                    if t.usuario else 'Desconhecido'
                ),
                'criado_em': t.criado_em.isoformat() if t.criado_em else '',
            }
            for t in tickets
        ]
    })


@ti_or_superuser_required
@require_POST
def ticket_mover_para_afazer(request, pk):
    """
    Move um ticket da inbox para a coluna "A fazer" do Kanban TI.

    Cria uma TarefaTI a partir do ticket (TarefaInteligencia) e marca o ticket
    de origem como já enviado, removendo-o da inbox. Retorna a lista atualizada
    para o modal.
    """
    ticket = get_object_or_404(
        TarefaInteligencia, pk=pk, destinado='TI', enviado_kanban_ti=False
    )

    TarefaTI.objects.create(
        titulo=ticket.titulo,
        descricao=ticket.descricao or '',
        status='a_fazer',
        prioridade=PRIORIDADE_TICKET_PARA_TI.get(ticket.prioridade, 'media'),
        data_limite=ticket.data_limite,
        imagem=ticket.imagem or None,
        usuario=ticket.usuario,
        ticket_origem=ticket,
    )

    ticket.enviado_kanban_ti = True
    ticket.save(update_fields=['enviado_kanban_ti'])

    return render(request, 'kanban_TI/partials/inbox_tickets.html', _contexto_triagem({
        'tickets_inbox': tickets_inbox_ti(),
    }))


@ti_or_superuser_required
def ticket_editar_partial(request, pk):
    """Formulário de edição do ticket de triagem — carregado pelo lápis (HTMX)."""
    ticket = get_object_or_404(
        TarefaInteligencia, pk=pk, destinado='TI', enviado_kanban_ti=False
    )
    return render(request, 'kanban_TI/partials/ticket_editar.html', _contexto_triagem({
        'ticket': ticket,
    }))


@ti_or_superuser_required
def ticket_recusar_partial(request, pk):
    """Formulário de recusa do ticket de triagem — carregado pelo botão recusar."""
    ticket = get_object_or_404(
        TarefaInteligencia, pk=pk, destinado='TI', enviado_kanban_ti=False
    )
    return render(request, 'kanban_TI/partials/ticket_recusar.html', {
        'ticket': ticket,
    })


@ti_or_superuser_required
@require_POST
def ticket_recusar(request, pk):
    """
    Recusa um ticket da inbox de triagem informando o motivo.

    Marca o ticket de origem (TarefaInteligencia) como recusado, guarda o
    motivo (exibido ao criador na central de tickets), remove-o da inbox e
    devolve a inbox atualizada para o HTMX substituir o corpo do modal.
    """
    ticket = get_object_or_404(
        TarefaInteligencia, pk=pk, destinado='TI', enviado_kanban_ti=False
    )

    motivo = (request.POST.get('motivo') or '').strip()

    ticket.status = 'recusado'
    ticket.motivo_recusa = motivo
    ticket.mensagem_status = (
        f'Ticket recusado: {motivo}' if motivo else 'Ticket recusado.'
    )
    ticket.data_limite = None
    ticket.data_conclusao = None
    # Sai da inbox de triagem — não volta a aparecer para nova avaliação.
    ticket.enviado_kanban_ti = True
    ticket.save(update_fields=[
        'status', 'motivo_recusa', 'mensagem_status',
        'data_limite', 'data_conclusao', 'enviado_kanban_ti',
    ])

    return inbox_tickets_partial(request)


@ti_or_superuser_required
@require_POST
def ticket_salvar_triagem(request, pk):
    """
    Cria a TarefaTI a partir do ticket de triagem com os dados do formulário
    (status, prazo, responsável, etiqueta/cor e prioridade) e remove o ticket da
    inbox. Reflete o andamento no ticket de origem (central de tickets) e
    devolve a inbox atualizada para o HTMX substituir o corpo do modal.
    """
    ticket = get_object_or_404(
        TarefaInteligencia, pk=pk, destinado='TI', enviado_kanban_ti=False
    )

    status = request.POST.get('status') or 'a_fazer'
    if status not in dict(TarefaTI.STATUS_CHOICES):
        status = 'a_fazer'
    prioridade = request.POST.get('prioridade') or None
    responsavel = request.POST.get('responsavel') or None
    etiqueta = request.POST.get('etiqueta') or 'azul'
    # Prazo agora é uma data (calendário, formato YYYY-MM-DD). Vazio mantém o atual.
    prazo = (request.POST.get('prazo') or '').strip()
    data_limite = prazo or ticket.data_limite or None

    tarefa = TarefaTI.objects.create(
        titulo=ticket.titulo,
        descricao=ticket.descricao or '',
        status=status,
        prioridade=prioridade,
        responsavel=responsavel,
        responsavel_cor=etiqueta,
        cor=etiqueta,
        data_limite=data_limite,
        imagem=ticket.imagem or None,
        usuario=ticket.usuario,
        ticket_origem=ticket,
    )

    # Reflete o andamento no ticket de origem (central) e remove da inbox.
    # O prazo (data_limite) aparece na central junto da mensagem de status.
    ticket.status = status
    ticket.data_limite = data_limite
    ticket.mensagem_status = 'Ticket avaliado' if status == 'a_fazer' else 'Ticket em andamento'
    ticket.enviado_kanban_ti = True
    ticket.save(update_fields=[
        'status', 'data_limite', 'mensagem_status', 'enviado_kanban_ti',
    ])

    # Caso entre direto como concluído, aplica conclusão/SLA de forma consistente.
    if status == 'concluido':
        aplicar_conclusao(tarefa, status)
        tarefa.save()
        sincronizar_ticket_origem(tarefa)

    return inbox_tickets_partial(request)


@ti_or_superuser_required
def exportar_tarefas(request):

    tarefas = TarefaTI.objects.all().order_by('-data_criacao')

    STATUS_ROTULOS = {
        'a_fazer': 'A Fazer',
        'em_progresso': 'Em Progresso',
        'validacao': 'Aguardando Validação',
        'concluido': 'Concluído',
    }

    PRIORIDADE_ROTULOS = {
        'baixa': 'Baixa',
        'media': 'Média',
        'alta': 'Alta',
    }

    estilo_cabecalho_fonte = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    estilo_cabecalho_fundo = PatternFill('solid', fgColor='1B1F2E')
    estilo_cabecalho_alinhamento = Alignment(horizontal='center', vertical='center', wrap_text=True)

    estilo_dado_fonte = Font(name='Arial', size=10, color='222222')
    estilo_dado_alinhamento = Alignment(vertical='center', wrap_text=True)

    estilo_borda = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    cores_status = {
        'a_fazer':      PatternFill('solid', fgColor='D6EAF8'),
        'em_progresso': PatternFill('solid', fgColor='FEF5E7'),
        'validacao':    PatternFill('solid', fgColor='FADBD8'),
        'concluido':    PatternFill('solid', fgColor='D5F5E3'),
    }

    cores_prioridade = {
        'baixa': PatternFill('solid', fgColor='D5F5E3'),
        'media': PatternFill('solid', fgColor='FEF5E7'),
        'alta':  PatternFill('solid', fgColor='FADBD8'),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = 'Chamados TI'
    ws.sheet_view.showGridLines = False

    colunas = [
        'Nº Chamado', 'Título', 'Descrição', 'Status',
        'Responsável', 'Prioridade', 'Data de Abertura',
        'Data de Conclusão', 'SLA (dias)',
    ]

    ws.row_dimensions[1].height = 28

    for col_idx, nome in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=col_idx, value=nome)
        celula.font = estilo_cabecalho_fonte
        celula.fill = estilo_cabecalho_fundo
        celula.alignment = estilo_cabecalho_alinhamento
        celula.border = estilo_borda

    for linha, tarefa in enumerate(tarefas, start=2):

        if tarefa.data_conclusao:
            sla = (tarefa.data_conclusao - tarefa.data_criacao).days
        else:
            sla = '—'

        valores = [
            f'INT-{tarefa.id:03d}',
            tarefa.titulo,
            tarefa.descricao or '—',
            STATUS_ROTULOS.get(tarefa.status, tarefa.status),
            tarefa.get_responsavel_display() if tarefa.responsavel else '—',
            PRIORIDADE_ROTULOS.get(tarefa.prioridade, tarefa.prioridade),
            tarefa.data_criacao.strftime('%d/%m/%Y') if tarefa.data_criacao else '—',
            tarefa.data_conclusao.strftime('%d/%m/%Y') if tarefa.data_conclusao else '—',
            sla,
        ]

        for col_idx, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha, column=col_idx, value=valor)
            celula.font = estilo_dado_fonte
            celula.alignment = estilo_dado_alinhamento
            celula.border = estilo_borda

        ws.cell(row=linha, column=4).fill = cores_status.get(tarefa.status)

        ws.cell(row=linha, column=6).fill = cores_prioridade.get(tarefa.prioridade)

        if linha % 2 == 0:
            fundo_linha = PatternFill('solid', fgColor='F4F6F7')
            for col_idx in range(1, len(colunas) + 1):
                celula = ws.cell(row=linha, column=col_idx)
                if col_idx not in (4, 6):
                    celula.fill = fundo_linha

    larguras = {'A': 14, 'B': 38, 'C': 48, 'D': 24, 'E': 18, 'F': 15, 'G': 20, 'H': 20, 'I': 14}
    for letra, largura in larguras.items():
        ws.column_dimensions[letra].width = largura

    ws.freeze_panes = 'A2'

    resumo = wb.create_sheet('Resumo')

    resumo['A1'] = 'Resumo dos Chamados'
    resumo['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    resumo['A1'].fill = PatternFill('solid', fgColor='1B1F2E')
    resumo.merge_cells('A1:B1')
    resumo.row_dimensions[1].height = 30

    dados_resumo = [
        ('Total de Chamados',        tarefas.count()),
        ('A Fazer',                  tarefas.filter(status='a_fazer').count()),
        ('Em Progresso',             tarefas.filter(status='em_progresso').count()),
        ('Aguardando Validação',    tarefas.filter(status='validacao').count()),
        ('Concluídos',               tarefas.filter(status='concluido').count()),
        ('Atrasados',                sum(1 for t in tarefas if t.esta_atrasada)),
    ]

    for i, (rotulo, valor) in enumerate(dados_resumo, start=2):
        resumo.cell(row=i, column=1, value=rotulo).font = Font(name='Arial', bold=True, size=10)
        resumo.cell(row=i, column=1).border = estilo_borda
        resumo.cell(row=i, column=2, value=valor).font = Font(name='Arial', size=10)
        resumo.cell(row=i, column=2).border = estilo_borda

    resumo.column_dimensions['A'].width = 28
    resumo.column_dimensions['B'].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"Chamados_TI_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'

    return response