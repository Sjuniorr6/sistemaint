import io
from datetime import datetime

from django.contrib.auth.models import Permission, User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from acompanhamento.models import Clientes
from reativacao.models import IdIccid, Reativacao


def _dia(ano, mes, dia):
    """Datetime aware ao meio-dia no fuso da operação (evita virar o dia no __date)."""
    return timezone.make_aware(datetime(ano, mes, dia, 12, 0))


class HistoricoReativacoesExcelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.cliente = Clientes.objects.create(
            nome='Cliente Teste', endereco='Rua A, 100', cnpj='00.000.000/0001-00'
        )
        cls.outro_cliente = Clientes.objects.create(
            nome='Outro Cliente', endereco='Rua B, 200', cnpj='11.111.111/0001-11'
        )

        # Três reativações em dias distintos (data_hora_criacao é auto_now_add,
        # então gravamos a data via update, que não dispara o auto_now_add).
        cls.r_jan = cls._reativacao_em(cls.cliente, _dia(2026, 1, 10), obs='janeiro')
        cls.r_fev = cls._reativacao_em(cls.cliente, _dia(2026, 2, 15), obs='fevereiro')
        cls.r_mar = cls._reativacao_em(cls.outro_cliente, _dia(2026, 3, 20), obs='marco')

        IdIccid.objects.create(
            reativacao=cls.r_fev, id_equipamentos='EQ1\nEQ2', ccid_equipamentos='CC1\nCC2', quantidade=2
        )

    @staticmethod
    def _reativacao_em(cliente, quando, obs=''):
        r = Reativacao.objects.create(nome=cliente, observacoes=obs)
        Reativacao.objects.filter(pk=r.pk).update(data_hora_criacao=quando)
        r.refresh_from_db()
        return r

    def _usuario_com_permissao(self):
        user = User.objects.create_user('operador', password='x')
        user.user_permissions.add(Permission.objects.get(codename='view_reativacao'))
        return user

    # ----- período na listagem -----

    def test_listagem_filtra_pelo_periodo(self):
        self.client.force_login(self._usuario_com_permissao())
        resp = self.client.get(
            reverse('reativacao_list'),
            {'data_inicio': '2026-02-01', 'data_fim': '2026-02-28'},
        )
        self.assertEqual(resp.status_code, 200)
        ids = {r.id for r in resp.context['reativacoes']}
        # Só fevereiro cai na janela — janeiro e março ficam de fora.
        self.assertEqual(ids, {self.r_fev.id})

    # ----- export -----

    def test_export_retorna_xlsx(self):
        self.client.force_login(self._usuario_com_permissao())
        resp = self.client.get(reverse('reativacao_export_excel'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', resp['Content-Disposition'])
        self.assertIn('reativacoes.xlsx', resp['Content-Disposition'])

        ws = load_workbook(io.BytesIO(resp.getvalue())).active
        # Cabeçalho + 3 registros
        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws.cell(row=1, column=1).value, 'ID')

    def test_export_respeita_o_periodo(self):
        self.client.force_login(self._usuario_com_permissao())
        resp = self.client.get(
            reverse('reativacao_export_excel'),
            {'data_inicio': '2026-02-01', 'data_fim': '2026-02-28'},
        )
        ws = load_workbook(io.BytesIO(resp.getvalue())).active
        # Cabeçalho + só o registro de fevereiro
        self.assertEqual(ws.max_row, 2)
        linha = [c.value for c in ws[2]]
        self.assertEqual(linha[0], self.r_fev.id)
        self.assertEqual(linha[7], 'fevereiro')   # coluna Observações
        self.assertEqual(linha[11], 2)            # coluna Quantidade (soma dos id_iccids)

    # ----- permissão -----

    def test_export_sem_permissao_e_bloqueado(self):
        # Usuário logado, porém SEM a permissão view_reativacao.
        sem_perm = User.objects.create_user('curioso', password='x')
        self.client.force_login(sem_perm)
        resp = self.client.get(reverse('reativacao_export_excel'))
        self.assertIn(resp.status_code, (302, 403))
