import requests
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .forms import ReativacaoForm, IdIccidFormSet, IdIccidForm
from .models import Reativacao, IdIccid, Clientes
from django.views.generic import ListView
from requisicao.models import Requisicoes
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic.edit import UpdateView
from .models import Reativacao
from .forms import ReativacaoForm
from django.urls import reverse_lazy
import logging
from django.contrib import messages
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReativacaoIdIccidCreateView(PermissionRequiredMixin, LoginRequiredMixin, View):
    permission_required = 'reativacao.add_reativacao'

    def get(self, request):
        reativacao_form = ReativacaoForm()
        id_iccid_formset = IdIccidFormSet(queryset=IdIccid.objects.none())
        return render(request, 'reativacao_id_iccid_form.html', {
            'reativacao_form': reativacao_form,
            'id_iccid_formset': id_iccid_formset
        })

    def post(self, request):
        reativacao_form = ReativacaoForm(request.POST)
        id_iccid_formset = IdIccidFormSet(request.POST)

        if reativacao_form.is_valid() and id_iccid_formset.is_valid():
            # Salva a Reativação associando o usuário logado
            reativacao = reativacao_form.save(commit=False)
            reativacao.usuario = request.user
            reativacao.save()

            # Processa cada formulário do formset
            for form in id_iccid_formset:
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                    # Recupera o conteúdo do campo id_equipamentos
                    id_equip_text = form.cleaned_data.get('id_equipamentos', '')
                    # Divide o texto em linhas e ignora as linhas em branco
                    linhas = [linha for linha in id_equip_text.splitlines() if linha.strip()]
                    quantidade = len(linhas)  # Cada linha não vazia conta como 1

                    instance = form.save(commit=False)
                    instance.reativacao = reativacao
                    instance.quantidade = quantidade  # Define a quantidade conforme a contagem de linhas
                    instance.save()

                    # Se desejar, também realiza a reativação via API
                    token = self.obter_token_acesso()
                    if token:
                        self.reativar_equipamento(instance.ccid_equipamentos, token)

            return redirect('reativacao_id_iccid_adicionar')

        return render(request, 'reativacao_id_iccid_form.html', {
            'reativacao_form': reativacao_form,
            'id_iccid_formset': id_iccid_formset
        })

    def obter_token_acesso(self):
        url = "https://api.1nce.com/management-api/v1/oauth/token"
        headers = {
            "accept": "application/json",
            "content-type": "application/json"
        }
        data = {
            "grant_type": "password",
            "username": "seu_usuario",
            "password": "sua_senha"
        }
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            print("Token de acesso obtido com sucesso!")
            return response.json().get("access_token")
        else:
            print(f"Erro ao obter token de acesso: {response.status_code}")
            print(response.text)
            return None

    def reativar_equipamento(self, ccid_equipamentos, token):
        # Remove quebras de linha para evitar problemas em headers (caso o valor seja usado em e-mails, por exemplo)
        ccid_limpo = ccid_equipamentos.replace("\r", " ").replace("\n", " ")
        url = "https://api.1nce.com/management-api/v1/sims"
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "Authorization": f"Bearer {token}"
        }
        data = [
            {
                "imei_lock": False,
                "status": "Enabled",
                "ccid_equipamentos": ccid_limpo
            }
        ]
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            print("Equipamento reativado com sucesso!")
        else:
            print(f"Erro ao reativar equipamento: {response.status_code}")
            print(response.text)
from django.http import JsonResponse  # Importar JsonResponse se for necessário
from django.http import HttpResponseNotAllowed
from django.contrib.auth.decorators import permission_required, login_required
from django.contrib import messages

class RequisicoesListView(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Requisicoes
    template_name = 'requisicoes_list.html'
    context_object_name = 'requisicoes'
    paginate_by = 8
    permission_required = 'reativacao.view_reativacao'

    def get_queryset(self):
        queryset = super().get_queryset().order_by('-id')  # Ordenação decrescente dos IDs
        nome = self.request.GET.get('nome')
        status = self.request.GET.get('status')

        if nome:
            queryset = queryset.filter(nome__nome__icontains=nome)
        if status:
            queryset = queryset.filter(status=status)
        return queryset

    

from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from .models import Reativacao, IdIccid, Clientes
from .forms import IdIccidForm
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin


from django.core.paginator import Paginator

def _filtrar_reativacoes(request):
    """Aplica os filtros da querystring (cliente, status, motivo e período)
    e devolve a queryset já com os relacionamentos resolvidos.

    Compartilhado pela listagem e pela exportação Excel para que ambas
    respeitem exatamente os mesmos filtros.
    """
    reativacoes = (
        Reativacao.objects
        .select_related('nome', 'usuario')
        .prefetch_related('id_iccids')
        .order_by('-id')
    )

    cliente_filtro = request.GET.get('cliente_filtro')
    status_reativacao_filtro = request.GET.get('status_reativacao_filtro')
    motivo_reativacao_filtro = request.GET.get('motivo_filtro')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if cliente_filtro:
        reativacoes = reativacoes.filter(nome__id=cliente_filtro)

    if status_reativacao_filtro:
        reativacoes = reativacoes.filter(status_reativacao=status_reativacao_filtro)

    if motivo_reativacao_filtro:
        reativacoes = reativacoes.filter(motivo_reativacao=motivo_reativacao_filtro)

    if data_inicio:
        reativacoes = reativacoes.filter(data_hora_criacao__date__gte=data_inicio)

    if data_fim:
        reativacoes = reativacoes.filter(data_hora_criacao__date__lte=data_fim)

    return reativacoes


class ReativacaoListView(PermissionRequiredMixin, LoginRequiredMixin, View):
    permission_required = 'reativacao.view_reativacao'

    def get(self, request):
        reativacoes = _filtrar_reativacoes(request)
        status_reativacao_filtro = request.GET.get('status_reativacao_filtro')

        # Dropdown de clientes: quando há filtro de status, mostra só os clientes
        # que possuem reativação naquele status. 'reativacao_nome' é o related_name
        # real da FK 'nome' em Reativacao; .distinct() evita cliente repetido quando
        # ele tem várias reativações no mesmo status (o join gera uma linha por match).
        clientes_choices = Clientes.objects.all()
        if status_reativacao_filtro:
            clientes_choices = clientes_choices.filter(
                reativacao_nome__status_reativacao=status_reativacao_filtro
            ).distinct()

        return render(request, 'reativacao_list.html', {
            'reativacoes': reativacoes,  # queryset completa
            'clientes_choices': clientes_choices,
            'status_reativacao_choices': Reativacao.STATUS_CHOICES,
            'motivos_choices': Reativacao.MOTIVO_CHOICES,
        })


class ReativacaoExportExcelView(PermissionRequiredMixin, LoginRequiredMixin, View):
    permission_required = 'reativacao.view_reativacao'

    COLUNAS = [
        ('ID', 12),
        ('Usuário', 18),
        ('Data de Criação', 20),
        ('Nome', 30),
        ('CNPJ', 20),
        ('Motivo', 16),
        ('Canal de Solicitação', 22),
        ('Observações', 30),
        ('ID Equipamento', 30),
        ('ICCID', 30),
        ('Status da Reativação', 20),
        ('Quantidade', 14),
    ]

    def _linha(self, registro):
        id_iccids = list(registro.id_iccids.all())
        ids_equip = "\n".join(e.id_equipamentos for e in id_iccids)
        iccids = "\n".join(e.ccid_equipamentos for e in id_iccids)
        quantidade = sum(e.quantidade or 0 for e in id_iccids)
        data_criacao = (
            registro.data_hora_criacao.strftime('%d/%m/%Y %H:%M')
            if registro.data_hora_criacao else ''
        )
        return [
            registro.id,
            registro.usuario.username if registro.usuario else '',
            data_criacao,
            str(registro.nome) if registro.nome else '',
            registro.cnpj or '',
            registro.motivo_reativacao or '',
            registro.canal_solicitacao or '',
            registro.observacoes or '',
            ids_equip,
            iccids,
            registro.status_reativacao or '',
            quantidade,
        ]

    def get(self, request):
        reativacoes = _filtrar_reativacoes(request)

        wb = Workbook()
        ws = wb.active
        ws.title = "Reativações"

        header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.append([label for label, _ in self.COLUNAS])
        for col_idx in range(1, len(self.COLUNAS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for registro in reativacoes:
            ws.append(self._linha(registro))
            for col_idx in range(1, len(self.COLUNAS) + 1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border

        for i, (_, width) in enumerate(self.COLUNAS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reativacoes.xlsx"'
        return response


def update_status(request):
    if request.method == 'POST':
        id_iccid = request.POST.get('id_iccid')
        novo_status = request.POST.get('status')
        
        # Atualiza o status no banco de dados
        reativacao = Reativacao.objects.get(id=id_iccid)
        reativacao.status_reativacao = novo_status
        reativacao.save()

        return JsonResponse({'success': True, 'status': novo_status})

    return JsonResponse({'success': False, 'error': 'Método não permitido'})

    


@method_decorator(login_required, name='dispatch')
class ReativacaoView( PermissionRequiredMixin,LoginRequiredMixin, View):
    permission_required = 'reativacao.view_reativacao'
    paginate_by = 5
    def post(self, request):
        id_iccid = get_object_or_404(IdIccid, pk=request.POST.get('id_iccid'))
        status = request.POST.get('status_reativacao')
        if status:
            id_iccid.status = status
            id_iccid.save()
            return JsonResponse({'success': True, 'status': id_iccid.status})
        return JsonResponse({'success': False}, status=400) # Substitua 'reativacao' pelo nome do seu aplicativo

    def get(self, request):
        return render(request, 'reativacao.html')

from django.shortcuts import get_object_or_404
from django.http import FileResponse, HttpResponse
from .pdf_utils import gerar_pdf_id_iccid
from .models import IdIccid
import os


class DownloadPdfView(View):

    def get(self, request, id_iccid):
        id_iccid_obj = get_object_or_404(IdIccid, pk=id_iccid)
        pdf_path = gerar_pdf_id_iccid(id_iccid_obj)
        if os.path.exists(pdf_path):
            return FileResponse(open(pdf_path, 'rb'), content_type='application/pdf')
        else:
            # Handle the case where the PDF was not generated successfully
            return HttpResponse("Erro ao gerar o PDF", status=500)

class ReativacaoUpdateView(UpdateView):
    model = Reativacao
    form_class = ReativacaoForm
    template_name = 'reativacao_update.html'
    success_url = reverse_lazy('reativacao_list')

class ReativacaoCompleteUpdateView(PermissionRequiredMixin, LoginRequiredMixin, View):
    permission_required = 'reativacao.change_reativacao'
    
    def get(self, request, pk):
        reativacao = get_object_or_404(Reativacao, pk=pk)
        form = ReativacaoForm(instance=reativacao)
        formset = IdIccidFormSet(instance=reativacao)
        return render(request, 'reativacao_complete_update.html', {
            'form': form,
            'formset': formset,
            'reativacao': reativacao,
        })


@login_required
@permission_required('reativacao.delete_reativacao', raise_exception=True)
def delete_reativacao(request, pk):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    reativacao = get_object_or_404(Reativacao, pk=pk)
    reativacao.delete()
    messages.success(request, 'Reativação excluída com sucesso.')
    return redirect('reativacao_list')

    def post(self, request, pk):
        reativacao = get_object_or_404(Reativacao, pk=pk)
        form = ReativacaoForm(request.POST, instance=reativacao)
        formset = IdIccidFormSet(request.POST, instance=reativacao)
        logger = logging.getLogger(__name__)

        is_form_valid = form.is_valid()
        is_formset_valid = formset.is_valid()

        if is_form_valid and is_formset_valid:
            form.save()
            formset.save()
            messages.success(request, 'Reativação atualizada com sucesso.')
            return redirect('reativacao_list')

        # Log detailed errors for debugging
        try:
            form_errors = form.errors.get_json_data()
        except Exception:
            form_errors = str(form.errors)
        try:
            formset_errors = [f.errors.get_json_data() for f in formset.forms]
        except Exception:
            formset_errors = [str(f.errors) for f in formset.forms]

        non_formset_errors = formset.non_form_errors()

        logger.error('ReativacaoCompleteUpdateView: form valid=%s, formset valid=%s', is_form_valid, is_formset_valid)
        logger.error('Form errors: %s', json.dumps(form_errors, ensure_ascii=False))
        logger.error('Formset errors: %s', json.dumps(formset_errors, ensure_ascii=False))
        logger.error('Formset non-form errors: %s', non_formset_errors)

        # Also show a short message to the user
        messages.error(request, 'Erro ao salvar. Verifique os campos destacados abaixo.')

        return render(request, 'reativacao_complete_update.html', {
            'form': form,
            'formset': formset,
            'reativacao': reativacao,
        })
        
