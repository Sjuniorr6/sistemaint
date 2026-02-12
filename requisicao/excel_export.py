from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Requisicoes


def gerar_excel_requisicoes(queryset):
    """
    Gera um arquivo Excel com os dados das requisições.
    
    Colunas:
    - Protocolo (ID)
    - Nome do cliente
    - CNPJ
    - Endereço
    - Contrato
    - Data
    - Motivo
    - Taxa de Envio
    - Comercial
    - Tipo de Produto
    - Quantidade
    - Carregador
    - Cabo
    - Envio
    - Valor Unitário
    - Valor Total
    - Observações
    """
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Requisições"
    
    # Definir estilos
    header_fill = PatternFill(start_color="665b1d", end_color="665b1d", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Colunas do Excel
    colunas = [
        "Protocolo",
        "Nome do Cliente",
        "CNPJ",
        "Endereço",
        "Contrato",
        "Data",
        "Motivo",
        "Taxa de Envio",
        "Comercial",
        "Tipo de Produto",
        "Quantidade",
        "Carregador",
        "Cabo",
        "Envio",
        "Valor Unitário",
        "Valor Total",
        "Observações"
    ]
    
    # Adicionar cabeçalho
    for col_num, coluna in enumerate(colunas, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = coluna
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Adicionar dados
    for row_num, requisicao in enumerate(queryset, 2):
        dados = [
            requisicao.id,  # Protocolo
            requisicao.nome.nome if requisicao.nome else "",  # Nome do Cliente
            requisicao.cnpj if requisicao.cnpj else "",  # CNPJ
            requisicao.nome.endereco if requisicao.nome else "",  # Endereço
            requisicao.contrato if requisicao.contrato else "",  # Contrato
            requisicao.data_alteracao.strftime("%d/%m/%Y %H:%M") if requisicao.data_alteracao else "",  # Data
            requisicao.motivo if requisicao.motivo else "",  # Motivo
            f"R$ {requisicao.taxa_envio:.2f}".replace(".", ",") if requisicao.taxa_envio else "R$ 0,00",  # Taxa de Envio
            requisicao.comercial if requisicao.comercial else "",  # Comercial
            requisicao.tipo_produto.nome if requisicao.tipo_produto else "",  # Tipo de Produto
            requisicao.numero_de_equipamentos if requisicao.numero_de_equipamentos else "",  # Quantidade
            requisicao.carregador if requisicao.carregador else "",  # Carregador
            requisicao.cabo if requisicao.cabo else "",  # Cabo
            requisicao.envio if requisicao.envio else "",  # Envio
            f"R$ {requisicao.valor_unitario:.2f}".replace(".", ",") if requisicao.valor_unitario else "R$ 0,00",  # Valor Unitário
            f"R$ {requisicao.valor_total:.2f}".replace(".", ",") if requisicao.valor_total else "R$ 0,00",  # Valor Total
            requisicao.observacoes if requisicao.observacoes else ""  # Observações
        ]
        
        for col_num, valor in enumerate(dados, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = valor
            cell.border = border
            
            # Alinhar números e valores monetários à direita
            if col_num in [1, 7, 11, 15, 16]:  # ID, Taxa, Quantidade, Valor Unitário, Valor Total
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar largura das colunas
    column_widths = [12, 25, 15, 35, 15, 20, 20, 15, 12, 25, 12, 15, 15, 15, 15, 15, 40]
    for col_num, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = width
    
    # Congelar primeira linha
    worksheet.freeze_panes = "A2"
    
    return workbook
