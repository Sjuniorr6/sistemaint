from typing import Any
from django.views.generic import (
    ListView,
    CreateView,
    DetailView,
    UpdateView,
    DeleteView,
)
from . import models, forms
from .forms import EstoqueantenistarForm
from django.urls import reverse_lazy
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from registrodemanutencao.models import registrodemanutencao
from requisicao.models import (
    Requisicoes,
    estoque_antenista,
    KanbanAuditLog,
    AuditLog,
    CampoAlterado,
)
from django.shortcuts import get_object_or_404
from django.db.models.signals import post_save, pre_save
from django.db import transaction
import contextlib
from django.http import HttpResponseRedirect
from django.db.models import Q
from django.conf import settings
from django.db import transaction
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.contrib.contenttypes.models import ContentType
import json
from django.views.generic import TemplateView
from registrodemanutencao.forms import FormulariosUpdateForm
from django.core.mail import send_mail
from django.contrib.auth.mixins import PermissionRequiredMixin, LoginRequiredMixin
from django.shortcuts import render, redirect
from django.forms import inlineformset_factory
from .forms import ControleForm
from franquia.models import registrodefranquia
from decimal import Decimal, InvalidOperation

# ------------------------------------------------------
class RequisicoesViews(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Requisicoes
    template_name = "requisicoes.html"
    context_object_name = "requisicoes"
    paginate_by = 12
    permission_required = "requisicao.view_requisicoes"

    def get_queryset(self):
        return Requisicoes.objects.filter(status__in=["Pendente"])


from django.http import JsonResponse
from .models import Clientes


def get_cliente_data(request, cliente_id):
    cliente = Clientes.objects.get(id=cliente_id)
    data = {
        "cnpj": cliente.cnpj,
        "inicio_de_contrato": cliente.inicio_de_contrato,
        "vigencia": cliente.vigencia,
        "contrato": cliente.tipo_contrato,
    }
    return JsonResponse(data)


import logging
from django.contrib import messages
from django.db import transaction
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from .models import Requisicoes, estoque_antenista
from .forms import RequisicaoForm

logger = logging.getLogger(__name__)


class RequisicaoCreateView(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Requisicoes
    template_name = "requisicao_create.html"
    form_class = RequisicaoForm
    success_url = reverse_lazy("requisicoes")
    permission_required = "requisicao.add_requisicoes"

    def get_queryset(self):
        return Requisicoes.objects.all().order_by("id")

    def form_valid(self, form):
        motivo = form.cleaned_data.get("motivo")
        antenista = form.cleaned_data.get("antenista")
        tipo_produto = form.cleaned_data.get("tipo_produto")
        numero_de_equipamentos = form.cleaned_data.get("numero_de_equipamentos")

        logger.info("Formulário válido: %s", form.is_valid())
        logger.info("Dados do formulário: %s", form.cleaned_data)

        if (
            motivo in ["Isca FAST", "Estoque Antenista"]
            and antenista
            and tipo_produto
            and numero_de_equipamentos
        ):
            try:
                with transaction.atomic():
                    requisicao = form.save(commit=False)
                    quantidade_requisitada = int(numero_de_equipamentos)
                    antenista_estoque, created = (
                        estoque_antenista.objects.get_or_create(
                            nome=antenista,
                            tipo_produto=tipo_produto,
                            defaults={"quantidade": 0},
                        )
                    )

                    if antenista_estoque.quantidade is None:
                        antenista_estoque.quantidade = 0

                    if motivo == "Isca FAST":
                        if antenista_estoque.quantidade >= quantidade_requisitada:
                            antenista_estoque.quantidade -= quantidade_requisitada
                        else:
                            messages.error(
                                self.request,
                                f"O antenista {antenista} não tem quantidade suficiente no estoque para o produto {tipo_produto}. Quantidade disponível: {antenista_estoque.quantidade}, quantidade requisitada: {quantidade_requisitada}.",
                            )
                            return self.form_invalid(form)
                    elif motivo == "Estoque Antenista":
                        antenista_estoque.quantidade += quantidade_requisitada

                    antenista_estoque.save()
                    requisicao.save()
                    return super().form_valid(form)
            except Exception as e:
                logger.error("Erro ao processar a requisição: %s", e)
                messages.error(
                    self.request, "Ocorreu um erro ao processar a requisição."
                )
                return self.form_invalid(form)
        else:
            response = super().form_valid(form)

            # Registrar log de criação
            AuditLog.registrar(
                objeto=form.instance,
                acao="criacao",
                usuario=self.request.user,
                status_novo=form.instance.status,
                detalhes={
                    "cliente": str(form.instance.nome) if form.instance.nome else None,
                    "tipo_produto": (
                        str(form.instance.tipo_produto)
                        if form.instance.tipo_produto
                        else None
                    ),
                    "quantidade": form.instance.numero_de_equipamentos,
                },
                observacao="Requisição criada",
                request=self.request,
            )

            return response


class RequisicaoDetailView(PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = models.Requisicoes
    template_name = "requisicao_detail.html"
    permission_required = "requisicao.view_requisicoes"


class RequisicaoUpdateView(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Requisicoes
    form_class = forms.RequisicaoForm
    template_name = "requisicao_update.html"
    context_object_name = "requisicao"
    success_url = reverse_lazy("requisicao_list")
    permission_required = "requisicao.change_requisicoes"

    def form_valid(self, form):
        # Capturar campos alterados
        if form.changed_data:
            campos_alterados = []
            for field_name in form.changed_data:
                campo_dict = {
                    "campo": field_name,
                    "anterior": getattr(self.get_object(), field_name),
                    "novo": form.cleaned_data[field_name],
                }
                campos_alterados.append(campo_dict)

        response = super().form_valid(form)

        # Registrar log de edição
        if form.changed_data:
            log = AuditLog.registrar(
                objeto=self.object,
                acao="edicao",
                usuario=self.request.user,
                detalhes={"total_campos_alterados": len(form.changed_data)},
                observacao=f"{len(form.changed_data)} campo(s) alterado(s)",
                request=self.request,
            )

            # Registrar cada campo alterado
            for campo in campos_alterados:
                CampoAlterado.objects.create(
                    audit_log=log,
                    nome_campo=campo["campo"],
                    valor_anterior=str(campo.get("anterior", "")),
                    valor_novo=str(campo.get("novo", "")),
                )

        return response


class Requisicao2UpdateView(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Requisicoes
    form_class = forms.RequisicaoForm
    template_name = "requisicao_update.html"
    context_object_name = "requisicao"
    success_url = reverse_lazy("requisicao_list")
    permission_required = "requisicao.change_requisicoes"


class requisicoesDeleteView(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Requisicoes
    template_name = "requisicao_delete.html"
    success_url = reverse_lazy("expedicaoListViews")
    permission_required = "requisicao.delete_requisicoes"

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Registrar log ANTES de deletar
        AuditLog.registrar(
            objeto=self.object,
            acao="exclusao",
            usuario=request.user,
            detalhes={
                "id_excluido": self.object.id,
                "dados_basicos": {
                    "status": self.object.status,
                    "cliente": str(self.object.nome) if self.object.nome else None,
                    "comercial": (
                        str(self.object.comercial) if self.object.comercial else None
                    ),
                    "tipo_produto": (
                        str(self.object.tipo_produto)
                        if self.object.tipo_produto
                        else None
                    ),
                    "quantidade": self.object.numero_de_equipamentos,
                },
            },
            observacao=f"Requisição #{self.object.id} excluída",
            request=request,
        )

        success_url = self.get_success_url()
        self.object.delete()
        return HttpResponseRedirect(success_url)

# ------------------------------------------------------


# ------------------------------------------------------
class configuracaodeleteview(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = models.Requisicoes
    template_name = "configuracao_delete.html"
    success_url = reverse_lazy("acompanhamento_requisicao")
    permission_required = "requisicao.delete_requisicoes"


from django.forms import modelformset_factory
from .models import ControleModel


class ControleModel(PermissionRequiredMixin, LoginRequiredMixin, CreateView):

    model = ControleModel
    form_class = ControleForm
    template_name = "t42_form.html"
    success_url = reverse_lazy("t42_view")
    permission_required = "t42.add_t42model"

    def form_valid(self, form):
        response = super().form_valid(form)
        total_forms = int(self.request.POST.get("total_forms", 0))

        for i in range(total_forms):
            id_equipamento = self.request.POST.get(f"id_equipamento-{i}-id_equipamento")
            iccid_equipamento = self.request.POST.get(
                f"iccid_equipamento-{i}-iccid_equipamento"
            )

            if id_equipamento and iccid_equipamento:
                ControleModel.objects.create(
                    nome=self.object.nome,
                    id_equipamento=id_equipamento,
                    iccid_equipamento=iccid_equipamento,
                )

        return response


from django.views.generic import CreateView, ListView
from django.contrib.auth.mixins import PermissionRequiredMixin, LoginRequiredMixin
from django.urls import reverse_lazy
from .models import ControleModel
from .forms import ControleForm


class ControleCreateView(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = ControleModel
    template_name = "controle.html"
    form_class = ControleForm
    success_url = reverse_lazy("controleList")
    permission_required = "requisicao.view_requisicoes"

    def form_valid(self, form):
        form.instance.usuario = self.request.user
        return super().form_valid(form)


class ControleListView(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = ControleModel
    template_name = "controle_list.html"
    context_object_name = "controles"
    permission_required = "requisicao.view_requisicoes"


class ConfiguracaoListView(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    template_name = "configuracao_list.html"
    context_object_name = "equipamentos"
    permission_required = "requisicao.view_requisicoes"

    def get_queryset(self):
        # Obter parâmetro de filtro por ID
        id_filtro = self.request.GET.get("id_filtro")

        requisicoes_queryset = Requisicoes.objects.filter(
            status__in=["Aprovado pelo CEO"]
        ).exclude(
            tipo_produto__nome__in=[
                "GS310",
                "GS340",
                "GS390",
                "GS8310 (4G)",
            ]
        )
        manutencao_queryset = registrodemanutencao.objects.filter(
            status__in=[
                "Aprovado Inteligência",
                "Aprovado pela Diretoria",
                "Aprovado pelo CEO",
            ]
        ).exclude(
            tipo_produto__nome__in=[
                "GS310",
                "GS340",
                "GS390",
                "GS8310 (4G)",
            ]
        )

        # Aplicar filtro por ID se fornecido
        if id_filtro:
            try:
                id_valor = int(id_filtro)
                requisicoes_queryset = requisicoes_queryset.filter(id=id_valor)
                manutencao_queryset = manutencao_queryset.filter(id=id_valor)
            except ValueError:
                # Se o ID não for um número válido, retornar queryset vazio
                return []

        # Combine os querysets
        combined_queryset = list(requisicoes_queryset) + list(manutencao_queryset)

        return combined_queryset


class ConfiguracaoUpdateView(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Requisicoes
    form_class = forms.requisicaoFormup
    template_name = "configuracao_update.html"
    context_object_name = "equipamento"
    success_url = reverse_lazy("ConfiguracaoListView")
    permission_required = "requisicao.change_requisicoes"


class ConfiguracaoUpdateView2(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = registrodemanutencao
    form_class = FormulariosUpdateForm
    template_name = "configuracao_update.html"
    context_object_name = "equipamento"
    success_url = reverse_lazy("ConfiguracaoListView")
    permission_required = "requisicao.change_requisicoes"


# ------------------------------------------------------


class tecnicoListView(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    template_name = "setor_tecnico.html"
    context_object_name = "equipamentos"
    permission_required = "requisicao.view_requisicoes"

    def get_queryset(self):
        valores_tipo_produto = [
            "GS310",
            "GS340",
            "GS390",
            "GS8310 (4G)",
            "PLUG AND PLAY",
        ]
        requisicao_queryset = Requisicoes.objects.filter(
            tipo_produto__nome__in=valores_tipo_produto
        )
        return requisicao_queryset


class tecnicoUpdateView(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Requisicoes
    form_class = forms.requisicaoFormup
    template_name = "setor_tecnico_update.html"
    context_object_name = "equipamento"
    success_url = reverse_lazy("tecnicoListView")
    permission_required = "requisicao.change_requisicoes"

    def form_valid(self, form):
        # Salvar o formulário primeiro
        response = super().form_valid(form)

        # Obter os valores dos campos
        id_equipamentos = form.cleaned_data.get("id_equipamentos", "")
        iccid = form.cleaned_data.get("iccid", "")

        if id_equipamentos and iccid:
            try:
                # Separar os valores por espaços (qualquer quantidade de espaços)
                ids = [id.strip() for id in id_equipamentos.split() if id.strip()]
                iccids = [
                    iccid_val.strip()
                    for iccid_val in iccid.split()
                    if iccid_val.strip()
                ]

                # Vincular os valores correspondentes
                for i, (equip_id, iccid_val) in enumerate(zip(ids, iccids), 1):
                    if i <= 10:  # Limite de 10 equipamentos conforme o modelo
                        field_name = f"iccid_equipamento{i}"
                        if hasattr(self.object, field_name):
                            setattr(self.object, field_name, iccid_val)

                # Salvar as alterações
                self.object.save()

                # Log de sucesso
                print(
                    f"Vinculação automática realizada: {len(ids)} equipamentos vinculados com ICCIDs"
                )

            except Exception as e:
                print(f"Erro ao vincular equipamentos com ICCIDs: {e}")

        return response


# ------------------------------------------------------


from datetime import timedelta
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import ListView
from .models import (
    Requisicoes,
)  # Certifique-se de que o modelo Requisicoes está importado


class ceoListViews(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Requisicoes
    template_name = "ceo_list.html"
    context_object_name = "ceo_list"
    paginate_by = 10
    permission_required = "requisicao.view_requisicoes"

    def get_queryset(self):
        return Requisicoes.objects.filter(
            status__in=["Pendente", "Aprovado pela Diretoria"]
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["total_pendente"] = Requisicoes.objects.filter(
            status="Pendente"
        ).count()
        context["total_aprovado_ceo"] = Requisicoes.objects.filter(
            status="Aprovado pelo CEO"
        ).count()
        context["total_configurado"] = Requisicoes.objects.filter(
            status="Configurado"
        ).count()

        # Verificando requisições que não foram alteradas nas últimas 24 horas, excluindo 'Enviado para o Cliente'
        threshold_time = timezone.now() - timedelta(hours=24)
        context["requisições_sem_alteracao"] = Requisicoes.objects.filter(
            data_alteracao__lt=threshold_time
        ).exclude(status="Enviado para o Cliente")

        # Contando requisições sem alteração
        context["count_requisicoes_sem_alteracao"] = context[
            "requisições_sem_alteracao"
        ].count()

        # Verificando requisições com data_entrega a menos de 48 horas
        threshold_delivery_time = timezone.now() + timedelta(hours=48)
        context["requisições_proximas_entrega"] = Requisicoes.objects.filter(
            data_entrega__lte=threshold_delivery_time
        )

        # Contando requisições próximas da entrega
        context["count_requisicoes_proximas_entrega"] = context[
            "requisições_proximas_entrega"
        ].count()

        # Adicionando dias restantes até a entrega
        for item in context["ceo_list"]:
            if item.data_entrega:
                dias_restantes = (item.data_entrega - timezone.now().date()).days
                item.dias_restantes = (
                    dias_restantes  # Armazena a contagem de dias restantes no objeto
                )
                item.dias_restantes_inclusivo = (
                    dias_restantes + 1
                )  # Adiciona 1 e armazena

        return context


class ceodetailview(PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = Requisicoes
    template_name = "ceo_detail.html"
    permission_required = "requisicao.view_requisicoes"


class CeoEntradaDetailView(PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = registrodemanutencao
    template_name = "ceo_detalheentrada.html"
    context_object_name = "manutencoes"
    permission_required = "requisicao.view_requisicoes"


# ------------------------------------------------------


# ------------------------------------------------------


class diretoriaListViews(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    template_name = "diretoria_list.html"
    context_object_name = "diretoria_list"
    permission_required = "requisicao.view_requisicoes"

    def get_queryset(self):
        requisicoes_queryset = Requisicoes.objects.filter(status__in=["", ""])
        manutencao_queryset = registrodemanutencao.objects.filter(status="Manutenção")

        # Combine os querysets
        combined_queryset = list(requisicoes_queryset) + list(manutencao_queryset)

        return combined_queryset


# ------------------------------------------------------


from django.views.generic import ListView
from formacompanhamento.models import (
    Formacompanhamento,
)  # Substitua 'Financeiro' pelo nome do seu modelo


class FinanceiroListViews(ListView):
    template_name = "financeiro_list.html"  # Substitua pelo nome do seu template
    context_object_name = "financeiro_list"
    paginate_by = 10

    def get_queryset(self):
        return Formacompanhamento.objects.all()  # Defina o queryset aqui


# ------------------------------------------------------
# class expedicaoListViews(PermissionRequiredMixin, LoginRequiredMixin, ListView):
#     model = Requisicoes
#     template_name = "expedicao_list.html"
#     context_object_name = "expedicao_list"
#     permission_required = "requisicao.view_requisicoes"

#     def get_queryset(self):
#         nome = self.request.GET.get("nome")

#         # Primeiro, busca as requisições
#         requisicoes_queryset = Requisicoes.objects.filter(status__in=["Configurado"])
#         if nome:
#             requisicoes_queryset = requisicoes_queryset.filter(
#                 nome__nome__icontains=nome
#             )

#         # Depois, busca as manutenções
#         manutencao_queryset = registrodemanutencao.objects.filter(
#             status__in=["Configurado"]
#         )
#         if nome:
#             manutencao_queryset = manutencao_queryset.filter(nome__nome__icontains=nome)

#         # Combina os querysets
#         combined_queryset = list(requisicoes_queryset) + list(manutencao_queryset)

#         return combined_queryset


class expedicaoListViews(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Requisicoes
    template_name = "expedicao_list.html"
    context_object_name = "expedicao_list"
    permission_required = "requisicao.view_requisicoes"

    def get_queryset(self):
        nome = self.request.GET.get("nome")
        id_filtro = self.request.GET.get("id_filtro")

        # Primeiro, busca as requisições
        requisicoes_queryset = Requisicoes.objects.filter(status__in=["Configurado"])
        if nome:
            requisicoes_queryset = requisicoes_queryset.filter(
                nome__nome__icontains=nome
            )

        if id_filtro:
            try:
                id_valor = int(id_filtro)
                requisicoes_queryset = requisicoes_queryset.filter(id=id_valor)
            except ValueError:
                # Se o ID não for um número válido, retornar requisicoes_queryset vazio
                return Requisicoes.objects.none()

        # Depois, busca as manutenções
        manutencao_queryset = registrodemanutencao.objects.filter(
            status__in=["Configurado"]
        )

        if nome:
            manutencao_queryset = manutencao_queryset.filter(nome__nome__icontains=nome)

        if id_filtro:
            try:
                id_valor = int(id_filtro)
                manutencao_queryset = manutencao_queryset.filter(id=id_valor)
            except ValueError:
                # Se o ID não for um número válido, retornar manutencao_queryset vazio
                return Requisicoes.objects.none()

        # Combina os querysets
        combined_queryset = list(requisicoes_queryset) + list(manutencao_queryset)

        return combined_queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["id_filter"] = self.request.GET.get("id_filtro", "")
        context["nome_filter"] = self.request.GET.get("nome", "")
        return context


# ------------------------------------------------------
# ------------------------------------------------------
class historicoListView(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Requisicoes
    template_name = "historico_list.html"
    context_object_name = "historico_list"
    paginate_by = 12
    permission_required = "requisicao.view_requisicoes"

    def get_queryset(self):
        queryset = Requisicoes.objects.all().order_by("-id")
        nome = self.request.GET.get("nome")
        status = self.request.GET.get("status")
        id_filtro = self.request.GET.get("id_filtro")

        if nome:
            queryset = queryset.filter(nome__nome__icontains=nome)

        if status:
            queryset = queryset.filter(status__icontains=status)

        if id_filtro:
            try:
                id_valor = int(id_filtro)
                queryset = queryset.filter(id=id_valor)
            except ValueError:
                # Se o ID não for um número válido, retornar queryset vazio
                return Requisicoes.objects.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Preserva os parâmetros GET para a paginação
        context["id_filter"] = self.request.GET.get("id_filtro", "")
        context["nome_filter"] = self.request.GET.get("nome", "")
        context["status_filter"] = self.request.GET.get("status", "")
        return context


# ------------------------------------------------------
# ------------------------------------------------------


# ------------------------------------------------------
def aprovar_requisicao(request, id):
    registro = get_object_or_404(Requisicoes, id=id)
    registro.status = "Aprovado pela Diretoria"
    registro.save()
    return redirect("#")


def reprovar_requisicao(request, id):
    registro = get_object_or_404(Requisicoes, id=id)
    registro.status = "Reprovado pela Diretoria"
    registro.save()
    return redirect("#")


# ------------------------------------------------------


def aprovar_FINANCEIRO(request, id):
    registro = get_object_or_404(Formacompanhamento, id=id)
    registro.status = "PAGO"
    registro.save()
    return redirect("FinanceiroListViews")


# ------------------------------------------------------
def reprovar_ceo(request, id):
    registro = get_object_or_404(Requisicoes, id=id)
    status_anterior = registro.status
    registro.status = "Reprovado pelo CEO"
    registro.save()

    # Registrar log de auditoria
    AuditLog.registrar(
        objeto=registro,
        acao="reprovacao",
        usuario=request.user,
        status_anterior=status_anterior,
        status_novo="Reprovado pelo CEO",
        observacao="Reprovado pelo CEO",
        request=request,
    )

    return redirect("ceoListViews")


def aprovar_ceo(request, id):
    registro = get_object_or_404(Requisicoes, id=id)
    status_anterior = registro.status
    registro.status = "Aprovado pelo CEO"
    registro.save()

    # Registrar log de auditoria
    AuditLog.registrar(
        objeto=registro,
        acao="aprovacao",
        usuario=request.user,
        status_anterior=status_anterior,
        status_novo="Aprovado pelo CEO",
        observacao="Aprovado pelo CEO",
        request=request,
    )

    subject = f"Requisicao Aprovada: {registro.id}"
    message = f"A manutenção {registro.id} foi aprovada com sucesso. {registro.nome} Status: {registro.status} criar Requisição"
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = ["sjuniorr6@gmail.com"]

    try:
        send_mail(subject, message, from_email, recipient_list)
        print("Email enviado com sucesso.")
    except Exception as e:
        print(f"Erro ao enviar email: {e}")

    return redirect("ceoListViews")


# ------------------------------------------------------


# ------------------------------------------------------
from django.http import HttpResponse


def configurado_expedicao(request, id):
    registro = get_object_or_404(Requisicoes, id=id)
    registro.status = "Configurado"
    registro.save()
    return redirect("ConfiguracaoListView")


def expedicao_expedido(request, id):
    registro = get_object_or_404(Requisicoes, id=id)
    status_anterior = registro.status
    registro.status = "Enviado para o Cliente"
    registro.save()

    # Registrar log de auditoria
    AuditLog.registrar(
        objeto=registro,
        acao="envio_cliente",
        usuario=request.user,
        status_anterior=status_anterior,
        status_novo="Enviado para o Cliente",
        observacao="Expedição finalizada - enviado ao cliente",
        request=request,
    )

    # Enviar e-mail quando status for 'Enviado para o Cliente'
    subject = f"Requisição Expedida - ID: {registro.id}"

    # Corpo do e-mail melhorado
    message = f"""
Prezados,

A requisição ID: {registro.id} foi expedida com sucesso e o pedido foi enviado ao cliente.

Informações da Requisição:
- Cliente: {registro.nome}
- Tipo de Produto: {registro.tipo_produto}
- Quantidade: {registro.numero_de_equipamentos}
- Comercial Responsável: {registro.comercial}

Qualquer dúvida que possa surgir, favor solicitar informações à recepção.

Atenciosamente,
Departamento de Inteligência
Golden Sat
    """

    from_email = settings.DEFAULT_FROM_EMAIL

    # E-mails obrigatórios
    recipient_list = [
        "quality@grupogoldensat.com.br",
        "comercial@grupogoldensat.com.br",
        "faturamento@grupogoldensat.com.br",
        "inteligencia@grupogoldensat.com.br",
    ]

    # Mapear comerciais específicos para seus e-mails (case-insensitive)
    comercial_emails = {
        "mayra": "mayra.monteiro@grupogoldensat.com.br",
        "aparecido": "comercial2@grupogoldensat.com.br",
        "marcio": "diretoria@grupogoldensat.com.br",
        "daniel": "superintendente@grupogoldensat.com.br",
    }

    # Adicionar e-mail do comercial se estiver na lista (comparação case-insensitive)
    if registro.comercial:
        comercial_nome = (
            str(registro.comercial).strip().lower()
        )  # Converte para minúsculo
        if comercial_nome in comercial_emails:
            email_comercial = comercial_emails[comercial_nome]
            recipient_list.append(email_comercial)
            print(
                f"E-mail do comercial {registro.comercial} adicionado: {email_comercial}"
            )

    try:
        send_mail(subject, message, from_email, recipient_list)
        print(f"Email enviado com sucesso para: {', '.join(recipient_list)}")
    except Exception as e:
        print(f"Erro ao enviar email: {e}")

    return redirect("expedicaoListViews")


def expedicao_expedido2(request, id):
    registro = get_object_or_404(registrodemanutencao, id=id)
    status_anterior = registro.status
    registro.status = "Enviado para o Cliente"
    registro.save()

    # Registrar log de auditoria
    AuditLog.registrar(
        objeto=registro,
        acao="manutencao_expedicao",
        usuario=request.user,
        status_anterior=status_anterior,
        status_novo="Enviado para o Cliente",
        observacao="Manutenção finalizada - enviado ao cliente",
        request=request,
    )

    return redirect("expedicaoListViews")


def expedir_requisicao(request, id):
    registro = get_object_or_404(Requisicoes, id=id)

    if request.method == "POST":
        # Captura dados do checklist de auditoria
        ids_auditados = request.POST.get("ids_auditados", "").strip()
        memoria_apagada = request.POST.get("memoria_apagada") == "on"
        verificacao_tp = request.POST.get("verificacao_tp") == "on"
        verificacao_plataforma = request.POST.get("verificacao_plataforma") == "on"
        customizacao_conforme = request.POST.get("customizacao_conforme") == "on"

        status_anterior = registro.status

        # Salva informações do checklist
        registro.ids_auditados = ids_auditados
        registro.memoria_apagada = memoria_apagada
        registro.verificacao_tp = verificacao_tp
        registro.verificacao_plataforma = verificacao_plataforma
        registro.customizacao_conforme = customizacao_conforme

        # Alterar o status do registro para "Configurado"
        registro.status = "Configurado"
        registro.save()

        # Registrar log de auditoria
        AuditLog.registrar(
            objeto=registro,
            acao="expedicao",
            usuario=request.user,
            status_anterior=status_anterior,
            status_novo="Configurado",
            detalhes={
                "tipo_expedicao": "total",
                "ids_auditados": ids_auditados,
                "memoria_apagada": memoria_apagada,
                "verificacao_tp": verificacao_tp,
                "verificacao_plataforma": verificacao_plataforma,
                "customizacao_conforme": customizacao_conforme,
                "quantidade": registro.numero_de_equipamentos,
            },
            observacao="Expedição total realizada com checklist de auditoria",
            request=request,
        )

        return redirect("kanban_gestao")

    # GET: renderiza template com modal de checklist
    return render(request, "expedir_confirmacao.html", {"requisicao": registro})


def expedir_requisicaotec(request, id):
    registro = get_object_or_404(Requisicoes, id=id)
    status_anterior = registro.status
    # Alterar o status do registro para "Configurado"
    registro.status = "Configurado"
    registro.save()

    # Registrar log de auditoria
    AuditLog.registrar(
        objeto=registro,
        acao="expedicao",
        usuario=request.user,
        status_anterior=status_anterior,
        status_novo="Configurado",
        detalhes={"tipo_expedicao": "tecnico"},
        observacao="Expedição realizada pelo técnico",
        request=request,
    )

    return redirect("tecnicoListView")


def expedir_manutencao(request, id):
    registro = get_object_or_404(registrodemanutencao, id=id)
    status_anterior = registro.status
    # Alterar o status do registro para "Configurado"
    registro.status = "Configurado"
    registro.save()

    # Registrar log de auditoria
    AuditLog.registrar(
        objeto=registro,
        acao="manutencao_expedicao",
        usuario=request.user,
        status_anterior=status_anterior,
        status_novo="Configurado",
        observacao="Manutenção expedida",
        request=request,
    )

    return redirect("ConfiguracaoListView")


def configurado_manutencao(request, id):
    registro = get_object_or_404(registrodemanutencao, id=id)
    status_anterior = registro.status
    registro.status = "Configurado"
    registro.save()

    # Registrar log de auditoria
    AuditLog.registrar(
        objeto=registro,
        acao="manutencao_status",
        usuario=request.user,
        status_anterior=status_anterior,
        status_novo="Configurado",
        observacao="Status alterado para Configurado",
        request=request,
    )

    return redirect("ConfiguracaoListView")


def expedicao_expedido_manutencao(request, id):
    registro = get_object_or_404(registrodemanutencao, id=id)
    registro.status = "Enviado para o Cliente"
    registro.save()
    return redirect("expedicaoListViews")


# ------------------------------------------------------
# View para aprovar uma requisição pela diretoria
# ------------------------------------------------------
def Reprovar_diretoria(request, id):
    registro = get_object_or_404(registrodemanutencao, id=id)
    registro.status = "Reprovado pela Diretoria"
    registro.save()
    return redirect("diretoriaListViews")


# View para reprovar uma requisição pela diretoria
def Aprovar_diretoria(request, id):
    registro = get_object_or_404(registrodemanutencao, id=id)
    registro.status = (
        "Aprovado pela Diretoria"  # Corrigido para "Aprovado pela Diretoria"
    )
    registro.save()

    subject = f"Manutenção Aprovada: {registro.id}"
    message = f"A manutenção {registro.id} foi aprovada com sucesso. {registro.nome} Status: {registro.status} criar Requisição"
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = ["sjuniorr6@gmail.com"]

    try:
        send_mail(subject, message, from_email, recipient_list)
        print("Email enviado com sucesso.")
    except Exception as e:
        print(f"Erro ao enviar email: {e}")

    return redirect("diretoriaListViews")


# protocolo de requisicao
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)
from reportlab.platypus.flowables import Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import os


def gerar_pdf_requisicao(requisicao):
    # Caminho para salvar o PDF
    pdf_path = os.path.join(settings.MEDIA_ROOT, f"requisicao-{requisicao.id}.pdf")

    # Criar documento PDF com margem superior ajustada
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=letter,
        topMargin=20,
        bottomMargin=20,
        leftMargin=20,
        rightMargin=20,
    )  # Ajuste de margens
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]

    # Caminho para a imagem do logo (ajuste o caminho conforme necessário)
    imagem_esquerda_path = os.path.join(
        settings.MEDIA_ROOT, "imagens_registros/SIDNEISIDNEISIDNEI.png"
    )

    # Criar a imagem do logo
    logo = Image(
        imagem_esquerda_path, width=100, height=100
    )  # Ajustar o tamanho do logo

    # Usar o estilo do título e ajustar a centralização
    title_paragraph = Paragraph("<b>Protocolo de Requisição</b>", title_style)
    title_paragraph.alignment = 1  # Alinhar o título ao centro

    # Adicionar logo e título em elementos
    elements.append(logo)
    elements.append(title_paragraph)
    elements.append(Spacer(1, 10))  # Reduzido o espaçamento após o título

    # Tabela com os dados
    table_data = [
        [Paragraph("<b>Nome:</b>", normal_style), requisicao.nome],
        [Paragraph("<b>Protocolo:</b>", normal_style), requisicao.id],
        [
            Paragraph("<b>Endereço:</b>", normal_style),
            Paragraph(requisicao.endereco, normal_style),
        ],  # Usando Paragraph para quebra automática
        [Paragraph("<b>Contrato:</b>", normal_style), requisicao.contrato],
        [Paragraph("<b>CNPJ:</b>", normal_style), requisicao.cnpj],
        [
            Paragraph("<b>Data:</b>", normal_style),
            requisicao.data.strftime("%d/%m/%Y") if requisicao.data else "N/A",
        ],
        [Paragraph("<b>Motivo:</b>", normal_style), requisicao.motivo],
        [Paragraph("<b>Taxa de Envio:</b>", normal_style), requisicao.taxa_envio],
        [Paragraph("<b>Comercial:</b>", normal_style), requisicao.comercial],
        [Paragraph("<b>Tipo de Produto:</b>", normal_style), requisicao.tipo_produto],
        [Paragraph("<b>Carregador:</b>", normal_style), requisicao.carregador],
        [Paragraph("<b>Cabo:</b>", normal_style), requisicao.cabo],
        [Paragraph("<b>TP:</b>", normal_style), requisicao.TP],
        [Paragraph("<b>Envio:</b>", normal_style), requisicao.envio],
        [
            Paragraph("<b>Quantidade:</b>", normal_style),
            requisicao.numero_de_equipamentos,
        ],
        [Paragraph("<b>Valor Unitário:</b>", normal_style), requisicao.valor_unitario],
        [Paragraph("<b>Customização:</b>", normal_style), requisicao.tipo_customizacao],
        [
            Paragraph("<b>Observações:</b>", normal_style),
            Paragraph(requisicao.observacoes, normal_style),
        ],  # Usando Paragraph para quebra automática
    ]

    # Criar uma tabela com duas colunas
    table = Table(table_data, colWidths=[200, 300])  # Ajuste para as colunas
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),  # Fonte ajustada
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),  # Ajuste de padding superior
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),  # Ajuste de padding inferior
                ("LEFTPADDING", (0, 0), (-1, -1), 5),  # Ajuste de padding lateral
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),  # Ajuste de padding lateral
            ]
        )
    )

    # Adicionar a frase no final de forma destacada
    highlight_style = ParagraphStyle(
        name="HighlightedText",
        parent=styles["Normal"],
        fontSize=12,
        alignment=1,  # Centraliza o texto
        fontName="Helvetica-Bold",
        spaceBefore=2,  # Espaço antes da frase
        textColor=colors.red,  # Cor vermelha para destacar
    )

    highlight_paragraph = Paragraph(
        "<b>NÃO ENVIAR ESSE PROTOCOLO AO CLIENTE</b>", highlight_style
    )
    elements.append(highlight_paragraph)  # Adiciona a frase ao PDF

    # Adiciona um espaçamento final
    elements.append(Spacer(1, 20))

    elements.append(table)  # Adiciona a tabela ao PDF
    elements.append(Spacer(1, 20))  # Espaço após a tabela

    # Adicionar quebra de página caso o conteúdo ocupe mais de uma página
    elements.append(PageBreak())

    # Salvar PDF
    doc.build(elements)
    return pdf_path


# protocolo de saida
from django.shortcuts import get_object_or_404
from .models import Requisicoes
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import os
from django.conf import settings
from textwrap import wrap
from reportlab.pdfbase.pdfmetrics import stringWidth


# Função para truncar texto dinamicamente para caber no espaço
def truncate_text_to_fit(text, max_width, font_name="Helvetica", font_size=10):
    if not text:
        return "Não Informado"
    while stringWidth(text, font_name, font_size) > max_width:
        text = text[:-1]
    return text + "..." if len(text) > 3 else text


def wrap_text(text, max_length=40):
    if not text:
        return "Não Informado"
    return "\n".join(wrap(str(text), max_length))


from reportlab.lib.colors import HexColor

# protocolo de saida

import os
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import A3
from reportlab.lib.pagesizes import landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

# Substitua "your_app" pelo nome do seu app


def gerar_pdf_saida(request, id):
    requisicao = get_object_or_404(Requisicoes, id=id)

    # Configuração inicial do PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=20 * mm,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        bottomMargin=30 * mm,
    )
    elements = []

    # Estilos de texto
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="Header",
            fontSize=14,
            alignment=1,  # Centralizado
            spaceAfter=10,
            fontName="Helvetica-Bold",
        )
    )
    styles.add(
        ParagraphStyle(
            name="Body",
            fontSize=9,
            alignment=0,  # Alinhado à esquerda
            spaceAfter=6,
            fontName="Helvetica",
        )
    )
    styles.add(
        ParagraphStyle(
            name="Footer",
            fontSize=8,
            alignment=1,  # Centralizado
            spaceBefore=20,
            fontName="Helvetica",
        )
    )
    styles.add(
        ParagraphStyle(
            name="TableHeader", fontSize=9, alignment=0, fontName="Helvetica-Bold"
        )
    )
    styles.add(
        ParagraphStyle(name="TableBody", fontSize=9, alignment=0, fontName="Helvetica")
    )

    # Define o tom de amarelo mais suave
    soft_yellow = HexColor("#FFFACD")

    # Cabeçalho com logotipo e QR Code
    try:
        logo_path = os.path.join(
            settings.MEDIA_ROOT, "imagens_registros/SIDNEISIDNEISIDNEI.png"
        )
        qr_code_path = os.path.join(settings.MEDIA_ROOT, "imagens_registros/qrcode.png")

        # Verifica se os arquivos existem antes de tentar abrir
        if not os.path.exists(logo_path):
            raise FileNotFoundError(f"Logo não encontrado: {logo_path}")
        if not os.path.exists(qr_code_path):
            raise FileNotFoundError(f"QR Code não encontrado: {qr_code_path}")

        logo = Image(logo_path, width=60, height=60)
        qr_code = Image(qr_code_path, width=60, height=60)

        header_table = Table(
            [
                [
                    logo,
                    Paragraph(
                        "<b>PROTOCOLO DE ENTREGA DE EQUIPAMENTOS</b>", styles["Header"]
                    ),
                    qr_code,
                ]
            ],
            colWidths=[80 * mm, 330 * mm, 80 * mm],
        )
        header_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (1, 0), (1, 0)),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ]
            )
        )
        elements.append(header_table)
    except (FileNotFoundError, Exception) as e:
        print(f"Erro ao carregar imagens do cabeçalho: {e}")
        elements.append(
            Paragraph("<b>PROTOCOLO DE ENTREGA DE EQUIPAMENTOS</b>", styles["Header"])
        )

    elements.append(Spacer(1, 10))

    # Função para texto formatado
    def para(text, style_name="TableBody"):
        return Paragraph(text, styles[style_name])

    # Informações principais
    fields = [
        [para("<b>Nº PEDIDO:</b>", "TableHeader"), para(str(requisicao.id))],
        [
            para("<b>DATA:</b>", "TableHeader"),
            para(requisicao.data.strftime("%d/%m/%Y") if requisicao.data else "N/A"),
        ],
        [
            para("<b>ENDEREÇO:</b>", "TableHeader"),
            para(str(requisicao.endereco) if requisicao.endereco else "N/A"),
        ],
        [
            para("<b>FORMA DE ENVIO:</b>", "TableHeader"),
            para(str(requisicao.envio) if requisicao.envio else "N/A"),
        ],
        [
            para("<b>CLIENTE:</b>", "TableHeader"),
            para(str(requisicao.nome) if requisicao.nome else "N/A"),
        ],
        [
            para("<b>QTD:</b>", "TableHeader"),
            para(
                str(requisicao.numero_de_equipamentos)
                if requisicao.numero_de_equipamentos
                else "Não Informado"
            ),
        ],
        [
            para("<b>COMERCIAL:</b>", "TableHeader"),
            para(
                str(requisicao.comercial) if requisicao.comercial else "Não Informado"
            ),
        ],
        [
            para("<b>CARREGADOR:</b>", "TableHeader"),
            para(
                str(requisicao.carregador) if requisicao.carregador else "Não Informado"
            ),
        ],
        [
            para("<b>TIPO DE PRODUTO:</b>", "TableHeader"),
            para(
                str(requisicao.tipo_produto)
                if requisicao.tipo_produto
                else "Não Informado"
            ),
        ],
        [
            para("<b>TP:</b>", "TableHeader"),
            para(str(requisicao.TP) if requisicao.TP else "Não Informado"),
        ],
        [
            para("<b>AOS CUIDADOS:</b>", "TableHeader"),
            para(
                str(requisicao.aos_cuidados)
                if requisicao.aos_cuidados
                else "Não Informado"
            ),
        ],
        [
            para("<b>CUSTOMIZAÇÃO:</b>", "TableHeader"),
            para(
                str(requisicao.tipo_customizacao)
                if requisicao.tipo_customizacao
                else "Não Informado"
            ),
        ],
        [
            para("<b>Tipo de Contrato:</b>", "TableHeader"),
            para(str(requisicao.contrato) if requisicao.contrato else "Não Informado"),
        ],
    ]

    table = Table(fields, colWidths=[50 * mm, 120 * mm])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (0, -1), soft_yellow),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements.append(table)
    elements.append(Spacer(1, 10))

    # IDs dos Equipamentos
    if requisicao.id_equipamentos:
        equipamentos = requisicao.id_equipamentos.split()
    else:
        equipamentos = ["Não Informado"]

    id_rows = [equipamentos[i : i + 5] for i in range(0, len(equipamentos), 5)]

    elements.append(Paragraph("<b>ID - EQUIPAMENTOS:</b>", styles["Body"]))

    for row in id_rows:
        row_table = Table([row], colWidths=[40 * mm] * 5)
        row_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        elements.append(row_table)
        elements.append(Spacer(1, 5))

    # Rodapé
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawString(50, 40, "NOME:")
        canvas.drawString(220, 40, "ASS.:")
        canvas.drawString(380, 40, "CPF:")
        canvas.drawString(500, 40, "DATA E HORA:")
        canvas.restoreState()

    # Geração do PDF
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="protocolo_saida_{id}.pdf"'
    buffer.close()
    return response


def enviar_email_com_pdf(request, id):
    requisicao = get_object_or_404(Requisicoes, id=id)
    pdf_path = gerar_pdf_requisicao(requisicao)

    subject = f"Requisição Criada: {requisicao.id}"
    message = f"A requisição {requisicao.id} foi criada com sucesso. Segue PDF para tratativa."
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = ["sjuniorr6@gmail.com"]

    email = EmailMessage(subject, message, from_email, recipient_list)
    email.attach_file(pdf_path)

    try:
        email.send()
        print("Email enviado com sucesso.")
    except Exception as e:
        print(f"Erro ao enviar email: {e}")

    return redirect("requisicoesListView")


from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from .models import Requisicoes


def download_pdf_requisicao(request, id):
    requisicao = get_object_or_404(Requisicoes, id=id)
    pdf_path = gerar_pdf_requisicao(requisicao)

    with open(pdf_path, "rb") as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="requisicao-{requisicao.id}.pdf"'
        )

        return response


from django.db import transaction
from django.db.models import F
from django.contrib.auth.mixins import PermissionRequiredMixin, LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from .models import estoque_antenista
from .forms import EstoqueantenistarForm
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.shortcuts import redirect


class RegistrarEstoqueantenistaView(
    PermissionRequiredMixin, LoginRequiredMixin, CreateView
):
    model = estoque_antenista
    form_class = EstoqueantenistarForm
    template_name = "estoque_antenista.html"
    success_url = reverse_lazy("RegistrarEstoqueantenistaView")
    permission_required = "tuper.add_estoque_tuper"

    def form_valid(self, form):
        nome = form.cleaned_data["nome"]
        tipo_produto = form.cleaned_data["tipo_produto"]
        quantidade = (
            form.cleaned_data["quantidade"] // 2
        )  # Retira metade do valor adicionado
        endereco = form.cleaned_data["endereco"]

        with transaction.atomic():
            estoques = estoque_antenista.objects.filter(
                nome=nome, tipo_produto=tipo_produto
            )
            if estoques.exists():
                estoque = estoques.order_by("-quantidade").first()
                estoque.quantidade = F("quantidade") + quantidade
                estoque.endereco = endereco
                estoque.save(update_fields=["quantidade", "endereco"])
            else:
                estoque = estoque_antenista(
                    nome=nome,
                    tipo_produto=tipo_produto,
                    quantidade=quantidade,
                    endereco=endereco,
                )
                estoque.save()

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        estoques = estoque_antenista.objects.all()
        estoque_dict = {}

        for estoque in estoques:
            key = (estoque.nome, estoque.tipo_produto)
            if key in estoque_dict:
                estoque_dict[key].quantidade += estoque.quantidade
            else:
                estoque_dict[key] = estoque

        for key, estoque in estoque_dict.items():
            estoque.save()
            estoque_antenista.objects.filter(nome=key[0], tipo_produto=key[1]).exclude(
                id=estoque.id
            ).delete()

        context["estoques"] = list(estoque_dict.values())
        return context


@permission_required("requisicao.change_estoque_antenista", raise_exception=True)
def zerar_estoque_antenista(request):
    """Zera (define para 0) a quantidade de todos os registros de estoque_antenista.

    Esta ação exige permissão `requisicao.change_estoque_antenista` e deve ser chamada via POST.
    """
    if request.method == "POST":
        estoque_antenista.objects.update(quantidade=0)
        messages.success(
            request, "Quantidade em estoque zerada para todos os registros."
        )
    else:
        messages.error(request, "Requisição inválida. Use POST para zerar o estoque.")
    return redirect("RegistrarEstoqueantenistaView")


@login_required
def delete_estoque_antenista(request, pk):
    """Apaga um registro de estoque_antenista via POST com verificação de permissão."""
    from django.shortcuts import get_object_or_404

    estoque = get_object_or_404(estoque_antenista, pk=pk)
    # verifica permissão
    if not request.user.has_perm("requisicao.delete_estoque_antenista"):
        messages.error(request, "Você não tem permissão para apagar esse registro.")
        return redirect("RegistrarEstoqueantenistaView")

    if request.method == "POST":
        estoque.delete()
        messages.success(request, "Registro de estoque apagado com sucesso.")
        return redirect("RegistrarEstoqueantenistaView")
    else:
        messages.error(request, "Método inválido para exclusão.")
        return redirect("RegistrarEstoqueantenistaView")


from .models import antenista_CARD, Antenista
from .forms import antenista_Form, AntenistaForm


class AntenistaCreateView(CreateView):
    # Cria um registro no modelo Antenista (nome + estado)
    model = Antenista
    form_class = AntenistaForm
    template_name = "novo_antenista.html"
    success_url = reverse_lazy("lista_antenistas")

    def form_valid(self, form):
        # Envia o e-mail após salvar o registro
        response = super().form_valid(form)  # Salva o registro

        # Configuração do e-mail
        send_mail(
            subject="Projeto Fast - Novo Registro",
            message="Um novo registro foi criado no sistema. Comercial favor tratar",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=["sjuniorr6@Gmail.com", "comercial@grupogoldensat.com.br"],
            fail_silently=False,  # Define como True para evitar erros visíveis ao usuário
        )

        return response


class AntenistaCardCreateView(LoginRequiredMixin, CreateView):
    """Cria um registro de antenista_CARD usando o template de Saída de Equipamentos.

    Mantém o template `novo_antenista.html` intacto (para cadastro de Antenista simples)
    e expõe uma view separada que usa `antenista_form.html` para o fluxo "Saída de Equipamentos".
    """

    model = antenista_CARD
    form_class = antenista_Form
    template_name = "antenista_form.html"
    success_url = reverse_lazy("lista_antenistas")

    def form_valid(self, form):
        # qualquer lógica extra pode ser adicionada aqui (enviar email, logs, etc.)
        return super().form_valid(form)


from django.views.generic.list import ListView
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .models import Antenista


class AntenistaListView(ListView):
    # Mantém a listagem existente de antenista_CARD (compatibilidade com export e template atual)
    model = antenista_CARD
    template_name = "antenista_list.html"
    context_object_name = "antenistas"


class AntenistaCadastradosListView(
    PermissionRequiredMixin, LoginRequiredMixin, ListView
):
    """Listagem simples dos antenistas cadastrados (modelo `Antenista`).

    Não substitui a listagem de `antenista_CARD` — fornece uma visão dos
    antenistas registrados (nome e estado), útil para administração.
    """

    model = Antenista
    template_name = "antenista_cadastrados_list.html"
    context_object_name = "antenistas_cadastrados"
    paginate_by = 50
    permission_required = "requisicao.view_antenista"


class AntenistaUpdateView(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    """Edita um Antenista cadastrado."""

    model = Antenista
    form_class = AntenistaForm
    template_name = "novo_antenista.html"  # Reusa o template de criação
    permission_required = "requisicao.change_antenista"
    success_url = reverse_lazy("lista_antenistas_cadastrados")


class AntenistaDeleteView(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    """Exclui um Antenista cadastrado (POST via formulário)."""

    model = Antenista
    template_name = "confirm_delete.html"
    permission_required = "requisicao.delete_antenista"
    success_url = reverse_lazy("lista_antenistas_cadastrados")

    def delete(self, request, *args, **kwargs):
        # Robust deletion: avoid calling get_object() early because it may trigger
        # related-table queries that fail when schema is inconsistent. Instead,
        # use the pk from the URL and attempt an ORM delete; if any
        # OperationalError occurs, fall back to a direct SQL DELETE on the model's
        # table so the UI can continue to manage records while migrations are fixed.
        from django.db import connection, transaction, utils as db_utils
        from django.contrib import messages

        pk = kwargs.get(self.pk_url_kwarg) or kwargs.get("pk")
        if not pk:
            messages.error(request, "PK não informado para exclusão.")
            return redirect(self.success_url)

        # First, try to delete via ORM but avoid calling get_object() which may
        # cause related-object queries.
        try:
            # Use a filtered queryset delete to minimize related-object collection
            # (QuerySet.delete() still uses collector, but this keeps code simple).
            qs = Antenista.objects.filter(pk=pk)
            if qs.exists():
                try:
                    qs.delete()
                    messages.success(request, "Registro apagado com sucesso.")
                    return redirect(self.success_url)
                except db_utils.OperationalError:
                    # fall through to raw delete
                    pass
            else:
                messages.info(request, "Registro não encontrado (já removido).")
                return redirect(self.success_url)
        except db_utils.OperationalError:
            # fall back to raw SQL delete below
            pass

        # Raw SQL fallback: delete directly from the model table using the pk.
        try:
            table = Antenista._meta.db_table
            with transaction.atomic():
                with connection.cursor() as cursor:
                    cursor.execute(f"DELETE FROM {table} WHERE id = ?", [pk])
            messages.success(request, "Registro apagado com sucesso (delete direto).")
            return redirect(self.success_url)
        except Exception:
            # If even the raw delete fails, raise to show the full error for debugging.
            raise


def export_antenistas_excel(request):
    """Exporta todos os registros de antenista_CARD para um arquivo Excel (.xlsx)."""
    # Query todos os registros
    antenistas = antenista_CARD.objects.all()

    # Cria workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Antenistas"

    # Cabeçalhos (mesma ordem da tabela)
    headers = [
        "Nome",
        "Tipo de Produto",
        "Telefone",
        "Cliente",
        "Quantidade",
        "Equipamentos",
        "Solicitante",
        "Valor Total",
        "Valor Prestador",
        "Valor Isca",
        "Valor Cliente",
        "Lucro",
        "Contrato",
        "Status",
        "Data de Criação",
    ]
    ws.append(headers)

    # Preenche linhas
    for a in antenistas:
        row = [
            str(a.nome) if a.nome is not None else "",
            str(a.tipo_produto) if a.tipo_produto is not None else "",
            a.telefone or "",
            a.cliente or "",
            a.quantidade if a.quantidade is not None else "",
            a.equipamentos or "",
            a.solicitante or "",
            float(a.valor_total) if a.valor_total is not None else 0,
            float(a.valor_prestador) if a.valor_prestador is not None else 0,
            float(a.valor_isca) if a.valor_isca is not None else 0,
            float(a.valor_cliente) if a.valor_cliente is not None else 0,
            float(a.lucro) if a.lucro is not None else 0,
            a.contrato or "",
            a.status or "",
            a.data_criacao.strftime("%d/%m/%Y") if a.data_criacao else "",
        ]
        ws.append(row)

    # Ajusta largura das colunas
    for i, column_cells in enumerate(ws.columns, 1):
        length = max((len(str(cell.value)) for cell in column_cells), default=0)
        ws.column_dimensions[get_column_letter(i)].width = min(max(length + 2, 10), 50)

    # Prepara resposta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=antenistas.xlsx"
    wb.save(response)
    return response

    # requisicao/views.py


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import antenista_CARD


def atualizar_status_atualizado(request, pk):
    antenista = get_object_or_404(antenista_CARD, pk=pk)
    if request.method == "POST":
        if antenista.status != "Atualizado":
            antenista.status = "Atualizado"
            antenista.save()
            messages.success(
                request, f"Status de {antenista.nome} atualizado para 'Atualizado'."
            )
        else:
            messages.info(
                request, f"O status de {antenista.nome} já está 'Atualizado'."
            )
        return redirect("lista_antenistas")
    else:
        messages.error(request, "Método HTTP inválido.")
        return redirect("lista_antenistas")


@login_required
def delete_antenista_card(request, pk):
    """Apaga um registro de antenista_CARD via POST.

    Requer que o usuário esteja autenticado e tenha permissão
    `requisicao.delete_antenista_card`.
    """
    antenista = get_object_or_404(antenista_CARD, pk=pk)
    # Verifica permissão explícita
    if not request.user.has_perm("requisicao.delete_antenista_card"):
        messages.error(request, "Você não tem permissão para apagar este registro.")
        return redirect("lista_antenistas")

    if request.method == "POST":
        antenista.delete()
        messages.success(request, "Registro apagado com sucesso.")
        return redirect("lista_antenistas")
    else:
        messages.error(request, "Método inválido para exclusão.")
        return redirect("lista_antenistas")


from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import UpdateView
from .models import Requisicoes
from .forms import RequisicaoForm  # Supondo que você tenha um form para Requisicao


class RequisicaoUpdateView(UpdateView):
    model = Requisicoes
    form_class = RequisicaoForm  # Você pode usar um ModelForm ou o form padrão
    template_name = "requisicao_update.html"
    success_url = reverse_lazy(
        "requisicoes_list"
    )  # Redireciona para a lista após a atualização

    def form_valid(self, form):
        form.instance.data_alteracao = timezone.now()
        return super().form_valid(form)


# ============== EXPEDIÇÃO PARCIAL ==============


@require_POST
@login_required
def expedir_requisicao_parcial(request):
    """
    Expede parcialmente uma requisição, criando uma nova com a quantidade restante.
    Apenas usuários do grupo 'Gestão Kanban' podem executar esta ação.
    """
    try:
        # Verifica se usuário pertence ao grupo Gestão Kanban
        if not request.user.groups.filter(name="Gestão Kanban").exists():
            return JsonResponse(
                {
                    "success": False,
                    "message": "Você não tem permissão para expedir requisições.",
                },
                status=403,
            )

        data = json.loads(request.body)
        requisicao_id = data.get("requisicao_id")
        quantidade_expedir = int(data.get("quantidade_expedir", 0))
        ids_auditados = data.get("ids_auditados", "").strip()
        verificacao_plataforma = data.get("verificacao_plataforma", False)
        customizacao_conforme = data.get("customizacao_conforme", False)

        requisicao = get_object_or_404(Requisicoes, id=requisicao_id)

        # Salva dados do checklist
        requisicao.ids_auditados = ids_auditados
        requisicao.verificacao_plataforma = verificacao_plataforma
        requisicao.customizacao_conforme = customizacao_conforme

        # Converte numero_de_equipamentos para int
        numero_equipamentos = (
            int(requisicao.numero_de_equipamentos)
            if requisicao.numero_de_equipamentos
            else 0
        )

        # Validações
        if quantidade_expedir <= 0:
            return JsonResponse(
                {"success": False, "message": "Quantidade deve ser maior que zero."}
            )

        if quantidade_expedir > numero_equipamentos:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Quantidade não pode ser maior que {numero_equipamentos}.",
                }
            )

        # Usa transaction para garantir atomicidade
        with transaction.atomic():
            # Calcula valores para faturamento
            valor_unitario = (
                float(requisicao.valor_unitario) if requisicao.valor_unitario else 0
            )
            taxa_envio = float(requisicao.taxa_envio) if requisicao.taxa_envio else 0
            valor_a_faturar = (valor_unitario * quantidade_expedir) + taxa_envio

            # Atualiza requisição original - MARCA COMO EXPEDIDA PARCIALMENTE
            requisicao.quantidade_expedida += quantidade_expedir
            requisicao.status = "Configurado"  # Muda status para expedido
            requisicao.kanban_status = None  # Remove do kanban

            # Adiciona informação de expedição parcial nas observações
            obs_parcial = f"\n[EXPEDIÇÃO PARCIAL] {quantidade_expedir} de {numero_equipamentos} expedidos. Valor a faturar: R$ {valor_a_faturar:.2f} (Qtd: {quantidade_expedir} x R$ {valor_unitario:.2f} + Taxa: R$ {taxa_envio:.2f})"
            if requisicao.observacoes:
                requisicao.observacoes += obs_parcial
            else:
                requisicao.observacoes = obs_parcial.strip()

            requisicao.save()

            # Registra log de auditoria da expedição parcial
            AuditLog.registrar(
                objeto=requisicao,
                acao="expedicao_parcial",
                usuario=request.user,
                status_anterior=requisicao.status,
                status_novo="Configurado",
                detalhes={
                    "tipo_expedicao": "parcial",
                    "quantidade_expedida": quantidade_expedir,
                    "quantidade_total": numero_equipamentos,
                    "quantidade_restante": numero_equipamentos - quantidade_expedir,
                    "ids_auditados": ids_auditados,
                    "verificacao_plataforma": verificacao_plataforma,
                    "customizacao_conforme": customizacao_conforme,
                    "valor_faturar": valor_a_faturar,
                },
                observacao=f"Expedição parcial: {quantidade_expedir} de {numero_equipamentos} equipamentos",
                request=request,
            )

            # Registra log de expedição parcial (KanbanAuditLog - mantém compatibilidade)
            KanbanAuditLog.objects.create(
                requisicao=requisicao,
                usuario=request.user,
                acao="expedicao_parcial",
                coluna_origem="auditoria",
                coluna_destino="expedido",
                quantidade_expedida=quantidade_expedir,
                observacao=f"Expedidos {quantidade_expedir} de {numero_equipamentos} equipamentos",
            )

            # Calcula quantidade restante
            quantidade_restante = numero_equipamentos - quantidade_expedir

            # Se ainda há quantidade restante, cria nova requisição
            if quantidade_restante > 0:
                # Calcula valor total para nova requisição
                valor_unitario = (
                    float(requisicao.valor_unitario) if requisicao.valor_unitario else 0
                )
                valor_total_novo = (
                    valor_unitario * quantidade_restante if valor_unitario > 0 else 0
                )

                # Cria nova requisição SEM disparar signals problemáticos
                try:
                    # Copia TODOS os dados importantes da requisição original
                    nova_requisicao = Requisicoes(
                        # Dados do cliente
                        nome=requisicao.nome,
                        endereco=requisicao.endereco or "",
                        contrato=requisicao.contrato or "",
                        cnpj=requisicao.cnpj or "",
                        inicio_de_contrato=requisicao.inicio_de_contrato,
                        vigencia=requisicao.vigencia or "",
                        motivo=requisicao.motivo or "",
                        aos_cuidados=requisicao.aos_cuidados or "",
                        # Dados do produto
                        tipo_produto=requisicao.tipo_produto,
                        numero_de_equipamentos=quantidade_restante,
                        tipo_customizacao=requisicao.tipo_customizacao
                        or "Sem customização",
                        carregador=requisicao.carregador or "",
                        cabo=requisicao.cabo or "",
                        # Dados comerciais e financeiros
                        comercial=requisicao.comercial,
                        valor_unitario=valor_unitario,
                        valor_total=valor_total_novo,
                        taxa_envio=requisicao.taxa_envio or 0,
                        tipo_fatura=requisicao.tipo_fatura,
                        forma_pagamento=requisicao.forma_pagamento or "",
                        # Dados de entrega
                        envio=requisicao.envio,
                        data_entrega=requisicao.data_entrega,
                        TP=requisicao.TP,
                        # Status e controle
                        observacoes=f"[EXPEDIÇÃO PARCIAL] Restante da requisição #{requisicao.id}",
                        status="Pendente",
                        kanban_status="a_fazer",
                        requisicao_original_id=requisicao,
                        id_equipamentos="",
                        iccid="",
                    )
                    # Marca para pular signals (evita envio de email e geração de PDF)
                    nova_requisicao._skip_signals = True
                    nova_requisicao.save()

                    # Registra log de criação da requisição complementar
                    AuditLog.registrar(
                        objeto=nova_requisicao,
                        acao="criacao",
                        usuario=request.user,
                        status_novo="Pendente",
                        detalhes={
                            "requisicao_origem_id": requisicao.id,
                            "tipo": "complementar",
                            "quantidade": quantidade_restante,
                            "motivo": f"Expedição parcial da requisição #{requisicao.id}",
                        },
                        observacao=f"Requisição complementar criada a partir da expedição parcial #{requisicao.id}",
                        request=request,
                    )

                except Exception as e_create:
                    # Se falhar, loga e continua
                    print(f"Erro ao criar nova requisição: {e_create}")
                    import traceback

                    traceback.print_exc()
                    # Retorna erro para o usuário
                    return JsonResponse(
                        {
                            "success": False,
                            "message": f"Erro ao criar requisição restante: {str(e_create)[:100]}",
                        },
                        status=500,
                    )

                # Registra log da nova requisição
                KanbanAuditLog.objects.create(
                    requisicao=nova_requisicao,
                    usuario=request.user,
                    acao="movimento",
                    coluna_origem=None,
                    coluna_destino="a_fazer",
                    observacao=f"Criada a partir da expedição parcial da requisição #{requisicao.id}",
                )

                return JsonResponse(
                    {
                        "success": True,
                        "message": f'Expedição parcial realizada! {quantidade_expedir} expedidos, {quantidade_restante} retornaram para "A Fazer" (Req #{nova_requisicao.id})',
                        "nova_requisicao_id": nova_requisicao.id,
                    }
                )
            else:
                # Expedição total - marca como expedida
                KanbanAuditLog.objects.create(
                    requisicao=requisicao,
                    usuario=request.user,
                    acao="expedicao_total",
                    coluna_origem=requisicao.kanban_status,
                    coluna_destino="concluido",
                    quantidade_expedida=quantidade_expedir,
                    observacao="Expedição total da requisição",
                )

                return JsonResponse(
                    {
                        "success": True,
                        "message": f"Expedição total realizada! Todos os {quantidade_expedir} equipamentos foram expedidos.",
                    }
                )

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Dados inválidos."}, status=400)
    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Erro ao processar expedição: {str(e)}"},
            status=500,
        )


@csrf_exempt
@require_http_methods(["POST"])
def salvar_ids_equipamentos(request, pk):
    """
    Salva IDs e ICCIDs dos equipamentos de uma requisição via AJAX
    """
    try:
        # Verifica permissões
        if not request.user.groups.filter(
            name__in=["Gestão Kanban", "Configuração Kanban"]
        ).exists():
            return JsonResponse(
                {"success": False, "error": "Você não tem permissão para esta ação."},
                status=403,
            )

        # Busca requisição
        requisicao = get_object_or_404(Requisicoes, pk=pk)

        # Parse JSON body
        data = json.loads(request.body)
        ids_equipamentos = data.get("ids_equipamentos", "").strip()
        iccid = data.get("iccid", "").strip()

        # Validações
        if not ids_equipamentos:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Por favor, informe os IDs dos equipamentos.",
                },
                status=400,
            )

        if not iccid:
            return JsonResponse(
                {"success": False, "error": "Por favor, informe os ICCIDs."}, status=400
            )

        # Conta quantidade de IDs
        ids_list = ids_equipamentos.split()
        iccid_list = iccid.split()

        if len(ids_list) != len(iccid_list):
            return JsonResponse(
                {
                    "success": False,
                    "error": f"Quantidade de IDs ({len(ids_list)}) diferente da quantidade de ICCIDs ({len(iccid_list)}).",
                },
                status=400,
            )

        # Atualiza requisição
        requisicao.id_equipamentos = ids_equipamentos
        requisicao.iccid = iccid
        requisicao.save()

        # Log de auditoria (KanbanAuditLog - mantém compatibilidade)
        KanbanAuditLog.objects.create(
            requisicao=requisicao,
            usuario=request.user,
            acao="inclusao_ids",
            observacao=f"IDs e ICCIDs incluídos: {len(ids_list)} equipamentos",
        )

        # Log de auditoria (AuditLog - sistema completo)
        AuditLog.registrar(
            objeto=requisicao,
            acao="ids_incluidos",
            usuario=request.user,
            detalhes={
                "quantidade_ids": len(ids_list),
                "quantidade_iccids": len(iccid_list),
                "ids": ids_equipamentos[:200],  # Limita tamanho para não explodir JSON
            },
            observacao=f"IDs e ICCIDs incluídos: {len(ids_list)} equipamentos",
            request=request,
        )

        return JsonResponse(
            {
                "success": True,
                "message": f"IDs e ICCIDs salvos com sucesso! ({len(ids_list)} equipamentos)",
            }
        )

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Dados inválidos."}, status=400)
    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Erro ao salvar IDs: {str(e)}"}, status=500
        )


# ============== KANBAN BOARD VIEWS ==============

from django.db.models import Count, F
from datetime import datetime, timedelta, time


class KanbanGestaoView(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    """
    View principal do Kanban Board para gestão de requisições.
    Exibe 3 colunas: A Fazer, Em Progresso, Auditoria
    Com estatísticas e cards drag & drop.
    """

    model = Requisicoes
    template_name = "kanban_gestao.html"
    context_object_name = "requisicoes"
    permission_required = "requisicao.view_requisicoes"

    def get_queryset(self):
        """
        Retorna requisições com status 'Aprovado pelo CEO' para gerenciamento no Kanban.
        Exclui produtos: GS310, GS340, GS390, GS8310 (4G), PLUG AND PLAY (mesmos da ConfiguracaoListView).
        Ordena por prioridade (DESC) e data (ASC - mais antigas primeiro).
        """
        return (
            Requisicoes.objects.filter(status="Aprovado pelo CEO")
            .exclude(
                tipo_produto__nome__in=[
                    "GS310",
                    "GS340",
                    "GS390",
                    "GS8310 (4G)",
                ]
            )
            .select_related("nome", "tipo_produto")
            .order_by("-prioridade", "data")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Separa requisições por coluna do Kanban
        todas_requisicoes = self.get_queryset()

        context["a_fazer"] = todas_requisicoes.filter(kanban_status="a_fazer")
        context["em_progresso"] = todas_requisicoes.filter(kanban_status="em_progresso")
        context["auditoria"] = todas_requisicoes.filter(kanban_status="auditoria")

        # Calcula estatísticas para o dashboard
        hoje = timezone.now()
        sete_dias_atras = hoje - timedelta(days=7)

        # Total de tarefas (apenas Aprovado pelo CEO)
        total_tarefas = todas_requisicoes.count()

        # Concluídas (as que já foram expedidas/enviadas ao cliente desde que foram aprovadas pelo CEO)
        concluidas = (
            Requisicoes.objects.filter(
                status__in=["Configurado", "Enviado para o cliente", "Expedido"]
            )
            .exclude(status="Aprovado pelo CEO")
            .count()
        )

        # Em progresso
        em_progresso_count = todas_requisicoes.filter(
            kanban_status="em_progresso"
        ).count()

        # Atrasadas (mais de 7 dias desde a última alteração e ainda não concluídas)
        atrasadas = todas_requisicoes.filter(data_alteracao__lt=sete_dias_atras).count()

        # Taxa de conclusão (baseado no total geral de requisições aprovadas)
        total_aprovadas_historico = Requisicoes.objects.filter(
            status__in=[
                "Aprovado pelo CEO",
                "Configurado",
                "Enviado para o cliente",
                "Expedido",
            ]
        ).count()
        taxa_conclusao = round(
            (concluidas / total_aprovadas_historico * 100)
            if total_aprovadas_historico > 0
            else 0
        )

        context["stats"] = {
            "total_tarefas": total_tarefas,
            "concluidas": concluidas,
            "em_progresso": em_progresso_count,
            "atrasadas": atrasadas,
            "taxa_conclusao": taxa_conclusao,
        }

        return context


@require_POST
@login_required
def update_kanban_status(request):
    """
    Endpoint AJAX para atualizar o status do Kanban quando um card é arrastado.
    Valida permissões do usuário e se IDs estão incluídos antes de permitir mover para 'auditoria'.
    Registra todas as movimentações no KanbanAuditLog.
    """
    try:
        data = json.loads(request.body)
        requisicao_id = data.get("requisicao_id")
        novo_status = data.get("novo_status")
        responsavel = data.get("responsavel_manutencao")  # Novo campo

        requisicao = get_object_or_404(Requisicoes, id=requisicao_id)
        status_anterior = requisicao.kanban_status

        # Validação: se tentou mover para a mesma coluna, ignora
        if status_anterior == novo_status:
            return JsonResponse(
                {
                    "success": True,
                    "message": "Card já está nesta coluna.",
                    "cor_card": requisicao.cor_card,
                }
            )

        # Verifica permissões do usuário
        is_gestao = request.user.groups.filter(name="Gestão Kanban").exists()
        is_config = request.user.groups.filter(name="Configuração Kanban").exists()

        # Se não pertence a nenhum grupo, bloqueia
        if not is_gestao and not is_config:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Você não tem permissão para mover cards no Kanban.",
                },
                status=403,
            )

        # Configuração só pode mover de "em_progresso" para "auditoria"
        if is_config and not is_gestao:
            if status_anterior != "em_progresso" or novo_status != "auditoria":
                return JsonResponse(
                    {
                        "success": False,
                        "error": 'Você só pode mover cards de "Em Progresso" para "Auditoria".',
                    },
                    status=403,
                )

            # Validação adicional: usuário configuração só pode mover se for o responsável atribuído ou super user
            if status_anterior == "em_progresso" and novo_status == "auditoria":
                if not request.user.is_superuser:
                    if not requisicao.responsavel_manutencao:
                        return JsonResponse(
                            {
                                "success": False,
                                "error": "Este card não possui responsável atribuído.",
                            },
                            status=403,
                        )

                    # Verifica se o username do usuário corresponde ao responsável
                    if request.user.username != requisicao.responsavel_manutencao:
                        # Busca o nome formatado do responsável para a mensagem
                        responsavel_nome = dict(
                            requisicao.RESPONSAVEL_MANUTENCAO_CHOICES
                        ).get(
                            requisicao.responsavel_manutencao,
                            requisicao.responsavel_manutencao,
                        )
                        return JsonResponse(
                            {
                                "success": False,
                                "error": f"Apenas {responsavel_nome} ou um gestor pode mover este card para Auditoria.",
                            },
                            status=403,
                        )

        # Validação: ao mover de "a_fazer" para "em_progresso", exige responsável
        if status_anterior == "a_fazer" and novo_status == "em_progresso":
            if not responsavel:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "Atribuição necessária",
                        "requer_atribuicao": True,
                        "requisicao_id": requisicao_id,
                    },
                    status=400,
                )
            # Atribui o responsável
            requisicao.responsavel_manutencao = responsavel

        # Validação: não pode mover para auditoria sem IDs de equipamentos
        # EXCEÇÃO: Produtos do tipo "CARREGADOR + CABO" não precisam de ID de equipamento
        if novo_status == "auditoria":
            # Verifica se é um produto do tipo CARREGADOR + CABO (normaliza espaços extras)
            nome_produto_normalizado = (
                requisicao.tipo_produto.nome.strip().upper()
                if requisicao.tipo_produto
                else ""
            )
            eh_carregador_cabo = (
                "CARREGADOR" in nome_produto_normalizado
                and "CABO" in nome_produto_normalizado
            )

            # Log temporário para debug
            print(
                f"DEBUG - Req #{requisicao.id}: Tipo Produto=[{requisicao.tipo_produto.nome if requisicao.tipo_produto else 'None'}] | É Carregador+Cabo? {eh_carregador_cabo}"
            )

            if not eh_carregador_cabo:
                # Para produtos normais, exige IDs de equipamentos
                if (
                    not requisicao.id_equipamentos
                    or requisicao.id_equipamentos.strip() == ""
                ):
                    return JsonResponse(
                        {
                            "success": False,
                            "error": "Não é possível mover para Auditoria sem incluir os IDs dos equipamentos.",
                        },
                        status=400,
                    )

                # Valida se a quantidade de IDs não é MAIOR que a quantidade de equipamentos
                ids_list = requisicao.id_equipamentos.strip().split()
                quantidade_ids = len(ids_list)
                quantidade_esperada = (
                    int(requisicao.numero_de_equipamentos)
                    if requisicao.numero_de_equipamentos
                    else 0
                )

                if quantidade_ids > quantidade_esperada:
                    return JsonResponse(
                        {
                            "success": False,
                            "error": f"Quantidade de IDs ({quantidade_ids}) é MAIOR que a quantidade de equipamentos ({quantidade_esperada}). Por favor, verifique os IDs incluídos.",
                        },
                        status=400,
                    )

                # Se quantidade de IDs é MENOR, permite mover (será expedição parcial)
                # Se quantidade de IDs é IGUAL, permite mover normalmente

                # Marca o card com cor especial quando IDs estão incluídos
                requisicao.cor_card = "ids-incluidos"
            else:
                # Para CARREGADOR + CABO, permite ir para auditoria sem IDs
                # Marca com uma cor especial para indicar que não precisa de IDs
                requisicao.cor_card = "carregador-cabo"

        # Armazena o status anterior para o signal
        requisicao._kanban_status_anterior = status_anterior
        requisicao._usuario_mudanca = request.user

        # Atualiza o status
        requisicao.kanban_status = novo_status
        requisicao.data_alteracao = timezone.now()
        requisicao.save()

        # Registra no audit log
        observacao = f"Card movido de {status_anterior} para {novo_status}"
        if responsavel:
            observacao += f" | Atribuído a: {responsavel}"

        # Registra no KanbanAuditLog (mantém compatibilidade)
        KanbanAuditLog.objects.create(
            requisicao=requisicao,
            usuario=request.user,
            acao="movimento",
            coluna_origem=status_anterior,
            coluna_destino=novo_status,
            observacao=observacao,
        )

        # Registra no AuditLog (sistema completo de auditoria)
        detalhes = {"coluna_origem": status_anterior, "coluna_destino": novo_status}

        # Se atribuiu responsável, adiciona aos detalhes
        if (
            status_anterior == "a_fazer"
            and novo_status == "em_progresso"
            and responsavel
        ):
            detalhes["responsavel_atribuido"] = responsavel
            acao_audit = "atribuicao"
            observacao_audit = (
                f"Card movido para Em Progresso. Responsável atribuído: {responsavel}"
            )
        else:
            acao_audit = "kanban_movido"
            observacao_audit = observacao

        AuditLog.registrar(
            objeto=requisicao,
            acao=acao_audit,
            usuario=request.user,
            status_anterior=status_anterior,
            status_novo=novo_status,
            detalhes=detalhes,
            observacao=observacao_audit,
            request=request,
        )

        return JsonResponse(
            {
                "success": True,
                "message": f"Requisição movida para {dict(requisicao.KANBAN_STATUS_CHOICES)[novo_status]}",
                "cor_card": requisicao.cor_card,
            }
        )

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON inválido"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@require_POST
@login_required
def toggle_prioridade(request, pk):
    """
    Endpoint AJAX para alternar a flag de prioridade de uma requisição.
    Cards prioritários aparecem no topo da coluna.
    """
    try:
        requisicao = get_object_or_404(Requisicoes, id=pk)

        # Inverte a prioridade
        requisicao.prioridade = not requisicao.prioridade
        requisicao.save()

        return JsonResponse(
            {
                "success": True,
                "prioridade": requisicao.prioridade,
                "message": f'Prioridade {"ativada" if requisicao.prioridade else "desativada"}',
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
def kanban_detalhes_requisicao(request, pk):
    """
    Endpoint para retornar HTML dos detalhes da requisição para o modal.
    """
    try:
        requisicao = get_object_or_404(Requisicoes, id=pk)

        # Renderiza o template parcial com os detalhes
        from django.template.loader import render_to_string

        html = render_to_string(
            "kanban_detalhes_modal.html", {"requisicao": requisicao}, request=request
        )

        return JsonResponse(
            {
                "success": True,
                "html": html,
                "quantidade_ids_incluidos": requisicao.get_quantidade_ids_incluidos(),
                "numero_de_equipamentos": requisicao.numero_de_equipamentos,
                "requisicao": {
                    "id_equipamentos": requisicao.id_equipamentos or "",
                    "iccid": requisicao.iccid or "",
                },
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


# ============================================================================
# VIEWS DE AUDITORIA
# ============================================================================


@login_required
def ver_logs_requisicao(request, id):
    """
    View para exibir todos os logs de auditoria de uma requisição
    """
    requisicao = get_object_or_404(Requisicoes, id=id)

    # Buscar todos os logs desta requisição
    content_type = ContentType.objects.get_for_model(Requisicoes)
    logs = (
        AuditLog.objects.filter(content_type=content_type, object_id=requisicao.id)
        .select_related("usuario")
        .prefetch_related("campos_alterados")
        .order_by("-data_hora")
    )

    return render(
        request, "requisicao/audit_logs.html", {"requisicao": requisicao, "logs": logs}
    )


@login_required
def ver_logs_manutencao(request, id):
    """
    View para exibir todos os logs de auditoria de uma manutenção
    """
    manutencao = get_object_or_404(registrodemanutencao, id=id)

    # Buscar todos os logs desta manutenção
    content_type = ContentType.objects.get_for_model(registrodemanutencao)
    logs = (
        AuditLog.objects.filter(content_type=content_type, object_id=manutencao.id)
        .select_related("usuario")
        .prefetch_related("campos_alterados")
        .order_by("-data_hora")
    )

    return render(
        request,
        "requisicao/audit_logs.html",
        {"requisicao": manutencao, "logs": logs},  # Usa mesmo template
    )


# ============================================================================
# API DE REQUISIÇÕES
# ============================================================================

from django.views.decorators.cache import cache_page
from rest_framework.decorators import api_view
from rest_framework.response import Response

@api_view(["GET"])
@cache_page(60 * 30)  # Cache de 30 minutos
def api_requisicoes(request):
    """
    API para retornar todas as requisições do sistema
    Atualiza a cada 30 minutos (cache)

    Retorna:
    - N° Pedido
    - Cliente
    - Contrato (Descartavel / Retornavel)
    - Modelo
    - Status
    - Comercial
    - Customização
    - Quantidade
    """
    requisicoes = (
        Requisicoes.objects.select_related("nome", "tipo_produto").all().order_by("-id")
    )

    dados = []
    for req in requisicoes:
        dados.append(
            {
                "numero_pedido": req.id,
                "data": req.data.strftime("%d/%m/%Y %H:%M:%S") if req.data else "",
                "cliente": req.nome.nome if req.nome else "",
                "contrato": req.contrato if req.contrato else "",
                "modelo": req.tipo_produto.nome if req.tipo_produto else "",
                "status": req.status if req.status else "",
                "comercial": req.comercial if req.comercial else "",
                "customizacao": req.tipo_customizacao if req.tipo_customizacao else "",
                "quantidade": (
                    req.numero_de_equipamentos if req.numero_de_equipamentos else "0"
                ),
            }
        )

    return Response({"total": len(dados), "requisicoes": dados})


# ===================== EXPORT EXCEL =====================
from .excel_export import gerar_excel_requisicoes
from datetime import datetime

@login_required
def export_historico_excel(request):
    """
    View para exportar o histórico de requisições em Excel.
    Preserva os filtros aplicados.
    """
    # Aplicar os mesmos filtros da lista
    queryset = Requisicoes.objects.all().order_by("-id")
    
    nome = request.POST.get("nome", "")
    status = request.POST.get("status", "")
    id_filtro = request.POST.get("id_filtro", "")
    
    if nome:
        queryset = queryset.filter(nome__nome__icontains=nome)
    
    if status:
        queryset = queryset.filter(status__icontains=status)
    
    if id_filtro:
        try:
            id_valor = int(id_filtro)
            queryset = queryset.filter(id=id_valor)
        except ValueError:
            queryset = Requisicoes.objects.none()
    
    # Gerar Excel
    workbook = gerar_excel_requisicoes(queryset)
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Nome do arquivo com data/hora
    timestamp = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    filename = f"Requisicoes_{timestamp}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    # Salvar workbook na resposta
    workbook.save(response)
    
    return response




